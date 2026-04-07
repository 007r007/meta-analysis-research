"""
Paper-02 第一轮筛选Excel生成脚本
输入：02-search/数据_2_四库合并去重后.ris（1827条）
输出：03-screen/数据_1_第一轮筛选.xlsx

字段：序号、筛选决定、排除原因、备注、标题、作者、年份、期刊、来源数据库、摘要
"""

import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE_DIR   = Path(r"E:\Meta-analysis writing project\projects\paper-02")
RIS_FILE   = BASE_DIR / "02-search" / "数据_2_四库合并去重后.ris"
OUT_EXCEL  = BASE_DIR / "03-screen" / "数据_1_第一轮筛选.xlsx"

# ── 样式 ─────────────────────────────────────────────────────
COL_HEADER = "4472C4"
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 解析RIS ──────────────────────────────────────────────────

def parse_ris(filepath):
    records = []
    current = {}
    with open(filepath, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            if line.startswith("ER  -") or line.startswith("ER -"):
                if current:
                    records.append(current)
                    current = {}
            elif len(line) >= 6 and line[2:6] == "  - ":
                tag = line[:2].strip()
                val = line[6:].strip()
                if tag in current:
                    if isinstance(current[tag], list):
                        current[tag].append(val)
                    else:
                        current[tag] = [current[tag], val]
                else:
                    current[tag] = val
    if current:
        records.append(current)
    return records


def get_field(rec, *tags):
    for tag in tags:
        v = rec.get(tag, "")
        if isinstance(v, list):
            v = " ".join(str(x) for x in v)
        if v:
            return str(v).strip()
    return ""


def get_year(rec):
    y = get_field(rec, "PY", "Y1", "DA")
    m = re.search(r"\d{4}", y)
    return int(m.group()) if m else ""


def get_authors(rec):
    au = rec.get("AU") or rec.get("A1") or ""
    if isinstance(au, list):
        authors = "; ".join(str(a) for a in au[:3])
        if len(au) > 3:
            authors += " et al."
        return authors
    return str(au)


# ── 加载RIS ──────────────────────────────────────────────────

print(f"读取RIS: {RIS_FILE}")
records = parse_ris(RIS_FILE)
print(f"共 {len(records)} 条文献")

# ── 生成Excel ─────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "第一轮筛选"

headers = [
    "#",            # A
    "筛选决定",     # B  ← 下拉
    "排除原因",     # C  ← 下拉
    "备注",         # D
    "标题",         # E
    "作者",         # F
    "年份",         # G
    "期刊",         # H
    "来源数据库",   # I
    "摘要",         # J
]

# 表头样式
header_fill = PatternFill("solid", fgColor=COL_HEADER)
header_font = Font(bold=True, color="FFFFFF", size=11)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 32

# 数据验证：筛选决定
dv_decision = DataValidation(
    type="list",
    formula1='"纳入,排除,不确定"',
    allow_blank=True,
    showDropDown=False,
)
dv_decision.sqref = f"B2:B{len(records)+1}"
ws.add_data_validation(dv_decision)

# 数据验证：排除原因（适配paper-02 PICOS）
exclude_reasons = (
    '"E1-年龄不符(非0-8岁),'
    'E2-暴露不符(非父母教育/SES),'
    'E3-无神经指标,'
    'E4-综述/元分析/protocol,'
    'E5-非英文,'
    'E6-无法获取全文,'
    'E7-样本重复报告,'
    'E8-其他"'
)
dv_reason = DataValidation(
    type="list",
    formula1=exclude_reasons,
    allow_blank=True,
    showDropDown=False,
)
dv_reason.sqref = f"C2:C{len(records)+1}"
ws.add_data_validation(dv_reason)

# 数据行
for i, rec in enumerate(records, 1):
    title   = get_field(rec, "TI", "T1")
    authors = get_authors(rec)
    year    = get_year(rec)
    journal = get_field(rec, "JO", "JF", "T2")
    db      = get_field(rec, "DB")
    abstract = get_field(rec, "AB", "N2")

    row_data = [i, "", "", "", title, authors, year, journal, db, abstract]
    row_num  = i + 1

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=(col in (5, 10)))
        if col in (1, 2, 3, 7, 9):
            cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.row_dimensions[row_num].height = 80

# 列宽
col_widths = [6, 10, 20, 20, 65, 28, 7, 28, 12, 90]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# 冻结首行
ws.freeze_panes = "E2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── 说明sheet ────────────────────────────────────────────────

ws2 = wb.create_sheet("筛选说明")
instructions = [
    ("操作说明", ""),
    ("1. 逐行阅读标题+摘要", ""),
    ("2. 在B列「筛选决定」选择：纳入 / 排除 / 不确定", ""),
    ("3. 排除时在C列选择排除原因代码", ""),
    ("4. 不确定的文献进入第二轮全文筛选", ""),
    ("", ""),
    ("排除代码说明（Paper-02 PICOS）", ""),
    ("E1", "年龄不符：样本非0–8岁（或无法确认包含0–8岁儿童）"),
    ("E2", "暴露不符：非父母教育水平/SES（如暴露=酒精/铅/屏幕时间等）"),
    ("E3", "无神经指标：无EEG/ERP/fNIRS/fMRI/DTI/MEG等直接神经测量"),
    ("E4", "综述/元分析/protocol/评论/书评"),
    ("E5", "非英文文献"),
    ("E6", "无法获取全文"),
    ("E7", "样本重复报告（同一数据集多篇，保留主要结果）"),
    ("E8", "其他不符合标准（请在备注列说明）"),
    ("", ""),
    ("PICOS标准", ""),
    ("P", "0–8岁儿童（婴儿、幼儿、学龄前、幼儿园、小学低年级）"),
    ("I/E", "父母教育水平（SES单一维度，不含复合SES指数）"),
    ("O", "任一神经指标：EEG/ERP/fNIRS/fMRI/DTI/MEG等"),
    ("S", "观察性研究（横断/纵向）或RCT；排除综述/元分析"),
    ("", ""),
    ("目标", f"1827条 → 预计第一轮保留约300–500条"),
]

for r, (key, val) in enumerate(instructions, 1):
    c1 = ws2.cell(row=r, column=1, value=key)
    c2 = ws2.cell(row=r, column=2, value=val)
    bold_keys = ("操作说明", "排除代码说明（Paper-02 PICOS）", "PICOS标准", "目标")
    if key in bold_keys:
        c1.font = Font(bold=True)
    c1.alignment = Alignment(vertical="top")
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.row_dimensions[r].height = 18

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 60

# ── 保存 ─────────────────────────────────────────────────────

OUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_EXCEL)
print(f"\nExcel已保存：{OUT_EXCEL}")
print(f"共 {len(records)} 条待筛选")
print("完成。")
