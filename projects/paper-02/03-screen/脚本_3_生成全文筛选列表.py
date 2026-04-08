"""
脚本_3_生成全文筛选列表.py
从数据_2_第一轮自动筛选结果.xlsx中提取B列为"保留2"或"不确定"的行，
生成全文筛选待下载列表。
输出：03-screen/数据_3_全文筛选待下载列表.xlsx
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE     = Path(r"E:\Meta-analysis writing project\projects\paper-02")
SRC      = BASE / "03-screen" / "数据_2_第一轮自动筛选结果.xlsx"
OUT      = BASE / "03-screen" / "数据_3_全文筛选待下载列表.xlsx"

# ── 读取源文件 ────────────────────────────────────────────────
print("读取源文件...")
wb_src = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws_src = wb_src.active

rows_keep = []
for row in ws_src.iter_rows(min_row=2, values_only=True):
    b = row[1] if len(row) > 1 else None
    if b in ('保留2', '不确定'):
        rows_keep.append(row)

wb_src.close()
print(f"提取到 {len(rows_keep)} 条（保留2+不确定）")

# ── 创建输出Excel ─────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '全文筛选待下载'

headers = ['序号', '标题', '作者', '年份', '期刊', '来源数据库', '筛选状态', 'DOI/链接', 'PDF状态']

thin        = Side(style='thin', color='AAAAAA')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', size=11)

# 颜色：保留2=绿，不确定=黄
fill_keep = PatternFill('solid', fgColor='C6EFCE')
fill_unc  = PatternFill('solid', fgColor='FFEB9C')

# 表头
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 28

# 数据行
# 源列映射：row[0]=#, row[1]=筛选结果, row[2]=排除原因, row[3]=标题,
#           row[4]=作者, row[5]=年份, row[6]=期刊, row[7]=来源数据库
for i, row in enumerate(rows_keep, 1):
    orig_no  = row[0]   # 原序号
    status   = row[1]   # 保留2 / 不确定
    title    = row[3] or ''
    authors  = row[4] or ''
    year     = row[5] or ''
    journal  = row[6] or ''
    source   = row[7] or ''

    row_data = [i, title, authors, year, journal, source, status, '', '待下载']
    fill     = fill_keep if status == '保留2' else fill_unc
    row_num  = i + 1

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = fill
        cell.border = border
        if col in (1, 4, 6, 7, 9):
            cell.alignment = Alignment(horizontal='center', vertical='top')
        else:
            cell.alignment = Alignment(vertical='top', wrap_text=(col in (2, 3)))

    ws.row_dimensions[row_num].height = 55

# 列宽
col_widths = [6, 70, 35, 8, 35, 14, 10, 35, 12]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

# 图例sheet
ws2 = wb.create_sheet('说明')
legend = [
    ('颜色', '状态', '说明'),
    ('绿色', '保留2', 'xhs+oo核查后确认保留，需下载全文'),
    ('黄色', '不确定', '边界案例，需全文判断'),
]
lc = ['4472C4', 'C6EFCE', 'FFEB9C']
for r, (rd, color) in enumerate(zip(legend, lc), 1):
    for c, val in enumerate(rd, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(bold=(r == 1), color='FFFFFF' if r == 1 else '000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws2.row_dimensions[r].height = 22
for col, width in enumerate([10, 12, 45], 1):
    ws2.column_dimensions[get_column_letter(col)].width = width

wb.save(OUT)
print(f"已保存：{OUT}")

# 统计
n_keep = sum(1 for r in rows_keep if r[1] == '保留2')
n_unc  = sum(1 for r in rows_keep if r[1] == '不确定')
print(f"  保留2：{n_keep} 条")
print(f"  不确定：{n_unc} 条")
print(f"  合计：{len(rows_keep)} 条")
print("完成。")
