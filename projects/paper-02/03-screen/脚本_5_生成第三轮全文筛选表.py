"""
脚本_5_生成第三轮全文筛选表.py
从数据_4_全文下载追踪.xlsx中读取"已下载"条目（108篇），
生成供xhs全文筛选用的数据_5_第三轮全文筛选.xlsx。

Sheet1: 第三轮全文筛选（主表，108行）
Sheet2: 进度统计（公式自动计算）
Sheet3: PICOS速查（paper-02专用）
"""

import openpyxl, sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

SRC = 'E:/Meta-analysis writing project/projects/paper-02/03-screen/数据_4_全文下载追踪.xlsx'
OUT = 'E:/Meta-analysis writing project/projects/paper-02/03-screen/数据_5_第三轮全文筛选.xlsx'

# ── 读取数据_4，只取已下载条目 ──────────────────────────────
src_wb = openpyxl.load_workbook(SRC)
ws_src = src_wb.active

# 列映射: 序号(0) 筛选状态(1) 标题(2) 作者(3) 年份(4) 期刊(5) DOI(6) 文件名(7) 下载状态(8)
rows = []
for r in range(2, ws_src.max_row + 1):
    vals = [ws_src.cell(r, c).value for c in range(1, 14)]
    dl_status = str(vals[8] or '')
    if '已下载' in dl_status:
        rows.append(vals)

print(f'读取已下载条目: {len(rows)} 篇')

# ── 样式工具 ────────────────────────────────────────────────
thin   = Side(style='thin', color='B8CCE4')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def style(cell, bg=None, bold=False, align='left', fc='000000', wrap=False, size=10):
    if bg:
        cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fc, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = border

def first_author(s):
    if not s:
        return ''
    p = str(s).split(';')[0].split(',')[0].strip().split()
    return p[-1] if p else str(s)[:15]

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════
# Sheet 1: 第三轮全文筛选（主表）
# ══════════════════════════════════════════════════════════
ws = wb.active
ws.title = '第三轮全文筛选'

cols = [
    ('#',           5),
    ('原序号',       7),
    ('年份',         6),
    ('第一作者',    14),
    ('标题',        55),
    ('期刊',        26),
    ('DOI',         32),
    ('PDF文件名',    40),
    ('全文筛选结果', 13),   # 下拉：纳入/排除/不确定
    ('排除代码',     12),   # 下拉：E1~E8
    ('排除原因说明', 30),   # 文字说明
    ('备注',        30),   # 同一样本/年龄实际范围等
]

# 表头
for ci, (h, w) in enumerate(cols, 1):
    c = ws.cell(1, ci, h)
    style(c, bg='2F5496', bold=True, align='center', fc='FFFFFF', size=10)
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 24
ws.freeze_panes = 'A2'

last_row = len(rows) + 1

for ri, row in enumerate(rows, 2):
    orig_no  = row[0]    # 原序号
    title    = row[2]
    author   = row[3]
    year     = row[4]
    journal  = row[5]
    doi      = row[6]
    filename = row[7]
    fa = first_author(str(author) if author else '')

    vals = [ri-1, orig_no, year, fa, title, journal, doi, filename,
            None, None, None, None]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v)
        align = 'center' if ci in (1, 2, 3, 9, 10) else 'left'
        style(c, align=align, wrap=(ci in (5, 11, 12)))
    ws.row_dimensions[ri].height = 32

# 下拉验证
dv1 = DataValidation(type='list', formula1='"纳入,排除,不确定"',
                     allow_blank=True, showDropDown=False)
dv1.sqref = f'I2:I{last_row}'
ws.add_data_validation(dv1)

dv2 = DataValidation(type='list',
    formula1='"E1-年龄不符,E2-暴露不符,E3-无神经指标,E4-综述/protocol,E5-非英文,E6-全文不可及,E7-样本重复,E8-其他"',
    allow_blank=True, showDropDown=False)
dv2.sqref = f'J2:J{last_row}'
ws.add_data_validation(dv2)

# 条件格式：整行变色
ws.conditional_formatting.add(f'A2:L{last_row}', FormulaRule(
    formula=['$I2="纳入"'],  fill=fill('C6EFCE'), font=Font(color='006100')))
ws.conditional_formatting.add(f'A2:L{last_row}', FormulaRule(
    formula=['$I2="排除"'],  fill=fill('FFC7CE'), font=Font(color='9C0006')))
ws.conditional_formatting.add(f'A2:L{last_row}', FormulaRule(
    formula=['$I2="不确定"'], fill=fill('FFEB9C'), font=Font(color='7F6000')))

ws.auto_filter.ref = f'A1:L{last_row}'

# ══════════════════════════════════════════════════════════
# Sheet 2: 进度统计
# ══════════════════════════════════════════════════════════
ws2 = wb.create_sheet('进度统计')
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 30

c = ws2.cell(1, 1, 'Paper-02 第三轮全文筛选 — 进度统计')
c.font = Font(bold=True, size=13, color='2F5496')
ws2.merge_cells('A1:C1')
ws2.row_dimensions[1].height = 28

sr = '第三轮全文筛选'
stats = [
    ('指标',          '数量',                                              '说明',                   True),
    ('输入总数',       f'=COUNTA({sr}!A2:A{last_row})',                    '进入第三轮的条目',        False),
    ('已完成判决',     f'=COUNTA({sr}!I2:I{last_row})',                    '已填写筛选结果',          False),
    ('待判决',         '=B4-B5',                                           '尚未填写',                False),
    ('完成进度',       '=IFERROR(B5/B4,0)',                                '',                       False),
    ('',              '',                                                  '',                       False),
    ('纳入',           f'=COUNTIF({sr}!I$2:I${last_row},"纳入")',          '最终纳入数',              False),
    ('排除',           f'=COUNTIF({sr}!I$2:I${last_row},"排除")',          '',                       False),
    ('不确定',         f'=COUNTIF({sr}!I$2:I${last_row},"不确定")',        '待进一步确认',            False),
    ('',              '',                                                  '',                       False),
    ('排除代码明细',   '',                                                  '',                       True),
    ('E1 年龄不符',    f'=COUNTIF({sr}!J$2:J${last_row},"E1-年龄不符")',   '样本非0-8岁/纯临床',     False),
    ('E2 暴露不符',    f'=COUNTIF({sr}!J$2:J${last_row},"E2-暴露不符")',   '非父母教育/教育效应不可提取', False),
    ('E3 无神经指标',  f'=COUNTIF({sr}!J$2:J${last_row},"E3-无神经指标")', '仅行为/认知/问卷结局',   False),
    ('E4 综述/protocol',f'=COUNTIF({sr}!J$2:J${last_row},"E4-综述/protocol")', '综述/元分析/protocol', False),
    ('E5 非英文',      f'=COUNTIF({sr}!J$2:J${last_row},"E5-非英文")',     '全文非英文',             False),
    ('E6 全文不可及',  f'=COUNTIF({sr}!J$2:J${last_row},"E6-全文不可及")', '先补下载再标E6',         False),
    ('E7 样本重复',    f'=COUNTIF({sr}!J$2:J${last_row},"E7-样本重复")',   '同样本保留主要文章',     False),
    ('E8 其他',        f'=COUNTIF({sr}!J$2:J${last_row},"E8-其他")',       '备注列说明',             False),
]

for ri, (a, b, c_, hdr) in enumerate(stats, 2):
    ca = ws2.cell(ri, 1, a)
    cb = ws2.cell(ri, 2, b)
    cc = ws2.cell(ri, 3, c_)
    bg = 'D9E1F2' if hdr else None
    for cell in (ca, cb, cc):
        cell.font = Font(bold=hdr, size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        if bg:
            cell.fill = fill(bg)
    if b == '=IFERROR(B5/B4,0)':
        cb.number_format = '0.0%'
        cb.fill = fill('E2EFDA')
        cb.font = Font(bold=True, size=10, color='375623')
    ws2.row_dimensions[ri].height = 18

ws2.freeze_panes = 'A2'

# ══════════════════════════════════════════════════════════
# Sheet 3: PICOS速查（paper-02专用）
# ══════════════════════════════════════════════════════════
ws3 = wb.create_sheet('PICOS速查')
ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 58
ws3.column_dimensions['C'].width = 38

c = ws3.cell(1, 1, 'PICOS 第三轮全文筛选速查 — Paper-02（v1.0，2026-04-09）')
c.font = Font(bold=True, size=12, color='2F5496')
ws3.merge_cells('A1:C1')
ws3.row_dimensions[1].height = 24

picos_data = [
    ('维度',     '纳入标准',  '排除 → 代码',  True,  18),
    ('P 研究对象',
     '儿童，年龄范围覆盖0–8岁（含婴儿期）\n若研究含多年龄组，有0–8岁子样本且单独报告结果即纳入',
     '样本年龄全部>8岁 → E1\n纯临床样本（早产/神经发育障碍/脑损伤/癫痫）除非有健康对照 → E1\n动物研究 → E1',
     False, 58),
    ('I/E 暴露',
     '父母教育水平作为主要暴露变量\n或在回归模型中有独立Beta/b系数及统计检验\n复合SES（教育+收入+职业），但教育的独立效应可单独提取',
     'SES仅为协变量，无教育独立效应量（无beta/b值）→ E2\nPCA合成SES主成分，教育权重隐含 → E2\n暴露为其他环境因素（缺铁/养育/虐待/干预等）→ E2',
     False, 72),
    ('O 结局',
     '任一神经测量指标：\nrsEEG（功率/连接）、ERP（MMN/P300/N400/Nc等）\nfNIRS、fMRI（任务态/静息态功能连接）\nDTI（FA/MD等）、structural MRI（皮层厚度/脑区体积/表面积/VBM）\nMEG/MRS（若有）',
     '仅报告行为/认知结局（无神经测量）→ E3\n仅主观问卷 → E3\n仅眼动/生理指标（心率/皮电）→ E3',
     False, 72),
    ('S 研究设计',
     '横断面观察性研究\n纵向观察性研究（队列/随访）\nRCT（若含SES×神经的子分析）',
     '综述/系统综述/元分析/叙述性综述 → E4\n研究方案（protocol）/评论/社论/信件 → E4\n非英文全文 → E5\n无法获取全文 → E6\n与已纳入研究完全相同样本重复报告 → E7',
     False, 68),
    ('', '', '', False, 10),
    ('SES编码说明', '关键判断：教育效应能否单独提取',  '', True, 18),
    ('edu-only',
     '单独报告父母教育效应 → 纳入',
     '备注ses_type=edu-only',
     False, 22),
    ('composite-separable',
     '复合SES，但回归中教育有独立Beta/b系数 → 纳入',
     '备注ses_type=composite-separable',
     False, 22),
    ('composite-inseparable',
     'PCA主成分/教育权重隐含/教育仅为协变量 → 排除E2',
     '备注ses_type=composite-inseparable',
     False, 22),
    ('', '', '', False, 10),
    ('排除代码', '含义',  '操作提示', True, 18),
    ('E1', '年龄不符/纯临床', '年龄范围、是否有健康对照组',              False, 18),
    ('E2', '暴露不符',        '父母教育是否有独立beta，SES编码方案',      False, 18),
    ('E3', '无神经指标',      'EEG/ERP/fNIRS/fMRI/DTI/MRI任一即可',      False, 18),
    ('E4', '综述/protocol',   '标题/摘要/全文确认是否实证研究',           False, 18),
    ('E5', '非英文',          '全文语言',                                 False, 18),
    ('E6', '全文不可及',      '已有PDF的不适用此代码',                    False, 18),
    ('E7', '样本重复',        '同一样本保留报告最完整的那篇',             False, 18),
    ('E8', '其他',            '备注列说明具体原因',                       False, 18),
    ('', '', '', False, 10),
    ('五步快速判断', '① 样本有0–8岁儿童且非纯临床？      否 → E1', '', False, 18),
    ('',             '② 父母教育作为主暴露或有独立效应？  否 → E2', '', False, 18),
    ('',             '③ 报告神经测量指标（EEG/MRI等）？  否 → E3', '', False, 18),
    ('',             '④ 实证研究（非综述/protocol）？    否 → E4', '', False, 18),
    ('',             '⑤ 英文全文可获取？                 否 → E5/E6', '', False, 18),
    ('',             '⑥ 全通过 → 纳入',                                '', False, 18),
]

for ri, (a, b, c_, hdr, h) in enumerate(picos_data, 2):
    ca = ws3.cell(ri, 1, a)
    cb = ws3.cell(ri, 2, b)
    cc = ws3.cell(ri, 3, c_)
    bg = 'D9E1F2' if hdr else None
    for cell in (ca, cb, cc):
        cell.font = Font(bold=hdr, size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        if bg:
            cell.fill = fill(bg)
    ws3.row_dimensions[ri].height = h

ws3.freeze_panes = 'A2'

wb.save(OUT)
print(f'生成成功：{OUT}')
print(f'数据行数：{len(rows)} 条（已下载PDF的条目）')
