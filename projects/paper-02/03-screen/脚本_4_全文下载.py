"""
脚本_4_全文下载.py
Paper-02 全文PDF自动下载
输入：03-screen/数据_3_全文筛选待下载列表.xlsx（133条，保留2+不确定）
RIS：02-search/数据_2_四库合并去重后.ris（用于匹配DOI）
输出：03-screen/全文PDF/  （PDF文件）
      03-screen/数据_4_全文下载追踪.xlsx （下载状态追踪）

下载策略（优先级）：
1. PubMed Central (PMC) — 开放获取，无验证码，国内可直连
2. CORE API — 开放获取聚合
3. Sci-Hub镜像 — 备选（当前有验证码，成功率低）
"""

import io
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ── 路径配置 ──────────────────────────────────────────────────
BASE      = Path(r"E:\Meta-analysis writing project\projects\paper-02")
EXCEL_IN  = BASE / "03-screen" / "数据_3_全文筛选待下载列表.xlsx"
RIS_FILE  = BASE / "02-search" / "数据_2_四库合并去重后.ris"
PDF_DIR   = BASE / "03-screen" / "全文PDF"
EXCEL_OUT = BASE / "03-screen" / "数据_4_全文下载追踪.xlsx"

PDF_DIR.mkdir(exist_ok=True)

SCIHUB_MIRRORS = [
    "https://sci-hub.st",
    "https://sci-hub.ru",
    "https://sci-hub.box",
    "https://sci-hub.red",
    "https://sci-hub.su",
]

# Clash代理配置（HTTP代理端口17890）
PROXIES = {
    "http":  "http://127.0.0.1:17890",
    "https": "http://127.0.0.1:17890",
}

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36"
}

# ── 工具函数 ──────────────────────────────────────────────────

def normalize_title(t: str) -> str:
    if not t: return ""
    t = unicodedata.normalize("NFKD", t)
    t = re.sub(r"[^\w\s]", " ", t.lower())
    return re.sub(r"\s+", " ", t).strip()


def load_ris_doi_map(ris_path: Path) -> dict:
    doi_map = {}
    cur_title = cur_doi = None
    with open(ris_path, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            if re.match(r'^TI\s+-\s+', line):
                cur_title = re.sub(r'^TI\s+-\s+', '', line).strip()
            elif re.match(r'^T1\s+-\s+', line) and not cur_title:
                cur_title = re.sub(r'^T1\s+-\s+', '', line).strip()
            elif re.match(r'^DO\s+-\s+', line):
                raw = re.sub(r'^DO\s+-\s+', '', line).strip()
                raw = re.sub(r'^https?://(dx\.)?doi\.org/', '', raw).strip()
                raw = re.sub(r'\s*\[doi\]$', '', raw, flags=re.I).strip()
                cur_doi = raw
            elif line.startswith("ER  -") or line.startswith("ER -"):
                if cur_title and cur_doi and re.match(r'^10\.\d{4,}/', cur_doi):
                    doi_map[normalize_title(cur_title)] = cur_doi
                cur_title = cur_doi = None
    print(f"[RIS] 解析完成，共 {len(doi_map)} 条有DOI记录")
    return doi_map


def find_doi(title: str, doi_map: dict) -> str | None:
    key = normalize_title(title)
    if key in doi_map: return doi_map[key]
    prefix = key[:100]
    for k, v in doi_map.items():
        if k[:100] == prefix: return v
    return None


def safe_filename(title: str, year, max_len=80) -> str:
    name = re.sub(r'[\\/:*?"<>|]', '_', str(title))
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > max_len:
        name = name[:max_len].strip()
    return f"{year}_{name}.pdf"


# ── 方法1：PMC下载（优先，无验证码） ─────────────────────────

def try_pmc(doi: str, save_path: Path) -> tuple[bool, str]:
    """DOI -> PMID -> PMCID -> PDF，通过Clash代理访问NCBI"""
    try:
        # Step1: DOI -> PMID
        r1 = requests.get(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi",
            params={"db": "pubmed", "term": f"{doi}[doi]", "retmode": "json"},
            headers=HEADERS, proxies=PROXIES, timeout=15
        )
        ids = r1.json().get("esearchresult", {}).get("idlist", [])
        if not ids:
            return False, ""
        pmid = ids[0]
        time.sleep(0.4)  # NCBI限速：10次/秒

        # Step2: PMID -> PMCID
        r2 = requests.get(
            "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/elink.fcgi",
            params={"dbfrom": "pubmed", "db": "pmc", "id": pmid, "retmode": "json"},
            headers=HEADERS, proxies=PROXIES, timeout=15
        )
        data = r2.json()
        pmcids = []
        for ls in data.get("linksets", []):
            for lsd in ls.get("linksetdbs", []):
                if lsd.get("dbto") == "pmc":
                    pmcids = lsd.get("links", [])
        if not pmcids:
            return False, ""
        pmcid = pmcids[0]
        time.sleep(0.4)

        # Step3: 下载PDF
        pdf_url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/PMC{pmcid}/pdf/"
        r3 = requests.get(pdf_url, headers=HEADERS, proxies=PROXIES, timeout=60,
                          stream=True, allow_redirects=True)
        if r3.status_code == 200:
            ct = r3.headers.get("Content-Type", "")
            if "pdf" in ct or "octet-stream" in ct:
                with open(save_path, "wb") as f:
                    for chunk in r3.iter_content(8192):
                        f.write(chunk)
                if save_path.stat().st_size > 10_000:
                    return True, f"PMC{pmcid}"
            # PMC返回HTML时，从页面找真实PDF链接
            if "html" in ct:
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(r3.text, "html.parser")
                for a in soup.find_all("a", href=True):
                    if ".pdf" in a["href"].lower():
                        real_pdf = a["href"]
                        if real_pdf.startswith("/"):
                            real_pdf = "https://www.ncbi.nlm.nih.gov" + real_pdf
                        r4 = requests.get(real_pdf, headers=HEADERS, proxies=PROXIES,
                                          timeout=60, stream=True)
                        if r4.status_code == 200 and "pdf" in r4.headers.get("Content-Type",""):
                            with open(save_path, "wb") as f:
                                for chunk in r4.iter_content(8192):
                                    f.write(chunk)
                            if save_path.stat().st_size > 10_000:
                                return True, f"PMC{pmcid}"
    except Exception as e:
        print(f"    [PMC] 异常: {e}")
    return False, ""


# ── 方法2：CORE API ───────────────────────────────────────────

def try_core(doi: str, save_path: Path) -> bool:
    try:
        r = requests.get(f"https://api.core.ac.uk/v3/works/doi:{doi}",
                         headers=HEADERS, proxies=PROXIES, timeout=20)
        if r.status_code != 200: return False
        data = r.json()
        dl_url = data.get("downloadUrl") or data.get("fullTextLink")
        if not dl_url:
            for lk in data.get("links", []):
                if lk.get("type") in ("download", "pdf") and lk.get("url"):
                    dl_url = lk["url"]
                    break
        if not dl_url: return False
        pdf_r = requests.get(dl_url, headers=HEADERS, proxies=PROXIES, timeout=40, stream=True)
        if pdf_r.status_code == 200 and "pdf" in pdf_r.headers.get("Content-Type", ""):
            with open(save_path, "wb") as f:
                for chunk in pdf_r.iter_content(8192):
                    f.write(chunk)
            return save_path.stat().st_size > 10_000
    except Exception as e:
        print(f"    [CORE] 异常: {e}")
    return False


# ── 方法3：Sci-Hub ────────────────────────────────────────────

def try_scihub(doi: str, save_path: Path) -> tuple[bool, str]:
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return False, ""
    for mirror in SCIHUB_MIRRORS:
        try:
            r = requests.get(f"{mirror}/{doi}", headers=HEADERS, proxies=PROXIES, timeout=30)
            if r.status_code != 200: continue
            soup = BeautifulSoup(r.text, "html.parser")
            # 检测验证码
            if "captcha" in r.text.lower() or "i'm not a robot" in r.text.lower():
                print(f"    [Sci-Hub {mirror}] 触发验证码，跳过")
                continue
            pdf_url = None
            iframe = soup.find("iframe", id="pdf")
            if iframe and iframe.get("src"): pdf_url = iframe["src"]
            if not pdf_url:
                embed = soup.find("embed", type="application/pdf")
                if embed and embed.get("src"): pdf_url = embed["src"]
            if not pdf_url:
                for a in soup.find_all("a", href=True):
                    if ".pdf" in a["href"].lower():
                        pdf_url = a["href"]
                        break
            if not pdf_url: continue
            if pdf_url.startswith("//"): pdf_url = "https:" + pdf_url
            elif pdf_url.startswith("/"): pdf_url = mirror + pdf_url
            pdf_r = requests.get(pdf_url, headers=HEADERS, proxies=PROXIES, timeout=60, stream=True)
            if pdf_r.status_code == 200:
                ct = pdf_r.headers.get("Content-Type", "")
                if "pdf" in ct or "octet-stream" in ct:
                    with open(save_path, "wb") as f:
                        for chunk in pdf_r.iter_content(8192):
                            f.write(chunk)
                    if save_path.stat().st_size > 10_000:
                        return True, mirror
        except Exception as e:
            print(f"    [Sci-Hub {mirror}] 异常: {e}")
    return False, ""


# ── 主流程 ────────────────────────────────────────────────────

def main():
    wb_in = openpyxl.load_workbook(EXCEL_IN, data_only=True)
    ws_in = wb_in.active
    papers = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if not row[1]: continue
        raw_doi = str(row[7] or "").strip()
        # 清理DOI字段
        raw_doi = re.sub(r'^https?://(dx\.)?doi\.org/', '', raw_doi).strip()
        if not re.match(r'^10\.\d{4,}/', raw_doi):
            raw_doi = ""
        papers.append({
            "seq":     row[0],
            "title":   str(row[1]).strip(),
            "authors": str(row[2] or ""),
            "year":    row[3] or "unknown",
            "journal": str(row[4] or ""),
            "source":  str(row[5] or ""),
            "status":  str(row[6] or ""),
            "doi_col": raw_doi,
        })
    wb_in.close()
    print(f"\n待下载：{len(papers)} 篇\n")

    doi_map = load_ris_doi_map(RIS_FILE)

    results = []
    for i, p in enumerate(papers, 1):
        title = p["title"]
        year  = p["year"]
        fname = safe_filename(title, year)
        fpath = PDF_DIR / fname

        doi = p["doi_col"] if p["doi_col"] else find_doi(title, doi_map)

        print(f"[{i:03d}/{len(papers)}] {title[:65]}...")
        print(f"    DOI: {doi or '未找到'}")

        if fpath.exists() and fpath.stat().st_size > 10_000:
            print("    -> 已存在，跳过")
            results.append({**p, "doi": doi or "", "filename": fname,
                             "dl_status": "已存在", "dl_source": "缓存"})
            continue

        if not doi:
            print("    -> 无DOI，需手动下载")
            results.append({**p, "doi": "", "filename": fname,
                             "dl_status": "无DOI-手动", "dl_source": ""})
            continue

        # 方法1：PMC（优先）
        ok, pmcid = try_pmc(doi, fpath)
        if ok:
            print(f"    -> [OK] PMC下载成功 ({pmcid})")
            results.append({**p, "doi": doi, "filename": fname,
                             "dl_status": "已下载", "dl_source": f"PMC({pmcid})"})
            time.sleep(0.5)
            continue

        # 方法2：CORE
        if try_core(doi, fpath):
            print("    -> [OK] CORE下载成功")
            results.append({**p, "doi": doi, "filename": fname,
                             "dl_status": "已下载", "dl_source": "CORE"})
            time.sleep(1)
            continue

        # 方法3：Sci-Hub
        ok, mirror = try_scihub(doi, fpath)
        if ok:
            print(f"    -> [OK] Sci-Hub下载成功 ({mirror})")
            results.append({**p, "doi": doi, "filename": fname,
                             "dl_status": "已下载", "dl_source": f"Sci-Hub({mirror})"})
        else:
            print("    -> [FAIL] 自动下载失败，需手动")
            results.append({**p, "doi": doi, "filename": fname,
                             "dl_status": "手动下载", "dl_source": ""})

        time.sleep(1)

    # 生成追踪Excel
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "全文下载追踪"
    headers = ["序号", "筛选状态", "标题", "作者", "年份", "期刊",
               "DOI", "文件名", "下载状态", "来源", "全文筛选结果", "排除原因", "备注"]
    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    hfill  = PatternFill("solid", fgColor="2E75B6")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    status_fill = {"已下载": green, "已存在": green,
                   "手动下载": yellow, "无DOI-手动": yellow}
    for r in results:
        row_data = [r["seq"], r["status"], r["title"], r["authors"],
                    r["year"], r["journal"], r["doi"], r["filename"],
                    r["dl_status"], r["dl_source"], "", "", ""]
        ws.append(row_data)
        rn = ws.max_row
        fill = status_fill.get(r["dl_status"], PatternFill("solid", fgColor="FFC7CE"))
        for col in range(1, len(headers) + 1):
            cell = ws.cell(rn, col)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(col in (3, 4)))
        ws.row_dimensions[rn].height = 50
    col_widths = [6, 10, 60, 30, 6, 28, 38, 50, 14, 22, 14, 16, 20]
    for col_i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    ws2 = wb_out.create_sheet("统计说明")
    n_ok  = sum(1 for r in results if r["dl_status"] in ("已下载", "已存在"))
    n_man = len(results) - n_ok
    pmc_cnt = sum(1 for r in results if "PMC" in r.get("dl_source",""))
    core_cnt = sum(1 for r in results if "CORE" in r.get("dl_source",""))
    sh_cnt = sum(1 for r in results if "Sci-Hub" in r.get("dl_source",""))
    for row in [
        ("自动下载成功", n_ok), ("  其中PMC", pmc_cnt),
        ("  其中CORE", core_cnt), ("  其中Sci-Hub", sh_cnt),
        ("需手动下载", n_man), ("总计", len(results)),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]:
        ws2.append(list(row))
    wb_out.save(EXCEL_OUT)

    print(f"\n{'='*60}")
    print(f"完成！自动下载：{n_ok}/{len(results)}")
    print(f"  PMC: {pmc_cnt} | CORE: {core_cnt} | Sci-Hub: {sh_cnt}")
    print(f"需手动下载：{n_man} 篇")
    print(f"PDF目录：{PDF_DIR}")
    print(f"追踪Excel：{EXCEL_OUT}")


if __name__ == "__main__":
    main()


import io
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Fix Windows GBK console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# ── 路径配置 ──────────────────────────────────────────────────
BASE     = Path(r"E:\Meta-analysis writing project\projects\paper-02")
EXCEL_IN = BASE / "03-screen" / "数据_3_全文筛选待下载列表.xlsx"
RIS_FILE = BASE / "02-search" / "数据_2_四库合并去重后.ris"
PDF_DIR  = BASE / "03-screen" / "全文PDF"
EXCEL_OUT = BASE / "03-screen" / "数据_4_全文下载追踪.xlsx"

PDF_DIR.mkdir(exist_ok=True)

# ── Sci-Hub镜像（按优先级） ───────────────────────────────────
SCIHUB_MIRRORS = [
    "https://sci-hub.st",   # ✅ 首选
    "https://sci-hub.ru",   # ✅
    "https://sci-hub.box",  # ✅
    "https://sci-hub.red",  # ✅ 备选
    "https://sci-hub.su",   # ✅ 备选
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36"
}

# ── 工具函数 ──────────────────────────────────────────────────

def normalize_title(t: str) -> str:
    """小写+去标点，用于模糊匹配"""
    if not t:
        return ""
    t = unicodedata.normalize("NFKD", t)
    t = re.sub(r"[^\w\s]", " ", t.lower())
    return re.sub(r"\s+", " ", t).strip()


def load_ris_doi_map(ris_path: Path) -> dict:
    """解析RIS，返回 title_normalized -> doi 字典"""
    doi_map = {}
    current_title = None
    current_doi = None

    with open(ris_path, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            # TI字段（标准RIS）
            if re.match(r'^TI\s+-\s+', line):
                current_title = re.sub(r'^TI\s+-\s+', '', line).strip()
            # T1字段（部分来源）
            elif re.match(r'^T1\s+-\s+', line) and not current_title:
                current_title = re.sub(r'^T1\s+-\s+', '', line).strip()
            # DO字段
            elif re.match(r'^DO\s+-\s+', line):
                raw = re.sub(r'^DO\s+-\s+', '', line).strip()
                current_doi = re.sub(r'\s*\[doi\]$', '', raw, flags=re.I).strip()
            elif line.startswith("ER  -") or line.startswith("ER -"):
                if current_title and current_doi:
                    key = normalize_title(current_title)
                    doi_map[key] = current_doi
                current_title = None
                current_doi = None

    print(f"[RIS] 解析完成，共 {len(doi_map)} 条有DOI记录")
    return doi_map


def find_doi(title: str, doi_map: dict) -> str | None:
    """精确匹配，失败则前100字符模糊匹配"""
    key = normalize_title(title)
    if key in doi_map:
        return doi_map[key]
    prefix = key[:100]
    for k, v in doi_map.items():
        if k[:100] == prefix:
            return v
    return None


def safe_filename(title: str, year, max_len=80) -> str:
    name = re.sub(r'[\\/:*?"<>|]', '_', str(title))
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > max_len:
        name = name[:max_len].strip()
    return f"{year}_{name}.pdf"


# ── CORE API下载 ──────────────────────────────────────────────

def try_core(doi: str, save_path: Path) -> bool:
    try:
        api_url = f"https://api.core.ac.uk/v3/works/doi:{doi}"
        r = requests.get(api_url, headers=HEADERS, timeout=20)
        if r.status_code != 200:
            return False
        data = r.json()
        dl_url = data.get("downloadUrl") or data.get("fullTextLink")
        if not dl_url:
            for lk in data.get("links", []):
                if lk.get("type") in ("download", "pdf") and lk.get("url"):
                    dl_url = lk["url"]
                    break
        if not dl_url:
            return False

        pdf_r = requests.get(dl_url, headers=HEADERS, proxies=PROXIES, timeout=40, stream=True)
        if pdf_r.status_code == 200 and "pdf" in pdf_r.headers.get("Content-Type", ""):
            with open(save_path, "wb") as f:
                for chunk in pdf_r.iter_content(8192):
                    f.write(chunk)
            return save_path.stat().st_size > 10_000
    except Exception as e:
        print(f"    [CORE] 异常: {e}")
    return False


# ── Sci-Hub下载 ───────────────────────────────────────────────

def try_scihub(doi: str, save_path: Path) -> tuple[bool, str]:
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print("    [Sci-Hub] 需要安装 beautifulsoup4：pip install beautifulsoup4")
        return False, ""

    for mirror in SCIHUB_MIRRORS:
        try:
            url = f"{mirror}/{doi}"
            r = requests.get(url, headers=HEADERS, timeout=30)
            if r.status_code != 200:
                continue

            soup = BeautifulSoup(r.text, "html.parser")
            pdf_url = None

            iframe = soup.find("iframe", id="pdf")
            if iframe and iframe.get("src"):
                pdf_url = iframe["src"]
            if not pdf_url:
                embed = soup.find("embed", type="application/pdf")
                if embed and embed.get("src"):
                    pdf_url = embed["src"]
            if not pdf_url:
                for a in soup.find_all("a", href=True):
                    if ".pdf" in a["href"].lower():
                        pdf_url = a["href"]
                        break

            if not pdf_url:
                continue

            if pdf_url.startswith("//"):
                pdf_url = "https:" + pdf_url
            elif pdf_url.startswith("/"):
                pdf_url = mirror + pdf_url

            pdf_r = requests.get(pdf_url, headers=HEADERS, proxies=PROXIES, timeout=60, stream=True)
            if pdf_r.status_code == 200:
                ct = pdf_r.headers.get("Content-Type", "")
                if "pdf" in ct or "octet-stream" in ct:
                    with open(save_path, "wb") as f:
                        for chunk in pdf_r.iter_content(8192):
                            f.write(chunk)
                    if save_path.stat().st_size > 10_000:
                        return True, mirror
        except Exception as e:
            print(f"    [Sci-Hub {mirror}] 异常: {e}")
            continue

    return False, ""


# ── 主流程 ────────────────────────────────────────────────────

def main():
    # 1. 读取待下载列表
    # 列：序号(0), 标题(1), 作者(2), 年份(3), 期刊(4), 来源数据库(5),
    #     筛选状态(6), DOI/链接(7), PDF状态(8)
    wb_in = openpyxl.load_workbook(EXCEL_IN, data_only=True)
    ws_in = wb_in.active

    papers = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if not row[1]:  # 标题为空跳过
            continue
        papers.append({
            "seq":     row[0],
            "title":   str(row[1]).strip(),
            "authors": str(row[2] or ""),
            "year":    row[3] or "unknown",
            "journal": str(row[4] or ""),
            "source":  str(row[5] or ""),
            "status":  str(row[6] or ""),
            "doi_col": str(row[7] or "").strip(),  # 用户可能已手填DOI
        })
    wb_in.close()
    print(f"\n待下载：{len(papers)} 篇\n")

    # 2. 加载RIS DOI字典
    doi_map = load_ris_doi_map(RIS_FILE)

    # 3. 逐篇处理
    results = []
    for i, p in enumerate(papers, 1):
        title = p["title"]
        year  = p["year"]
        fname = safe_filename(title, year)
        fpath = PDF_DIR / fname

        # DOI优先用户手填，否则从RIS匹配
        # 清理：去除 https://doi.org/ 前缀；过滤非DOI文本（Conference paper/Article等）
        raw_doi = p["doi_col"]
        if raw_doi:
            raw_doi = re.sub(r'^https?://(dx\.)?doi\.org/', '', raw_doi).strip()
            if not re.match(r'^10\.\d{4,}/', raw_doi):
                raw_doi = ""  # 不像DOI格式，当作空值
        doi = raw_doi if raw_doi else find_doi(title, doi_map)

        print(f"[{i:03d}/{len(papers)}] {title[:65]}...")
        print(f"    DOI: {doi or '未找到'}")

        # 已存在
        if fpath.exists() and fpath.stat().st_size > 10_000:
            print("    → 已存在，跳过")
            results.append({**p, "doi": doi or "", "filename": fname,
                             "dl_status": "已存在", "dl_source": "缓存"})
            continue

        # 无DOI
        if not doi:
            print("    → 无DOI，需手动下载")
            results.append({**p, "doi": "", "filename": fname,
                             "dl_status": "无DOI-手动", "dl_source": ""})
            continue

        # CORE优先
        if try_core(doi, fpath):
            print("    → ✅ CORE下载成功")
            results.append({**p, "doi": doi, "filename": fname,
                             "dl_status": "已下载", "dl_source": "CORE"})
            time.sleep(1)
            continue

        # Sci-Hub
        ok, mirror = try_scihub(doi, fpath)
        if ok:
            print(f"    → ✅ Sci-Hub下载成功 ({mirror})")
            results.append({**p, "doi": doi, "filename": fname,
                             "dl_status": "已下载", "dl_source": f"Sci-Hub({mirror})"})
        else:
            print("    → ❌ 自动下载失败，需手动")
            results.append({**p, "doi": doi, "filename": fname,
                             "dl_status": "手动下载", "dl_source": ""})

        time.sleep(2)

    # 4. 生成追踪Excel
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "全文下载追踪"

    headers = ["序号", "筛选状态", "标题", "作者", "年份", "期刊",
               "DOI", "文件名", "下载状态", "来源", "全文筛选结果", "排除原因", "备注"]

    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red    = PatternFill("solid", fgColor="FFC7CE")
    hfill  = PatternFill("solid", fgColor="2E75B6")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    status_fill = {
        "已下载":     green,
        "已存在":     green,
        "手动下载":   yellow,
        "无DOI-手动": yellow,
    }

    for r in results:
        row_data = [
            r["seq"], r["status"], r["title"], r["authors"],
            r["year"], r["journal"], r["doi"], r["filename"],
            r["dl_status"], r["dl_source"], "", "", ""
        ]
        ws.append(row_data)
        rn = ws.max_row
        fill = status_fill.get(r["dl_status"], red)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(rn, col)
            cell.fill = fill
            if col in (3, 4):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")
        ws.row_dimensions[rn].height = 50

    col_widths = [6, 10, 60, 30, 6, 28, 38, 50, 14, 22, 14, 16, 20]
    for col_i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # 统计sheet
    ws2 = wb_out.create_sheet("统计说明")
    n_ok  = sum(1 for r in results if r["dl_status"] in ("已下载", "已存在"))
    n_man = len(results) - n_ok
    summary = [
        ("颜色", "含义"),
        ("绿色", "PDF已自动下载或已存在"),
        ("黄色", "需手动下载（无DOI或自动失败）"),
        ("", ""),
        ("统计", "数量"),
        ("自动下载成功", n_ok),
        ("需手动下载",   n_man),
        ("总计",        len(results)),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for row in summary:
        ws2.append(list(row))

    wb_out.save(EXCEL_OUT)

    print(f"\n{'='*60}")
    print(f"完成！自动下载：{n_ok}/{len(results)}")
    print(f"需手动下载：{n_man} 篇")
    print(f"PDF目录：{PDF_DIR}")
    print(f"追踪Excel：{EXCEL_OUT}")


if __name__ == "__main__":
    main()
