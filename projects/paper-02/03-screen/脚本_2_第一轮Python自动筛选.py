"""
脚本_2_第一轮Python自动筛选.py
Paper-02 第一轮自动关键词筛选
输入：02-search/数据_2_四库合并去重后.ris（1827条）
输出：03-screen/数据_2_第一轮自动筛选结果.xlsx
      03-screen/结果_1_第一轮筛选统计.json

筛选逻辑（paper-02 PICOS）：
- E1：年龄不符（无0-8岁相关词）
- E2：暴露不符（无父母教育/SES词）
- E3：无神经指标（无EEG/ERP/fNIRS/fMRI/DTI等词）
- E4：综述/元分析/protocol
- 保留：通过所有检查
"""

import json
import re
from collections import Counter
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 路径配置 ─────────────────────────────────────────────────
BASE      = Path(r"E:\Meta-analysis writing project\projects\paper-02")
RIS_FILE  = BASE / "02-search" / "数据_2_四库合并去重后.ris"
OUT_DIR   = BASE / "03-screen"
OUT_EXCEL = OUT_DIR / "数据_2_第一轮自动筛选结果.xlsx"
OUT_STATS = OUT_DIR / "结果_1_第一轮筛选统计.json"
OUT_DIR.mkdir(exist_ok=True)

# ── 颜色定义 ─────────────────────────────────────────────────
COLORS = {
    "保留": "C6EFCE",   # 绿
    "E1":   "FFCCCC",   # 红  — 年龄不符
    "E2":   "FFD7A8",   # 橙  — 暴露不符
    "E3":   "FFEB9C",   # 黄  — 无神经指标
    "E4":   "D9D9D9",   # 灰  — 综述/元分析
}

# ── 筛选词表 ─────────────────────────────────────────────────

# 概念C：0-8岁年龄相关词（任一命中 → 年龄可能符合）
AGE_WORDS = [
    r'\binfant\b', r'\binfants\b', r'\bnewborn\b', r'\bnewborns\b',
    r'\bneonatal\b', r'\bneonate\b', r'\bneonates\b', r'\bperinatal\b',
    r'\btoddler\b', r'\btoddlers\b',
    r'\bpreschool\b', r'\bpre-school\b', r'\bpreschooler\b',
    r'\bkindergarten\b',
    r'\bearly childhood\b', r'\bearly brain development\b',
    r'\byoung children\b',
    r'school.aged child',   # school-aged children
    r'\bage[sd]? [0-8]\b',  # aged 0-8 / ages 3
    r'\b[0-8].year.old\b',  # 3-year-old
    r'\bbaby\b', r'\bbabies\b',
]

# 概念A：父母教育/SES词（任一命中 → 暴露可能符合）
SES_WORDS = [
    r'parental education', r'maternal education', r'paternal education',
    r'educational attainment', r'years of education', r'educational level',
    r'parental schooling', r'socioeconomic status', r'\bses\b',
    r'family income', r'household income',
    r'\bpoverty\b', r'child poverty', r'income.to.needs',
    r'\blow.income\b', r'household wealth',
    r'neighborhood disadvantage', r'socioeconomic background',
    r'socioeconomic disadvantage', r'social class',
    r'parental socioeconomic', r'family socioeconomic',
]

# 概念B（精确层）：有明确神经测量方法的词（任一命中 → 确定有神经测量）
NEURAL_SPECIFIC = [
    r'\beeg\b', r'electroencephalograph', r'electrophysiolog',
    r'\berp\b', r'event.related potential',
    r'\bfmri\b', r'functional mri', r'structural mri', r'neuroimaging',
    r'brain imaging',
    r'\brseeg\b', r'spectral power', r'alpha power', r'theta power',
    r'\bdti\b', r'diffusion tensor', r'functional connectivity',
    r'resting.state',
    r'\bfnirs\b', r'near.infrared spectroscopy', r'\bnirs\b',
    r'\bmeg\b', r'magnetoencephalograph',
    r'visual evoked potential', r'mismatch negativity', r'\bmmn\b',
    r'\bn400\b', r'\bp300\b', r'\bnc\b',
    r'cortical thickness', r'cortical surface area',
    r'gray matter', r'grey matter', r'white matter',
    r'brain structure', r'brain morpholog',
    r'voxel.based morphometry', r'\bvbm\b',
]

# 概念B（宽泛层）：单独出现不足以判定有神经测量（需同时命中精确层才保留）
NEURAL_BROAD = [
    r'brain development',
    r'brain function',
    r'brain activity',
]

# 合并，用于判断"是否有任何神经相关词"
NEURAL_WORDS = NEURAL_SPECIFIC + NEURAL_BROAD

# E4：综述/元分析识别词（标题精确，摘要宽松）
E4_TITLE = [
    r'\bsystematic review\b', r'\bmeta.?analysis\b',
    r'\bnarrative review\b', r'\bscoping review\b',
    r'\bliterature review\b', r'\beditorial\b',
    r'\bcommentary\b', r'\bletter to\b', r'\bcase report\b',
    r'\bprotocol\b',
]
E4_ABSTRACT = [
    r'this systematic review', r'this meta.?analysis',
    r'this scoping review', r'this narrative review',
    r'we conducted a (systematic )?review', r'we performed a meta',
]

# ── 工具函数 ──────────────────────────────────────────────────

def parse_ris(filepath):
    records = []
    current = {}
    with open(filepath, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\n')
            if line.startswith('ER  -') or line.startswith('ER -'):
                if current:
                    records.append(current)
                    current = {}
            elif len(line) >= 6 and line[2:6] == '  - ':
                tag = line[:2].strip()
                val = line[6:].strip()
                if tag in current:
                    if isinstance(current[tag], list):
                        current[tag].append(val)
                    else:
                        current[tag] = [current[tag], val]
                else:
                    current[tag] = val
    if current:
        records.append(current)
    return records


def get_field(rec, *tags):
    for tag in tags:
        v = rec.get(tag, '')
        if isinstance(v, list):
            v = ' '.join(str(x) for x in v)
        if v:
            return str(v).strip()
    return ''


def get_year(rec):
    y = get_field(rec, 'PY', 'Y1', 'DA')
    m = re.search(r'\d{4}', y)
    return int(m.group()) if m else 0


def get_authors(rec):
    au = rec.get('AU') or rec.get('A1') or ''
    if isinstance(au, list):
        s = '; '.join(str(a) for a in au[:3])
        return s + ' et al.' if len(au) > 3 else s
    return str(au)


def any_match(text, patterns):
    t = text.lower()
    return any(re.search(p, t) for p in patterns)


# ── 筛选逻辑 ─────────────────────────────────────────────────

def screen(rec):
    title    = get_field(rec, 'TI', 'T1')
    abstract = get_field(rec, 'AB', 'N2')
    combined = (title + ' ' + abstract).lower()

    # E4：综述/元分析（优先判断，避免误排实质性文献）
    if any_match(title.lower(), E4_TITLE):
        return 'E4', '标题为综述/元分析/protocol'
    if any_match(abstract.lower(), E4_ABSTRACT):
        return 'E4', '摘要显示为综述'

    # E1：年龄不符（标题+摘要中无0-8岁相关词）
    if not any_match(combined, AGE_WORDS):
        return 'E1', '无0-8岁年龄相关词'

    # E2：暴露不符（无父母教育/SES词）
    if not any_match(combined, SES_WORDS):
        return 'E2', '无父母教育/SES相关词'

    # E3：无神经指标（必须命中至少一个精确层神经词；宽泛词单独不足）
    if not any_match(combined, NEURAL_WORDS):
        return 'E3', '无任何神经相关词'
    if not any_match(combined, NEURAL_SPECIFIC):
        return 'E3', '仅宽泛神经词(brain development/function/activity)，无具体测量指标'

    return '保留', ''


# ── 主流程 ───────────────────────────────────────────────────

print('读取文献...')
records = parse_ris(RIS_FILE)
print(f'共 {len(records)} 条')

print('开始筛选...')
results = []
for rec in records:
    decision, reason = screen(rec)
    results.append({'rec': rec, 'decision': decision, 'reason': reason})

counts = Counter(r['decision'] for r in results)
print('\n筛选结果：')
for k in ['保留', 'E1', 'E2', 'E3', 'E4']:
    print(f'  {k}: {counts.get(k, 0)}')
print(f'  总计: {len(results)}')

# ── 写Excel ──────────────────────────────────────────────────

print('\n生成Excel...')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '第一轮自动筛选'

headers = ['#', '筛选结果', '排除原因', '标题', '作者', '年份',
           '期刊', '来源数据库', '摘要']

thin   = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border
ws.row_dimensions[1].height = 30

for i, item in enumerate(results, 1):
    rec      = item['rec']
    decision = item['decision']
    reason   = item['reason']

    title    = get_field(rec, 'TI', 'T1')
    authors  = get_authors(rec)
    year     = get_year(rec) or ''
    journal  = get_field(rec, 'JO', 'JF', 'T2')
    source   = get_field(rec, 'DB')
    abstract = get_field(rec, 'AB', 'N2')
    if len(abstract) > 500:
        abstract = abstract[:500] + '...'

    row_data = [i, decision, reason, title, authors, year, journal, source, abstract]
    fill     = PatternFill('solid', fgColor=COLORS.get(decision, 'FFFFFF'))
    row_num  = i + 1

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=(col in (4, 9)))
        if col in (1, 2, 6, 8):
            cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.row_dimensions[row_num].height = 60

col_widths = [6, 10, 28, 60, 30, 8, 30, 12, 80]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

# 图例sheet
ws2 = wb.create_sheet('图例说明')
legend = [
    ('颜色',  '结果', '说明'),
    ('绿色',  '保留', '进入第二轮xhs+人工摘要筛选'),
    ('红色',  'E1',   '年龄不符：无0-8岁相关词'),
    ('橙色',  'E2',   '暴露不符：无父母教育/SES词'),
    ('黄色',  'E3',   '无神经测量指标词'),
    ('灰色',  'E4',   '综述/元分析/protocol'),
]
legend_colors = ['4472C4', 'C6EFCE', 'FFCCCC', 'FFD7A8', 'FFEB9C', 'D9D9D9']
for r, (row_data, color) in enumerate(zip(legend, legend_colors), 1):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(bold=(r == 1), color='FFFFFF' if r == 1 else '000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws2.row_dimensions[r].height = 22
for col, width in enumerate([10, 10, 40], 1):
    ws2.column_dimensions[get_column_letter(col)].width = width

wb.save(OUT_EXCEL)
print(f'Excel已保存：{OUT_EXCEL}')

# ── 统计JSON ─────────────────────────────────────────────────
stats = {
    'total_input': len(records),
    'retained': counts.get('保留', 0),
    'E1_age': counts.get('E1', 0),
    'E2_exposure': counts.get('E2', 0),
    'E3_neural': counts.get('E3', 0),
    'E4_review': counts.get('E4', 0),
    'excluded_total': len(records) - counts.get('保留', 0),
}
with open(OUT_STATS, 'w', encoding='utf-8') as f:
    json.dump(stats, f, ensure_ascii=False, indent=2)
print(f'统计已保存：{OUT_STATS}')
print('\n完成。')
