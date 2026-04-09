"""
脚本_1_生成数据提取表模板.py
按 v2 字段方案（34列）生成数据_7_数据提取表_v1.xlsx
含下拉菜单、冻结首行、列宽适配、条件格式
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUT = 'E:/Meta-analysis writing project/projects/paper-02/04-extract/数据_7_数据提取表_v1.xlsx'

# ── 样式工具 ────────────────────────────────────────────
thin = Side(style='thin', color='B8CCE4')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def hdr(cell, bg, text=None):
    if text:
        cell.value = text
    cell.fill = fill(bg)
    cell.font = Font(bold=True, color='FFFFFF', size=9)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

def data_cell(cell):
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = border
    cell.font = Font(size=9)

# ── 表头定义（34列） ─────────────────────────────────────
# (字段名, 说明, 宽度, 分组色)
HEADERS = [
    # A. 基本信息 — 深蓝 2F5496
    ('A1 序号',           '筛选表原序号',                                    7,  '2F5496'),
    ('A2 第一作者',        '姓氏',                                           12,  '2F5496'),
    ('A3 年份',           '发表年',                                           7,  '2F5496'),
    ('A4 国家/地区',       '样本来源国',                                      10,  '2F5496'),
    ('A5 研究设计',        '横断/纵向/RCT子分析',                             14,  '2F5496'),
    ('A6 样本量(总N)',     '最终分析样本',                                     9,  '2F5496'),
    ('A7 儿童年龄',        '均值±SD或范围',                                   14,  '2F5496'),
    ('A8 追踪时间点数',    '横断=1；纵向填时间点数',                           10,  '2F5496'),
    ('A9 年龄分组',        'neonatal/infant/toddler/preschool/early-school',  16,  '2F5496'),
    # B. SES — 深绿 375623
    ('B1 教育测量方式',    '年数/等级/学历分类/复合SES',                       18,  '375623'),
    ('B2 教育报告方式',    '单独报告/复合SES分项/两者均有',                    16,  '375623'),
    ('B3 教育编码',        '连续变量/二分/多分类',                             14,  '375623'),
    ('B4 效应提取类型',    'beta/r/group-comparison（数值填D3）',              16,  '375623'),
    ('B5 其他SES指标',     '同时报告的收入/职业等',                            16,  '375623'),
    # C. 神经测量 — 深紫 4B2E83
    ('C1 测量模态',        'rsEEG/ERP/fNIRS/fMRI/sMRI/DTI/MEG',              14,  '4B2E83'),
    ('C2 具体指标',        '频段/成分/结构指标名称',                           20,  '4B2E83'),
    ('C3 任务类型',        '静息态/任务态(任务名)/结构',                       18,  '4B2E83'),
    ('C4 分析方法',        '时频/ICA/GLM/VBM/tractography等',                 18,  '4B2E83'),
    ('C5 脑区/电极',       '主要分析脑区或电极位置',                           18,  '4B2E83'),
    ('C6 同一样本标记',    '是/否',                                            10,  '4B2E83'),
    ('C7 队列/数据集名称', '大型队列名称（如BFY/ABCD）',                       22,  '4B2E83'),
    # D. 主要结局 — 深红 7B1E1E
    ('D1 教育×神经关系',  '显著正相关/显著负相关/无显著/混合',                16,  '7B1E1E'),
    ('D2 效应方向描述',    '一句话核心发现（写作时直接引用）',                  40,  '7B1E1E'),
    ('D3 效应量/统计值',   '原始值：β=0.34/r=0.21/d=0.52',                   18,  '7B1E1E'),
    ('D3b 效应方向',       'positive/negative/null/mixed',                    14,  '7B1E1E'),
    ('D4 统计方法',        '回归/ANOVA/相关/LMM等',                           14,  '7B1E1E'),
    ('D5 协变量控制',      '分析中控制了哪些变量',                             24,  '7B1E1E'),
    ('D6 调节/中介',       '是否检验中介/调节及结论',                          20,  '7B1E1E'),
    # E. 机制路径 — 棕 7F4E18
    ('E1 机制路径类型',    '认知刺激/语言输入/慢性应激/营养/其他/未检验',      22,  '7F4E18'),
    ('E2 路径检验方法',    '中介分析/调节分析/路径分析/未检验',                18,  '7F4E18'),
    ('E3 路径结论',        '显著中介/不显著/部分中介/未检验',                  18,  '7F4E18'),
    # F. NOS偏倚风险 — 深灰 404040
    ('F1 NOS-选择(/4)',    '样本代表性+暴露+结局确认',                          12,  '404040'),
    ('F2 NOS-可比(/2)',    '协变量控制',                                        10,  '404040'),
    ('F3 NOS-结局(/3)',    '结局评估+随访+失访',                                10,  '404040'),
    ('F4 NOS总分(/9)',     '自动求和',                                           9,  '404040'),
    # G. 备注 — 中灰 595959
    ('G1 特殊说明',        '早产对照/年龄跨越/数据不可提取等边界情况',          28,  '595959'),
    ('G2 提取者',          'oo/cc',                                             8,  '595959'),
]

# 下拉菜单配置
DROPDOWNS = {
    'A5':  '"横断,纵向,RCT子分析"',
    'A9':  '"neonatal(0-1m),infant(1-12m),toddler(1-3y),preschool(3-5y),early-school(6-8y)"',
    'B2':  '"单独报告,复合SES分项,两者均有"',
    'B3':  '"连续变量,二分,多分类"',
    'B4':  '"beta,r,group-comparison,other"',
    'C1':  '"rsEEG,ERP,fNIRS,fMRI,sMRI,DTI,MEG"',
    'C3':  '"静息态,任务态-oddball,任务态-go-nogo,任务态-language,任务态-attention,任务态-social,任务态-other,结构"',
    'C6':  '"是,否"',
    'D1':  '"显著正相关,显著负相关,无显著,混合"',
    'D3b': '"positive,negative,null,mixed"',
    'E1':  '"认知刺激,语言输入,慢性应激,营养,其他,未检验"',
    'E2':  '"中介分析,调节分析,路径分析,未检验"',
    'E3':  '"显著中介,不显著,部分中介,未检验"',
    'G2':  '"oo,cc"',
}

# ── 建表 ────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '数据提取'

# 第1行：分组标签（合并单元格）
groups = [
    ('A. 基本信息', 1, 9, '2F5496'),
    ('B. SES/父母教育', 10, 14, '375623'),
    ('C. 神经测量', 15, 21, '4B2E83'),
    ('D. 主要结局', 22, 28, '7B1E1E'),
    ('E. 机制路径', 29, 31, '7F4E18'),
    ('F. NOS偏倚风险', 32, 35, '404040'),
    ('G. 备注', 36, 37, '595959'),
]

for label, c_start, c_end, color in groups:
    ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
    c = ws.cell(1, c_start, label)
    c.fill = fill(color)
    c.font = Font(bold=True, color='FFFFFF', size=10)
    c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 20

# 第2行：字段名表头
col_letters = {}  # 字段名前缀 → 列号
for ci, (fname, fdesc, fwidth, fcolor) in enumerate(HEADERS, 1):
    c = ws.cell(2, ci)
    hdr(c, fcolor, fname)
    ws.column_dimensions[get_column_letter(ci)].width = fwidth
    # 记录列字母映射（取字段名前2字符如A1,B2等）
    col_letters[fname[:2]] = ci

ws.row_dimensions[2].height = 36

# 第3行：说明行（淡色背景）
for ci, (fname, fdesc, fwidth, fcolor) in enumerate(HEADERS, 1):
    c = ws.cell(3, ci, fdesc)
    c.fill = fill('F2F2F2')
    c.font = Font(italic=True, size=8, color='595959')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.border = border
ws.row_dimensions[3].height = 28

# 预留200行数据区
MAX_ROW = 203
for r in range(4, MAX_ROW + 1):
    for ci in range(1, len(HEADERS) + 1):
        data_cell(ws.cell(r, ci))
    ws.row_dimensions[r].height = 18

# F4自动求和公式（F4=F1+F2+F3）
f1_col = col_letters.get('F1', 32)
f2_col = col_letters.get('F2', 33)
f3_col = col_letters.get('F3', 34)
f4_col = col_letters.get('F4', 35)
for r in range(4, MAX_ROW + 1):
    ws.cell(r, f4_col).value = (
        f'=IFERROR({get_column_letter(f1_col)}{r}'
        f'+{get_column_letter(f2_col)}{r}'
        f'+{get_column_letter(f3_col)}{r},"")'
    )
    ws.cell(r, f4_col).font = Font(size=9, bold=True)
    ws.cell(r, f4_col).alignment = Alignment(horizontal='center', vertical='center')

# ── 下拉菜单 ─────────────────────────────────────────────
for field_prefix, formula in DROPDOWNS.items():
    col_idx = col_letters.get(field_prefix)
    if not col_idx:
        continue
    col_letter = get_column_letter(col_idx)
    dv = DataValidation(
        type='list',
        formula1=formula,
        allow_blank=True,
        showDropDown=False
    )
    dv.sqref = f'{col_letter}4:{col_letter}{MAX_ROW}'
    ws.add_data_validation(dv)

# ── 冻结前3行+A列 ────────────────────────────────────────
ws.freeze_panes = 'B4'

# ── 自动筛选（第2行表头行） ──────────────────────────────
ws.auto_filter.ref = f'A2:{get_column_letter(len(HEADERS))}2'

# ── Sheet2: 字段说明速查 ─────────────────────────────────
ws2 = wb.create_sheet('字段说明')
ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 30

ws2.cell(1, 1, 'Paper-02 数据提取表 v1 — 字段说明速查（v2方案，2026-04-09）')
ws2.cell(1, 1).font = Font(bold=True, size=12, color='2F5496')
ws2.merge_cells('A1:D1')
ws2.row_dimensions[1].height = 24

ws2.cell(2, 1, '列号').font = Font(bold=True)
ws2.cell(2, 2, '字段名').font = Font(bold=True)
ws2.cell(2, 3, '说明').font = Font(bold=True)
ws2.cell(2, 4, '下拉选项').font = Font(bold=True)

dropdown_display = {k: v.strip('"') for k, v in DROPDOWNS.items()}

for ri, (fname, fdesc, _, fcolor) in enumerate(HEADERS, 3):
    prefix = fname[:2]
    ws2.cell(ri, 1, prefix)
    ws2.cell(ri, 2, fname[3:] if len(fname) > 3 else fname)
    ws2.cell(ri, 3, fdesc)
    ws2.cell(ri, 4, dropdown_display.get(prefix, '自由填写'))
    for ci in range(1, 5):
        ws2.cell(ri, ci).font = Font(size=9)
        ws2.cell(ri, ci).alignment = Alignment(vertical='center', wrap_text=True)
        ws2.cell(ri, ci).border = border
    ws2.row_dimensions[ri].height = 16

# ── Sheet3: 多模态拆行说明 ───────────────────────────────
ws3 = wb.create_sheet('多模态拆行规则')
ws3.column_dimensions['A'].width = 16
ws3.column_dimensions['B'].width = 60

rules = [
    ('多模态拆行规则', ''),
    ('适用情况', '同一篇文章同时报告EEG+MRI或其他多模态结果'),
    ('A1-A8', '两行填写完全相同的基本信息'),
    ('A9', '同一篇文章两行填写相同年龄分组'),
    ('C1', '第1行填rsEEG，第2行填sMRI（各填对应模态）'),
    ('C2-C5', '按各自模态分别填写'),
    ('C6', '两行均填"是"，C7填相同序号'),
    ('D列', 'D1-D6按各模态结果分别填写'),
    ('统计文献总数', '按A1去重 = 76篇；数据行数 > 76正常'),
    ('示例', 'A1均填#23，A2均填Neville，C1第1行rsEEG/第2行sMRI'),
]

ws3.cell(1, 1, 'Paper-02 多模态研究拆行规则（oo-cc 2026-04-09确认）')
ws3.cell(1, 1).font = Font(bold=True, size=11, color='2F5496')
ws3.merge_cells('A1:B1')
ws3.row_dimensions[1].height = 22

for ri, (k, v) in enumerate(rules, 2):
    ws3.cell(ri, 1, k).font = Font(bold=(ri == 2), size=9)
    ws3.cell(ri, 2, v).font = Font(size=9)
    ws3.cell(ri, 2).alignment = Alignment(wrap_text=True)
    ws3.row_dimensions[ri].height = 18

wb.save(OUT)
print(f'生成成功：{OUT}')
print(f'共 {len(HEADERS)} 列，含下拉菜单 {len(DROPDOWNS)} 列')
print('Sheet列表：数据提取 / 字段说明 / 多模态拆行规则')
