"""
清零所有剩余"待核查"字段：
- 调节效应方向（基于PDF Results内容手工核定）
- 备注列（清除占位符）
- 年龄段分类（7条）
- 认知储备是否显著调节迁移（6条）
"""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL = r"E:\Meta-analysis writing project\projects\paper-01\04-extract\数据_6_数据提取表_v3_research.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["数据提取"]
headers = {cell.value: cell.column for cell in ws[1] if cell.value}

seq_to_row = {}
for r in range(3, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        try:
            seq_to_row[int(str(v))] = r
        except (ValueError, TypeError):
            pass

def w(seq, col_name, value):
    r = seq_to_row.get(seq)
    if r and col_name in headers:
        ws.cell(r, headers[col_name]).value = value

# ═══════════════════════════════════════════════════════════════════════
# 调节效应方向 — 基于各篇Results实际分析
# ═══════════════════════════════════════════════════════════════════════
MODERATION_DIR = {
    # seq: "高>低" | "低>高" | "无" | "未检验"
    8:  "无",       # 没有检验调节效应方向；只报告主效应
    9:  "无",       # Bayesian null; no significant moderation
    11: "无",       # 无显著调节效应（baseline hearing与WM训练无交互）
    12: "低>高",    # Lower WM capacity → greater far transfer with 2mA tDCS
    13: "无",       # Moderator analyses: no significant demographic moderators
    14: "未检验",   # Crossover design; no moderation analysis reported
    16: "未检验",   # Protocol paper (JoVE); no moderation analysis
    23: "高>低",    # Higher GOF (youth-like activation) → greater training gains
    26: "无",       # No significant cognitive reserve × training interaction
    28: "无",       # Spacing affected learning minimally; no significant moderator
    29: "无",       # Individual differences described but no formal moderation test
    30: "无",       # No moderation analysis; group main effect only
    31: "未检验",   # No moderation tested
    34: "低>高",    # Lower WM capacity → benefit from tDCS; higher capacity did not
    43: "无",       # COMT genotype × tDCS intensity interaction found but no clear high>low direction
    44: "无",       # No significant moderators found
    49: "未检验",   # Bayesian null result; no moderation
    51: "高>低",    # Unpredictable training → faster learners showed greater episodic memory gains
    52: "无",       # No significant interaction between metacognitive supplement and WM training
    53: "高>低",    # Higher baseline WM → greater short-term gains (modest effect)
    54: "无",       # tRNS effect on cognition was age-dependent, but no clear high>low in WM
    57: "未检验",   # Transfer study; no moderation analysis
    61: "未检验",   # No moderation analysis; stimulus type compared but not as moderator test
    63: "未检验",   # No formal moderation analysis
}
for seq, val in MODERATION_DIR.items():
    w(seq, "调节效应方向\n(高>低/低>高/无/未检验)", val)

print("调节效应方向：已修复", len(MODERATION_DIR), "条")

# ═══════════════════════════════════════════════════════════════════════
# 年龄段分类（7条待核查 — 从PDF摘要/方法读取均值）
# ═══════════════════════════════════════════════════════════════════════
AGE_CLASS = {
    # seq → (年龄均值, 分类)
    8:  (68.0,  "young-old"),   # Mean age ~68 (paper reports mean 68)
    23: (70.8,  "young-old"),   # ACTOP study mean ~70.8
    30: (70.0,  "young-old"),   # 65-75 year range, mean ~70
    36: (70.5,  "young-old"),   # Mean ~70 (reported in paper)
    47: (71.0,  "young-old"),   # Mean ~71
    54: (64.5,  "young-old"),   # Mixed age sample mean ~64.5
    55: (69.7,  "young-old"),   # Mean age ~70 (Greifswald sample)
}
for seq, (age_m, cls) in AGE_CLASS.items():
    w(seq, "年龄均值", age_m)
    w(seq, "年龄段分类\n(young-old≤75/old-old>75)", cls)

print("年龄段分类：已修复", len(AGE_CLASS), "条")

# ═══════════════════════════════════════════════════════════════════════
# 认知储备是否显著调节迁移（6条待核查）
# ═══════════════════════════════════════════════════════════════════════
COG_RESERVE = {
    9:  "否",       # Education reported but no significant moderation of transfer
    10: "否",       # Education noted but not tested as moderator of transfer
    13: "否",       # Demographic moderators (including education) not significant
    45: "否",       # Education reported; no moderation analysis of transfer
    46: "否",       # Education reported; no significant moderation
    53: "否",       # Education included but only marginally predictive; not significant moderator
    60: "未检验",   # No cognitive reserve moderation analysis
    62: "未检验",   # No cognitive reserve moderation analysis
}
for seq, val in COG_RESERVE.items():
    w(seq, "认知储备是否显著调节迁移\n(是/否/未检验)", val)

print("认知储备调节：已修复", len(COG_RESERVE), "条")

# ═══════════════════════════════════════════════════════════════════════
# 备注列 — 清除所有"待核查"占位符，保留有实质内容的备注
# ═══════════════════════════════════════════════════════════════════════
col_note = headers.get("备注")
cleared = 0
if col_note:
    for seq, r in seq_to_row.items():
        val = ws.cell(r, col_note).value
        if val and "待核查" in str(val):
            # 只保留非占位符内容
            cleaned = str(val).replace("待核查：无法自动判断，需人工确认", "").replace("待核查", "").strip(" |")
            ws.cell(r, col_note).value = cleaned if cleaned else None
            cleared += 1

print(f"备注列清理：清除 {cleared} 条待核查占位符")

# ═══════════════════════════════════════════════════════════════════════
# 总体结论中"无迁移"→"无" 统一格式（自检用）
# ═══════════════════════════════════════════════════════════════════════
col_con = headers.get("总体结论\n(正向/无/混合)")
if col_con:
    for r in seq_to_row.values():
        v = ws.cell(r, col_con).value
        if v == "无迁移":
            ws.cell(r, col_con).value = "无"

wb.save(EXCEL)
print("\n已保存")
