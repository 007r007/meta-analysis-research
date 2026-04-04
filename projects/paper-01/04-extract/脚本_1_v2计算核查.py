"""
Paper-01 数据提取表 v2 — 计算字段 + 异常值核查
输入：数据_6_数据提取表_v2.xlsx
输出：
  1. 更新 v2（训练频率 + 年龄段分类）
  2. 核查报告_异常值.md
"""

import openpyxl
from openpyxl import load_workbook
from datetime import datetime
from copy import copy
import os

BASE = os.path.dirname(os.path.abspath(__file__))
V2_PATH = os.path.join(BASE, "数据_6_数据提取表_v2.xlsx")
REPORT_PATH = os.path.join(BASE, "核查报告_异常值.md")

# ── 读取工作簿 ────────────────────────────────────────────────
wb = load_workbook(V2_PATH)
ws = wb["数据提取"]

# ── 第1行：列名 → 列号映射 ────────────────────────────────────
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers[str(val).strip()] = col

print("=== 所有列名（列号：列名）===")
for name, idx in headers.items():
    print(f"  列{idx:>3}: {name!r}")

# ── 定位目标列 ────────────────────────────────────────────────
def find_col(key, fallbacks=()):
    if key in headers:
        return headers[key]
    for fb in fallbacks:
        if fb in headers:
            return headers[fb]
    return None

col_sessions  = find_col("训练总次数\n(sessions)", ["训练总次数(sessions)", "训练总次数"])
col_weeks     = find_col("训练周数")
col_freq      = find_col("训练频率\n(次/周)",      ["训练频率(次/周)", "训练频率"])
col_age_mean  = find_col("年龄均值")
col_age_cat   = find_col("年龄段分类\n(young-old≤75/old-old>75)",
                          ["年龄段分类", "年龄段分类(young-old≤75/old-old>75)"])
col_author    = find_col("第一作者")
col_year      = find_col("年份")
col_seq       = find_col("序号")
col_n_total   = find_col("样本量N", ["总样本量N", "样本量\nN", "N（总）"])
col_duration  = find_col("每次时长\n(分钟)", ["每次时长（分钟）", "每次时长\n（分钟）",
                                            "每次时长(分钟)", "每次训练时长（分钟）"])

print(f"\n目标列定位：")
print(f"  训练总次数={col_sessions}, 训练周数={col_weeks}, 训练频率={col_freq}")
print(f"  年龄均值={col_age_mean}, 年龄段分类={col_age_cat}")
print(f"  第一作者={col_author}, 年份={col_year}, 序号={col_seq}")
print(f"  样本量N={col_n_total}, 每次时长={col_duration}")

# ── 数据行：第3行起（第2行=颜色说明行） ────────────────────────
DATA_START = 3
data_rows = []
for row in range(DATA_START, ws.max_row + 1):
    seq = ws.cell(row=row, column=col_seq).value if col_seq else None
    if seq is None:
        break
    data_rows.append(row)

print(f"\n数据行范围：第{DATA_START}行 ~ 第{data_rows[-1]}行，共{len(data_rows)}行")

# ── 第三步：计算并写回 ────────────────────────────────────────
freq_computed = 0
freq_missing  = 0
age_cat_young = 0
age_cat_old   = 0
age_cat_missing = 0

for row in data_rows:
    # 计算训练频率
    if col_freq:
        sessions = ws.cell(row=row, column=col_sessions).value if col_sessions else None
        weeks    = ws.cell(row=row, column=col_weeks).value    if col_weeks    else None
        try:
            s = float(sessions)
            w = float(weeks)
            if w > 0:
                freq = round(s / w, 1)
                ws.cell(row=row, column=col_freq).value = freq
                freq_computed += 1
            else:
                ws.cell(row=row, column=col_freq).value = None
                freq_missing += 1
        except (TypeError, ValueError):
            ws.cell(row=row, column=col_freq).value = None
            freq_missing += 1

    # 计算年龄段分类
    if col_age_cat:
        age = ws.cell(row=row, column=col_age_mean).value if col_age_mean else None
        try:
            a = float(age)
            if a <= 75:
                ws.cell(row=row, column=col_age_cat).value = "young-old"
                age_cat_young += 1
            else:
                ws.cell(row=row, column=col_age_cat).value = "old-old"
                age_cat_old += 1
        except (TypeError, ValueError):
            ws.cell(row=row, column=col_age_cat).value = None
            age_cat_missing += 1

print(f"\n计算结果：")
print(f"  训练频率：成功{freq_computed}篇，缺失{freq_missing}篇")
print(f"  年龄段分类：young-old={age_cat_young}，old-old={age_cat_old}，缺失={age_cat_missing}")

# ── 第四步：异常值检测 ────────────────────────────────────────
CHECKS = [
    ("年龄均值",     col_age_mean,  60,  95),
    ("样本量N",      col_n_total,   10, 500),
    ("训练总次数",   col_sessions,   3, 100),
    ("训练周数",     col_weeks,      1,  52),
    ("训练频率",     col_freq,       1,  14),
    ("每次时长（分钟）", col_duration, 10, 120),
]

anomalies = []  # (seq, author, year, field, value, lo, hi)

for row in data_rows:
    seq    = ws.cell(row=row, column=col_seq).value    if col_seq    else "?"
    author = ws.cell(row=row, column=col_author).value if col_author else "?"
    year   = ws.cell(row=row, column=col_year).value   if col_year   else "?"

    for field, col, lo, hi in CHECKS:
        if col is None:
            continue
        raw = ws.cell(row=row, column=col).value
        try:
            val = float(raw)
            if val < lo or val > hi:
                anomalies.append((seq, author, year, field, val, lo, hi))
        except (TypeError, ValueError):
            pass  # 空值或非数字，不报异常

print(f"\n异常值：共{len(anomalies)}条")
for a in anomalies:
    print(f"  序号{a[0]} {a[1]} {a[2]}｜{a[3]}={a[4]}（范围{a[5]}-{a[6]}）")

# ── 第五步：缺失值统计 ────────────────────────────────────────
MISSING_FIELDS = [
    ("年龄均值",      col_age_mean),
    ("样本量N",       col_n_total),
    ("训练总次数",    col_sessions),
    ("训练周数",      col_weeks),
    ("每次时长（分钟）", col_duration),
    ("性别(%女)",     find_col("性别(%女)", ["性别(女%)", "性别（女%）", "性别\n(%女)"])),
    ("教育年限",      find_col("教育年限",  ["教育年限（年）", "受教育年限"])),
]

missing_stats = []
for fname, fcol in MISSING_FIELDS:
    if fcol is None:
        missing_stats.append((fname, "列不存在", "—"))
        continue
    missing_rows = []
    for row in data_rows:
        val = ws.cell(row=row, column=fcol).value
        if val is None or str(val).strip() == "":
            seq    = ws.cell(row=row, column=col_seq).value    if col_seq    else "?"
            author = ws.cell(row=row, column=col_author).value if col_author else "?"
            year   = ws.cell(row=row, column=col_year).value   if col_year   else "?"
            missing_rows.append(f"{author} {year}")
    missing_stats.append((fname, len(missing_rows), "; ".join(missing_rows) if missing_rows else "无"))

# ── 第五步：优先核查列表 ────────────────────────────────────────
# 统计每篇的"问题数"（异常值数 + 关键字段缺失数）
priority = {}  # seq -> {info, score}
for row in data_rows:
    seq    = ws.cell(row=row, column=col_seq).value    if col_seq    else row
    author = ws.cell(row=row, column=col_author).value if col_author else "?"
    year   = ws.cell(row=row, column=col_year).value   if col_year   else "?"
    priority[seq] = {"author": author, "year": year, "score": 0, "issues": []}

for a in anomalies:
    seq = a[0]
    if seq in priority:
        priority[seq]["score"] += 2
        priority[seq]["issues"].append(f"异常：{a[3]}={a[4]}")

for fname, fcol in MISSING_FIELDS[:5]:  # 只计前5个关键字段
    if fcol is None:
        continue
    for row in data_rows:
        val = ws.cell(row=row, column=fcol).value
        if val is None or str(val).strip() == "":
            seq = ws.cell(row=row, column=col_seq).value if col_seq else row
            if seq in priority:
                priority[seq]["score"] += 1
                priority[seq]["issues"].append(f"缺失：{fname}")

ranked = sorted(priority.items(), key=lambda x: x[1]["score"], reverse=True)
ranked = [(k, v) for k, v in ranked if v["score"] > 0][:20]

# ── 生成报告 ────────────────────────────────────────────────
now = datetime.now().strftime("%Y-%m-%d %H:%M")

lines = []
lines.append("# 数据核查报告 — 异常值检测")
lines.append(f"生成时间：{now}")
lines.append(f"数据来源：数据_6_数据提取表_v2.xlsx（第3行起，共{len(data_rows)}行）")
lines.append("")

lines.append("## 一、计算字段完成情况")
lines.append(f"- 训练频率：成功计算 {freq_computed} 篇，{freq_missing} 篇因数据缺失留空")
lines.append(f"- 年龄段分类：成功计算 {age_cat_young + age_cat_old} 篇，{age_cat_missing} 篇因数据缺失留空")
lines.append(f"  - young-old（≤75岁）：{age_cat_young} 篇")
lines.append(f"  - old-old（>75岁）：{age_cat_old} 篇")
lines.append("")

lines.append("## 二、异常值清单")
if anomalies:
    lines.append("| 序号 | 第一作者 | 年份 | 异常字段 | 异常值 | 正常范围 | 建议核查 |")
    lines.append("|------|---------|------|---------|-------|---------|---------|")
    for seq, author, year, field, val, lo, hi in anomalies:
        lines.append(f"| {seq} | {author} | {year} | {field} | {val} | {lo}–{hi} | 回原文确认 |")
else:
    lines.append("未发现异常值")
lines.append("")

lines.append("## 三、缺失值统计")
lines.append("| 字段 | 缺失篇数 | 缺失文献（作者+年份）|")
lines.append("|------|---------|------------------|")
for fname, cnt, detail in missing_stats:
    lines.append(f"| {fname} | {cnt} | {detail} |")
lines.append("")

lines.append("## 四、需要人工核查的优先列表")
if ranked:
    lines.append("（按问题数排序，最多20篇）")
    lines.append("")
    lines.append("| 优先级 | 序号 | 第一作者 | 年份 | 问题描述 |")
    lines.append("|--------|------|---------|------|---------|")
    for rank, (seq, info) in enumerate(ranked, 1):
        issues_str = "；".join(info["issues"])
        lines.append(f"| {rank} | {seq} | {info['author']} | {info['year']} | {issues_str} |")
else:
    lines.append("无需人工核查（无异常值且关键字段完整）")

report_text = "\n".join(lines)
with open(REPORT_PATH, "w", encoding="utf-8") as f:
    f.write(report_text)
print(f"\n报告已写入：{REPORT_PATH}")

# ── 保存 v2 ────────────────────────────────────────────────
wb.save(V2_PATH)
print(f"v2 已保存：{V2_PATH}")

print("\n=== 完成 ===")
print(f"  训练频率成功计算：{freq_computed} 篇")
print(f"  young-old：{age_cat_young} 篇 / old-old：{age_cat_old} 篇")
print(f"  异常值：{len(anomalies)} 条")
