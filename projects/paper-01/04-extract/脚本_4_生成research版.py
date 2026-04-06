"""
从 v3 复制生成 research 版和 xhs 版
- 保留表头（第1行）和颜色说明行（第2行）
- 保留序号、第一作者、年份（第1-3列，让执行者知道是哪篇）
- 清空其余所有数据字段
- 新增第56列"数据填入状态"（默认"未填入"）
- 新增第57列"数据核查状态"（默认"未核查"）
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil, os

BASE    = os.path.dirname(os.path.abspath(__file__))
V3_PATH = os.path.join(BASE, "数据_6_数据提取表_v3.xlsx")
R_PATH  = os.path.join(BASE, "数据_6_数据提取表_v3_research.xlsx")
X_PATH  = os.path.join(BASE, "数据_6_数据提取表_v3_xhs.xlsx")

C_STATUS_HDR  = "375623"   # 深绿：状态列表头字体
C_STATUS_BG   = "E2EFDA"   # 浅绿：状态列背景
C_UNFILLED    = "FFF2CC"   # 浅黄：未填入
C_UNCHECKED   = "FAFAFA"   # 近白：未核查
C_HEADER      = "2F5496"

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def make_copy(src, dst, label):
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["数据提取"]

    orig_cols = ws.max_column   # v3 是55列
    col56 = orig_cols + 1       # 56
    col57 = orig_cols + 2       # 57

    # ── 第1行：新增两列表头 ──────────────────────────
    for col_i, name in [(col56, "数据填入状态"), (col57, "数据核查状态")]:
        c = ws.cell(row=1, column=col_i, value=name)
        c.fill      = PatternFill("solid", fgColor=C_HEADER)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="微软雅黑")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(col_i)].width = 14

    # ── 第2行：说明行 ────────────────────────────────
    for col_i, label_text in [(col56, "自动"), (col57, "自动")]:
        c = ws.cell(row=2, column=col_i, value=label_text)
        c.fill      = PatternFill("solid", fgColor=C_STATUS_BG)
        c.font      = Font(size=8, color="444444", name="微软雅黑")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

    # ── 第3行起：清空数据，保留序号/作者/年份，填状态 ──
    DATA_START = 3
    for row in range(DATA_START, ws.max_row + 1):
        seq = ws.cell(row=row, column=1).value
        if seq is None:
            break

        # 清空第4列及以后的所有原有数据列
        for col_i in range(4, orig_cols + 1):
            ws.cell(row=row, column=col_i).value = None

        # 填状态列
        c56 = ws.cell(row=row, column=col56, value="未填入")
        c56.fill      = PatternFill("solid", fgColor=C_UNFILLED)
        c56.font      = Font(size=9, name="微软雅黑")
        c56.alignment = Alignment(horizontal="center", vertical="center")
        c56.border    = thin_border()

        c57 = ws.cell(row=row, column=col57, value="未核查")
        c57.fill      = PatternFill("solid", fgColor=C_UNCHECKED)
        c57.font      = Font(size=9, name="微软雅黑")
        c57.alignment = Alignment(horizontal="center", vertical="center")
        c57.border    = thin_border()

    wb.save(dst)
    print(f"✅ {label} 已生成：{dst}")
    print(f"   列数：{col57}（原{orig_cols}列 + 状态2列）")

    # 验证行数
    wb2 = openpyxl.load_workbook(dst)
    ws2 = wb2["数据提取"]
    data_rows = sum(1 for r in range(3, ws2.max_row+1)
                    if ws2.cell(r, 1).value is not None)
    print(f"   数据行数：{data_rows} 篇")

make_copy(V3_PATH, R_PATH, "research版")
make_copy(V3_PATH, X_PATH, "xhs版")

print("\n两个空白表已就绪，可以开始双盲提取。")
