import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('04-extract/数据_6_数据提取表.xlsx')
ws = wb['数据提取']

# Build seq->row mapping
seq_row = {}
for ri in range(3, ws.max_row + 1):
    seq = ws.cell(ri, 1).value
    if seq is not None:
        seq_row[int(seq)] = ri

# Extraction data based on PDF reading
# Columns: F=6 country, G=7 sample_source, H=8 N, I=9 age_mean, J=10 age_sd
# K=11 pct_female, L=12 edu_years, M=13 screen_tool, N=14 screen_score
# O=15 train_type, P=16 task_name, Q=17 adaptive, R=18 sessions, S=19 duration_min
# T=20 train_weeks, U=21 active_control, V=22 combined_type, W=23 platform
# X=24 near_transfer, Y=25 near_outcomes, Z=26 far_transfer, AA=27 far_outcomes
# AB=28 followup, AC=29 followup_months, AD=30 effect_size, AE=31 conclusion
# AF=32 neuro_imaging, AG=33 imaging_type, AH=34 neuro_finding
# AI=35 baseline_wm, AJ=36 mod_test, AK=37 mod_direction, AL=38 age_subgroup
# AM=39 cog_reserve, AN=40 pub_status, AO=41 notes

data = {}

# SEQ 29: Burki et al. 2014
data[29] = {6:'Switzerland', 7:'大学', 8:65, 9:70.8, 10:5.2, 12:13.2, 13:'未报告', 14:'未报告',
     15:'n-back', 16:'verbal n-back', 17:'是', 18:10, 19:30, 20:'2-4周',
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'spatial n-back', 26:'是', 27:'fluid intelligence (Raven)', 28:'否', 29:'未报告',
     30:'未报告', 31:'混合',
     32:'否', 35:'高/低', 36:'是', 37:'低>高', 38:'是', 39:'教育年限', 40:'期刊',
     41:'潜增长曲线模型分析训练曲线；含年轻与老年成人比较'}

# SEQ 30: Cantarella et al. 2017
data[30] = {6:'Italy', 7:'社区', 8:36, 9:69.5, 10:3.0, 12:11.0, 13:'未报告', 14:'未报告',
     15:'span', 16:'verbal WM span (CWMS)', 17:'是', 18:6, 19:30, 20:3,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'WM criterion task (CWMS)', 26:'是', 27:'Everyday Problem Test; TIADL; Cattell; Raven RSPM',
     28:'否', 29:'未报告', 30:'d>=0.8 (large effect)', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'评估日常生活能力迁移效应'}

# SEQ 31: Carretti et al. 2013
data[31] = {6:'Italy', 7:'社区', 8:37, 9:69.0, 10:3.5, 11:57, 12:10.5, 13:'未报告', 14:'未报告',
     15:'span', 16:'verbal WM span (CWMS)', 17:'是', 18:6, 19:30, 20:3,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'WM updating task', 26:'是', 27:'Language comprehension; Cattell fluid intelligence',
     28:'是', 29:'6', 30:'η²=0.15-0.26 (transfer)', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'6个月随访；语言理解力迁移'}

# SEQ 32: Park et al. 2014 (Korea)
data[32] = {6:'Korea', 7:'医院', 8:40, 9:69.7, 10:4.5, 11:68, 13:'MMSE', 14:'未报告',
     15:'other', 16:'CACT (computer-assisted cognitive training)', 17:'否', 18:10, 20:2,
     21:'否', 22:'tDCS', 23:'Maxmedica CACT',
     24:'是', 25:'Verbal WM task; Digit span forward', 26:'否',
     28:'是', 29:'1', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'tDCS联合CACT；anodal vs sham tDCS'}

# SEQ 34: Assecondi et al. 2022 (UK)
data[34] = {6:'UK', 7:'社区', 8:28, 9:69.5, 10:5.5, 13:'MoCA', 14:'未报告',
     15:'n-back', 16:'adaptive spatial n-back', 17:'是', 18:5, 19:20, 20:1,
     21:'否', 22:'tDCS', 23:'自制',
     24:'是', 25:'WM composite score (5 tasks)', 26:'否',
     28:'否', 29:'未报告', 30:'未报告', 31:'混合',
     32:'否', 35:'高/低', 36:'是', 37:'低>高', 38:'是', 39:'未报告', 40:'期刊',
     41:'低WM capacity者从tDCS中获益更多；YO vs OO分组分析'}

# SEQ 35: Gunther et al. 2003 (Austria)
data[35] = {6:'Austria', 7:'医院', 8:19, 9:83.5, 10:4.0, 11:79, 13:'未报告', 14:'未报告',
     15:'other', 16:'computer-assisted cognitive training (multi-domain)', 17:'否', 20:14,
     21:'否', 22:'无', 23:'自制',
     24:'是', 25:'Primary WM; secondary WM (verbal/visual)', 26:'是',
     27:'Information processing speed; learning; interference tendency',
     28:'是', 29:'5', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'单组前后测；养老院居民；老老年（75-91岁）'}

# SEQ 36: Shatil et al. 2014
data[36] = {6:'Czech Republic', 7:'社区', 8:119, 9:70.5, 10:7.0, 13:'未报告', 14:'未报告',
     15:'mixed', 16:'CogniFit iTV (21 cognitive tasks incl. WM)', 17:'是', 18:25,
     21:'是', 22:'无', 23:'CogniFit',
     24:'是', 25:'WM tasks (validated)', 26:'是', 27:'Executive function tasks',
     28:'否', 29:'未报告', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'互动电视（iTV）平台认知训练'}

# SEQ 39: Buschkuehl et al. 2008
data[39] = {6:'Switzerland', 7:'社区', 8:80, 9:80.0, 10:4.0, 13:'未报告', 14:'未报告',
     15:'other', 16:'WM process-based training (visual/spatial)', 17:'是', 18:24, 19:45, 20:12,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'Visual WM (VLMT)', 26:'是', 27:'Visual episodic memory',
     28:'是', 29:'12', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'老老年（~80岁）；1年随访；身体锻炼主动对照组'}

# SEQ 40: Booth et al. 2023
data[40] = {6:'UK', 7:'社区', 8:71, 9:66.0, 10:5.5, 13:'未报告', 14:'未报告',
     15:'n-back', 16:'adaptive verbal + spatial n-back', 17:'是', 18:16, 20:'4-8周',
     21:'是', 22:'无', 23:'自制（居家在线）',
     24:'是', 25:'Digit span', 26:'是', 27:'Abstract relational reasoning',
     28:'否', 29:'未报告', 30:'未报告', 31:'无迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'训练时间表（密集vs分散）；居家无监督在线训练'}

# SEQ 41: Heinzel et al. 2017
data[41] = {6:'Germany', 7:'社区', 8:40, 9:66.0, 10:4.0, 13:'未报告', 14:'未报告',
     15:'n-back', 16:'adaptive n-back', 17:'是', 18:12, 20:4,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'multimodal dual-task (visual+auditory)', 26:'否',
     28:'否', 29:'未报告', 30:'p<0.05', 31:'正向迁移',
     32:'是', 33:'fMRI', 34:'DLPFC活动变化预测双任务表现；左侧DLPFC→听觉双任务成本',
     35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'双任务迁移；fMRI子研究；先导研究'}

# SEQ 43: Stephens et al. 2017 (COMT)
data[43] = {6:'USA', 7:'社区', 8:137, 13:'未报告', 14:'未报告',
     15:'other', 16:'Visual and Spatial WM training', 17:'是',
     21:'否', 22:'tDCS', 23:'自制',
     24:'是', 25:'Visual WM; Spatial WM', 26:'否',
     28:'是', 29:'1', 30:'未报告', 31:'混合',
     32:'否', 35:'高/低', 36:'是', 37:'混合', 38:'否', 39:'未报告', 40:'期刊',
     41:'COMT val158met基因型调节tDCS-WM训练效益；不同tDCS强度（1/1.5/2mA）'}

# SEQ 44: Sutton et al. 2025
data[44] = {6:'UK', 7:'社区', 8:103, 13:'未报告', 14:'未报告',
     15:'mixed', 16:'commercial brain training app (multi-domain)', 17:'是', 20:12,
     21:'是', 22:'无', 23:'商业脑训练app',
     24:'是', 25:'WM tasks (trained domain)', 26:'是',
     27:'Processing speed; attention; language',
     28:'否', 29:'未报告', 30:'未报告', 31:'无迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'商业脑训练app；仅练习效应，无迁移；3个月干预'}

# SEQ 45: Tusch et al. 2016 (Cogmed, ERP)
data[45] = {6:'USA', 7:'社区', 8:35, 13:'未报告', 14:'未报告',
     15:'other', 16:'Cogmed (adaptive WM training)', 17:'是', 18:25, 19:40, 20:5,
     21:'是', 22:'无', 23:'Cogmed',
     24:'是', 25:'n-back (0-back, 1-back, 2-back)', 26:'否',
     28:'否', 29:'未报告', 30:'未报告', 31:'混合',
     32:'是', 33:'ERP', 34:'P3a和P3b波幅增大（自适应训练组）；P3波幅与任务表现正相关',
     35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'Cogmed自适应vs非自适应；ERP神经指标变化'}

# SEQ 46: McAvinue et al. 2013
data[46] = {6:'Ireland', 7:'社区', 8:36, 9:70.0, 10:4.0, 13:'未报告', 14:'未报告',
     15:'span', 16:'Baddeley WM model-based (auditory+visuospatial span)', 17:'是', 20:5,
     21:'是', 22:'无', 23:'自制（在线）',
     24:'是', 25:'Auditory STM span; visuospatial STM', 26:'是',
     27:'Long-term episodic memory',
     28:'是', 29:'3;6', 30:'未报告', 31:'混合',
     32:'否', 35:'未报告', 36:'是', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'心理压力与训练时间交互关系；auditory STM span扩展但WM容量无提升'}

# SEQ 47: Goghari & Lawlor-Savage 2018 (self-perceived)
data[47] = {6:'Canada', 7:'社区', 8:97, 13:'未报告', 14:'未报告',
     15:'other', 16:'computerized WM training (web-based)', 17:'是', 20:8,
     21:'是', 22:'无', 23:'商业软件',
     24:'是', 25:'Self-perceived cognitive failures (CFQ)', 26:'否',
     28:'否', 29:'未报告', 30:'未报告', 31:'混合',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'主观感知认知获益；WM vs 逻辑规划训练比较；自我报告结局'}

# SEQ 48: Borella, Carbone, Spironelli 2025
data[48] = {6:'Italy', 7:'社区', 8:30, 9:69.0, 10:3.5, 13:'未报告', 14:'未报告',
     15:'span', 16:'verbal WM span (CWMS)', 17:'是', 18:6, 19:30, 20:3,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'n-back task', 26:'否',
     28:'是', 29:'6', 30:"Cohen's d medium-large", 31:'正向迁移',
     32:'是', 33:'ERP', 34:'训练后及6个月随访时TG出现左侧化ERP激活；近迁移仅ERP层面显著',
     35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'先导研究；ERP神经相关；6个月随访'}

# SEQ 49: Goghari & Lawlor-Savage 2017 (Bayesian)
data[49] = {6:'Canada', 7:'社区', 8:97, 13:'未报告', 14:'未报告',
     15:'other', 16:'web-based WM or logic & planning training', 17:'是', 20:8,
     21:'是', 22:'无', 23:'商业软件',
     24:'是', 25:'WM tasks (near transfer)', 26:'是',
     27:'Planning; reasoning; processing speed; verbal fluency',
     28:'否', 29:'未报告', 30:'未报告 (Bayesian)', 31:'无迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'Bayesian分析；WM vs 逻辑规划训练；认知天花板效应可能'}

# SEQ 50: Zinke et al. 2012 (old-old)
data[50] = {6:'Germany', 7:'社区', 8:40, 9:86.8, 10:4.0, 13:'未报告', 14:'未报告',
     15:'span', 16:'WM span tasks (5 tasks)', 17:'否', 18:10,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'trained WM tasks', 26:'是', 27:'Executive functions (2 tests)',
     28:'否', 29:'未报告', 30:'d=0.5-0.8 (trained tasks)', 31:'无迁移',
     32:'否', 35:'高/低', 36:'是', 37:'低>高', 38:'否', 39:'未报告', 40:'期刊',
     41:'老老年（~87岁）；低基线WM容量者获益更多；无迁移效应'}

# SEQ 51: Basak & O'Connell 2016
data[51] = {6:'USA', 7:'社区', 8:60, 9:66.0, 10:5.0, 13:'未报告', 14:'未报告',
     15:'other', 16:'memory updating task (predictable vs unpredictable)', 17:'是',
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'WM updating task (novel)', 26:'是', 27:'Episodic memory',
     28:'是', 29:'1.5', 30:'未报告', 31:'混合',
     32:'否', 35:'未报告', 36:'是', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'不可预测vs可预测记忆更新训练；个体学习率预测迁移'}

# SEQ 52: Jaeggi et al. 2023 (EngAge)
data[52] = {6:'USA', 7:'社区', 8:119, 13:'未报告', 14:'未报告',
     15:'n-back', 16:'n-back + metacognitive program (EngAge)', 17:'是', 18:20,
     21:'是', 22:'无', 23:'自制（居家）',
     24:'是', 25:'non-trained WM measures', 26:'是',
     27:'Inhibitory control; episodic memory',
     28:'是', 29:'未报告', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'WM训练+EngAge元认知干预；元认知附加效应不显著；可行性研究'}

# SEQ 53: Borella et al. 2017 (individual characteristics, 4 pooled studies)
data[53] = {6:'Italy', 7:'社区', 8:150, 9:69.0, 10:4.0, 12:11.0, 13:'未报告', 14:'未报告',
     15:'span', 16:'verbal WM span (CWMS, Borella 2010 procedure)', 17:'是', 18:6, 19:30, 20:3,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'visuospatial WM; STM tasks', 26:'是',
     27:'fluid intelligence; processing speed; inhibitory measures',
     28:'是', 29:'未报告', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'高/低', 36:'是', 37:'低>高', 38:'是', 39:'教育年限', 40:'期刊',
     41:'4项研究合并分析；年龄、教育、词汇量、基线WM预测训练效益'}

# SEQ 54: Brambilla et al. 2021 (tRNS)
data[54] = {6:'UK', 7:'社区', 8:42, 9:70.0, 10:6.5, 13:'未报告', 14:'未报告',
     15:'other', 16:'executive function training (cognitive flexibility+inhibitory control+WM)', 17:'是', 18:5, 19:30, 20:1,
     21:'否', 22:'TMS', 23:'自制',
     24:'是', 25:'non-verbal logical reasoning', 26:'是',
     27:'Attention; memory; executive functions',
     28:'是', 29:'1', 30:'未报告', 31:'混合',
     32:'否', 35:'未报告', 36:'是', 37:'未检验', 38:'是', 39:'未报告', 40:'期刊',
     41:'tRNS联合执行功能训练；年龄与逻辑推理改善相关（1mA组）'}

# SEQ 55: Antonenko et al. 2022 (tDCS + letter updating)
data[55] = {6:'Germany', 7:'社区', 8:56, 9:70.0, 10:4.0, 13:'未报告', 14:'未报告',
     15:'other', 16:'letter updating task', 17:'是', 18:9, 19:20, 20:3,
     21:'否', 22:'tDCS', 23:'自制',
     24:'是', 25:'n-back (near transfer)', 26:'是',
     27:'executive/memory tasks',
     28:'是', 29:'1;7', 30:'未报告', 31:'混合',
     32:'否', 35:'未报告', 36:'是', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'tDCS-assisted letter updating；n-back近迁移显著'}

# SEQ 57: Zelinski et al. 2014 (IMPACT, SEM)
data[57] = {6:'USA', 7:'社区', 8:487, 9:74.0, 10:6.0, 13:'未报告', 14:'未报告',
     15:'span', 16:'syllable span + auditory discrimination (IMPACT program)', 17:'是',
     21:'是', 22:'无', 23:'Posit Science (IMPACT)',
     24:'是', 25:'WM factor score', 26:'是', 27:'List memory; text memory',
     28:'否', 29:'未报告', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'是', 37:'未检验', 38:'是', 39:'教育年限', 40:'期刊',
     41:'结构方程模型分析训练-迁移增益关系；IMPACT研究二次分析；N=487'}

# SEQ 59: Spironelli & Borella 2021 (resting EEG)
data[59] = {6:'Italy', 7:'社区', 8:24, 9:69.0, 10:3.5, 13:'未报告', 14:'未报告',
     15:'span', 16:'verbal WM span (CWMS)', 17:'是', 18:6, 19:30, 20:3,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'CWMS criterion; inhibitory control (intrusion errors)', 26:'否',
     28:'是', 29:'6', 30:'d=0.5-0.9', 31:'正向迁移',
     32:'是', 33:'EEG', 34:'训练后TG前额left ROI高b/a比率增大；与WM表现正相关；随访持续',
     35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'静息态EEG；额叶振荡活动；先导研究'}

# SEQ 60: Nguyen et al. 2026 (Australia, single vs multi-domain)
data[60] = {6:'Australia', 7:'社区', 8:66, 9:69.58, 10:7.04, 13:'未报告', 14:'未报告',
     15:'mixed', 16:'single-domain WM (n-back) or multidomain EF training', 17:'是', 20:4,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'n-back; global EF accuracy', 26:'是',
     27:'Prospective memory (Virtual Week); fluid intelligence',
     28:'否', 29:'未报告', 30:'未报告 (Bayesian)', 31:'混合',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'单域WM vs 多域EF训练比较；Bayesian分析；主观认知自评提升'}

# SEQ 61: Cantarella et al. 2021 (visuospatial, emotional stimuli)
data[61] = {6:'Italy', 7:'社区', 8:40, 9:69.0, 10:4.0, 13:'未报告', 14:'未报告',
     15:'span', 16:'visuospatial WM training (neutral vs emotional stimuli)', 17:'是', 18:6, 19:30, 20:3,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'visuospatial WM criterion task', 26:'是',
     27:'verbal WM; visuospatial STM; reasoning',
     28:'否', 29:'未报告', 30:'未报告', 31:'混合',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'训练刺激材料情绪效价对迁移效应的影响（中性vs正性图片）'}

# SEQ 62: Pergher et al. 2020 (Belgium, strategy use, EEG)
data[62] = {6:'Belgium', 7:'社区', 8:26, 13:'未报告', 14:'未报告',
     15:'n-back', 16:'n-back (strategy vs no strategy)', 17:'是',
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'n-back variant', 26:'是', 27:'P300 transfer measures',
     28:'否', 29:'未报告', 30:'未报告', 31:'混合',
     32:'是', 33:'EEG', 34:'P300变化；策略使用影响训练组间差异；个体特征影响P300',
     35:'未报告', 36:'是', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'N-back训练中策略使用的影响；EEG P300；个体差异调节'}

# SEQ 63: Lange & Suss 2015 (Germany, multicomponent)
data[63] = {6:'Germany', 7:'社区', 8:91, 13:'未报告', 14:'未报告',
     15:'mixed', 16:'adaptive multicomponent WM training (phonological+visuospatial+CE)', 17:'是',
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'WM tasks (near transfer)', 26:'是',
     27:'Short-term memory; processing speed; reasoning',
     28:'否', 29:'未报告', 30:'未报告', 31:'无迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'主动对照组与训练组在近迁移上表现相似；日常生活评估（ambulatory assessment）'}

# SEQ 67: Chai & Gao et al. 2026 (China, older women, cognitive reappraisal)
data[67] = {6:'China', 7:'社区', 8:69, 9:70.0, 10:5.0, 11:100, 13:'未报告', 14:'未报告',
     15:'n-back', 16:'adaptive WM updating (numerical updating + 1-back)', 17:'是', 18:20,
     21:'是', 22:'无', 23:'自制',
     24:'是', 25:'WM updating task', 26:'是',
     27:'Cognitive reappraisal (emotion regulation task)',
     28:'否', 29:'未报告', 30:'未报告', 31:'正向迁移',
     32:'否', 35:'未报告', 36:'否', 37:'未检验', 38:'否', 39:'未报告', 40:'期刊',
     41:'仅老年女性样本；认知重评情绪调节能力作为远迁移结局'}

# Write data to Excel
filled = 0
for seq, fields in data.items():
    if seq not in seq_row:
        print("WARNING: seq %d not found" % seq)
        continue
    row = seq_row[seq]
    for col, val in fields.items():
        ws.cell(row, col).value = val
    filled += 1

wb.save('04-extract/数据_6_数据提取表.xlsx')
print("Saved. Filled %d records." % filled)
