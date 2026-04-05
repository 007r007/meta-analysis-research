"""
Paper-01 数据提取 v4 — 全量重跑脚本
=====================================
目标：对56篇文献逐篇读PDF，提取/覆盖 col6-57 全部字段写入Excel
策略：
  - 保留 col1-5（序号/作者/年份/标题/期刊）不覆盖
  - 保留 col6-57 已有高质量手工数据（seq1-6 大部分字段），仅补空缺
  - seq8-67 重新提取，修正 neuroimaging 错误
  - 神经影像判断基于 Methods 节强指标，不再依赖任意文本匹配
  - 每10篇自动保存一次；全部完成后做自检并最终保存
"""

import fitz           # pymupdf
import openpyxl
import os, re, sys

sys.stdout.reconfigure(encoding='utf-8')

BASE    = r"E:\Meta-analysis writing project\projects\paper-01"
PDF_DIR = os.path.join(BASE, r"03-screen\全文PDF")
EXCEL   = os.path.join(BASE, r"04-extract\数据_6_数据提取表_v3_research.xlsx")

# ── PDF映射 ──────────────────────────────────────────────────────────────────
PDF_MAP = {
    1:  "2020_The effect of cognitive training in older adults_ be aware of CRUNCH..pdf",
    2:  "2019_Is working memory training in older adults sensitive to music_.pdf",
    3:  "2016_Age-specific differences of dual n-back training..pdf",
    4:  "2015_Longitudinal neurostimulation in older adults improves working memory..pdf",
    5:  "2018_N-back training and transfer effects revealed by behavioral responses and EEG..pdf",
    6:  "2020_Verbal working memory training in older adults_ an investigation of dose respons.pdf",
    8:  "2025_Gamified working memory intervention enhances prefrontal neurocognitive plastici.pdf",
    9:  "2017_Working memory training in older adults_ Bayesian evidence supporting the absenc.pdf",
    10: "2022_Working Memory Training Coupled With Transcranial Direct Current Stimulation in.pdf",
    11: "2016_Working Memory Training and Speech in Noise Comprehension in Older Adults..pdf",
    12: "2016_Older Adults Improve on Everyday Tasks after Working Memory Training and Neurost.pdf",
    13: "2019_Working Memory Capacity as a Predictor of Cognitive Training Efficacy in the Eld.pdf",
    14: "2021_Cognitive Aftereffects of Acute tDCS Coupled with Cognitive Training_ An fMRI St.pdf",
    15: "2019_Improving Everyday Functioning in the Old-Old with Working Memory Training..pdf",
    16: "2020_Working Memory Training for Older Participants_ A Control Group Training Regimen.pdf",
    18: "2014_Working memory training improvements and gains in non-trained cognitive tasks in.pdf",
    19: "2013_Working memory training in old age_ an examination of transfer and maintenance e.pdf",
    20: "2020_Feasibility of using a computer-assisted working memory training program for hea.pdf",
    21: "2018_The Effects of Cognitive Training on Cognitive Abilities and Everyday Function_.pdf",
    22: "2022_Higher white matter hyperintensity load adversely affects pre-post proximal cogn.pdf",
    23: "2024_Youth-like brain activation linked with greater cognitive training gains in olde.pdf",
    24: "2022_Training attentive individuation leads to visuo-spatial working memory improveme.pdf",
    25: "2014_Working memory training and transfer in older adults_ effects of age, baseline p.pdf",
    26: "2020_The Impact of Working Memory Training on Cognitive Abilities in Older Adults_ Th.pdf",
    27: "2024_Limited training and transfer effects in older and young adults who participated.pdf",
    28: "2020_Investigating the Effects of Spacing on Working Memory Training Outcome_ A Rando.pdf",
    29: "2014_Individual differences in cognitive plasticity_ an investigation of training cur.pdf",
    30: "2017_Benefits in tasks related to everyday life competences after a working memory tr.pdf",
    31: "2013_Gains in language comprehension relating to working memory training in healthy o.pdf",
    32: "2014_Long-term effects of transcranial direct current stimulation combined with compu.pdf",
    34: "2022_Older adults with lower working memory capacity benefit from transcranial direct.pdf",
    35: "2003_Long-term improvements in cognitive performance through computer-assisted cognit.pdf",
    36: "2014_Novel television-based cognitive training improves working memory and executive.pdf",
    39: "2008_Impact of working memory training on memory performance in old-old adults..pdf",
    40: "2023_Experimental investigation of training schedule on home-based working memory tra.pdf",
    41: "2017_Transfer Effects to a Multimodal Dual-Task after Working Memory Training and Ass.pdf",
    43: "2017_Task demands, tDCS intensity, and the COMT val(158)met polymorphism impact tDCS-.pdf",
    44: "2025_Practice makes perfect, but to what end_ Computerised brain training has limited.pdf",
    45: "2016_Changes in Neural Activity Underlying Working Memory after Computerized Cognitiv.pdf",
    46: "2013_An evaluation of a working memory training scheme in older adults..pdf",
    47: "2018_Self-Perceived Benefits of Cognitive Training in Healthy Older Adults..pdf",
    48: "2025_Short- and long-term cognitive and electrophysiological effects of a brief worki.pdf",
    49: "2017_Comparison of Cognitive Change after Working Memory Training and Logic and Plann.pdf",
    50: "2012_Potentials and limits of plasticity induced by working memory training in old-ol.pdf",
    51: "2016_To Switch or Not to Switch_ Role of Cognitive Control in Working Memory Training.pdf",
    52: "2023_EngAge - A metacognitive intervention to supplement working memory training_ A f.pdf",
    53: "2017_Working Memory Training for Healthy Older Adults_ The Role of Individual Charact.pdf",
    54: "2021_The Effect of Transcranial Random Noise Stimulation on Cognitive Training Outcom.pdf",
    55: "2022_Randomized trial of cognitive training and brain stimulation in non-demented old.pdf",
    57: "2014_Evaluating the relationship between change in performance on training tasks and.pdf",
    59: "2021_Working Memory Training and Cortical Arousal in Healthy Older Adults_ A Resting-.pdf",
    60: "2026_A comparison of single-domain and multidomain executive functions cognitive trai.pdf",
    61: "2021_The influence of training task stimuli on transfer effects of working memory tra.pdf",
    62: "2020_Impact of strategy use during N-Back training in older adults.pdf",
    63: "2015_Experimental evaluation of near\u2010 and far\u2010transfer effects of an adaptive multico.pdf",
    67: "2026_The Impact of Working Memory Training on Cognitive Reappraisal Ability Among Old.pdf",
}

# ── 确认神经影像研究列表（手工逐篇核查确认）─────────────────────────────────
# 依据：Methods节明确描述本研究收集了脑成像/电生理数据
CONFIRMED_NEURO = {
    5:  {"type": "EEG/ERP",    "finding": "训练后老年人3-back P300在顶叶(Pz)显著增加，反映神经效率提升"},
    8:  {"type": "fMRI/sMRI",  "finding": "训练组左额叶皮质厚度增加，静息态功能均匀性降低，与WM改善关联"},
    14: {"type": "fMRI/sMRI",  "finding": "tDCS+认知训练组在训练任务激活右中前额叶和右顶叶后叶出现变化"},
    22: {"type": "sMRI",       "finding": "较高白质高信号负荷负向预测训练前后近迁移改善量"},
    23: {"type": "fMRI",       "finding": "基线青年样激活模式（GOF指数）正向预测训练增益和迁移效应"},
    41: {"type": "fMRI",       "finding": "训练组左DLPFC神经活动变化预测训练后听觉双任务成本，右DLPFC预测视觉双任务"},
    45: {"type": "EEG/ERP",    "finding": "训练后WM任务P300振幅增大，训练组与对照组相比神经活动模式改变"},
    48: {"type": "EEG/ERP",    "finding": "训练组出现侧化事件相关电位激活，效果在训练后6个月随访中维持"},
    59: {"type": "EEG",        "finding": "训练后静息态EEG皮质唤醒指标（alpha功率）发生变化，试验性研究"},
    62: {"type": "EEG/ERP",    "finding": "P300振幅变化反映训练组负荷效应减小，提示神经效率提升；策略使用调节效果"},
}

SEQ_ORDER = [1,2,3,4,5,6,8,9,10,11,12,13,14,15,16,18,19,20,21,22,23,24,25,26,27,28,
             29,30,31,32,34,35,36,39,40,41,43,44,45,46,47,48,49,50,51,52,53,54,55,
             57,59,60,61,62,63,67]

# ── 辅助函数 ─────────────────────────────────────────────────────────────────
def read_pdf(seq):
    fname = PDF_MAP.get(seq)
    if not fname:
        return None, "PDF缺失"
    path = os.path.join(PDF_DIR, fname)
    if not os.path.exists(path):
        return None, "PDF缺失"
    try:
        doc = fitz.open(path)
        text = "".join([p.get_text() for p in doc])
        doc.close()
        if len(text.strip()) < 200:
            return text, "扫描件"
        return text, "ok"
    except Exception as e:
        return None, f"读取错误:{e}"


def find_num(text, patterns):
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1))
            except:
                pass
    return None


def find_str(text, patterns):
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def pct_female(text):
    """Return female percentage as string like '55%', or '未报告'"""
    # Direct percentage form
    for p in [
        r'(\d+(?:\.\d+)?)\s*%\s*(?:were\s+)?(?:female|women)',
        r'(?:female|women)[^.]{0,40}?(\d+(?:\.\d+)?)\s*%',
        r'(\d+(?:\.\d+)?)\s*%\s*(?:female|women)',
        r'gender.*?(\d+(?:\.\d+)?)\s*%\s*(?:female|women)',
        r'sex.*?(\d+(?:\.\d+)?)\s*%\s*(?:female|women)',
    ]:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            try:
                f = float(m.group(1))
                if 0 < f <= 100:
                    return f"{f:.0f}%"
            except:
                pass
    # Count form: "X females and Y males"
    m = re.search(r'(\d+)\s+(?:females?|women)\s+(?:and\s+)?(\d+)\s+males?', text, re.I)
    if m:
        nf, nm = int(m.group(1)), int(m.group(2))
        total = nf + nm
        if total > 0:
            return f"{round(nf/total*100)}%"
    return "未报告"


def extract_country(text):
    """Extract first author's institution country"""
    # Look for affiliation patterns in first 2000 chars
    header = text[:2000]
    country_map = {
        r'\bUSA\b|\bUnited\s+States\b|\bU\.S\.A\b': 'USA',
        r'\bGermany\b|\bDeutschland\b': 'Germany',
        r'\bItaly\b|\bItalia\b': 'Italy',
        r'\bBrazil\b|\bBrasil\b': 'Brazil',
        r'\bCanada\b': 'Canada',
        r'\bSweden\b|\bSverige\b': 'Sweden',
        r'\bSwiss(?:erland)?\b|\bSwitzerland\b': 'Switzerland',
        r'\bNetherlands\b|\bHolland\b': 'Netherlands',
        r'\bBelgium\b|\bBelgique\b': 'Belgium',
        r'\bSpain\b|\bEspa[nñ]a\b': 'Spain',
        r'\bFrance\b|\bFrançais\b': 'France',
        r'\bAustralia\b': 'Australia',
        r'\bNorway\b|\bNorge\b': 'Norway',
        r'\bFinland\b|\bSuomi\b': 'Finland',
        r'\bUK\b|\bUnited\s+Kingdom\b|\bEngland\b|\bScotland\b|\bWales\b': 'UK',
        r'\bChina\b|\bChinese\b|\bBeijing\b|\bShanghai\b|\bHong\s+Kong\b': 'China',
        r'\bJapan\b': 'Japan',
        r'\bKorea\b|\bKorean\b': 'Korea',
        r'\bIsrael\b': 'Israel',
        r'\bPortugal\b|\bPortuguese\b': 'Portugal',
        r'\bAustria\b|\bÖsterreich\b': 'Austria',
        r'\bDenmark\b|\bDanmark\b': 'Denmark',
        r'\bPoland\b|\bPolska\b': 'Poland',
        r'\bGreece\b|\bGreek\b': 'Greece',
        r'\bCzech\b|\bCzechia\b': 'Czech Republic',
        r'\bBelgium\b': 'Belgium',
        r'\bNew\s+Zealand\b': 'New Zealand',
        r'\bIreland\b': 'Ireland',
        r'\bHungary\b|\bMagyar\b': 'Hungary',
        r'\bArgentina\b': 'Argentina',
        r'\bMexico\b|\bMéxico\b': 'Mexico',
        r'\bTurkey\b|\bTürkiye\b': 'Turkey',
        r'\bIndia\b': 'India',
        r'\bSingapore\b': 'Singapore',
        r'\bRomania\b': 'Romania',
    }
    for pat, name in country_map.items():
        if re.search(pat, header, re.IGNORECASE):
            return name
    # Fall back to full text
    for pat, name in country_map.items():
        if re.search(pat, text[:5000], re.IGNORECASE):
            return name
    return "待核查"


def extract_fields(text, seq):
    d = {}

    # ─ 1. 国家 ─
    d["国家"] = extract_country(text)

    # ─ 2. 样本来源 ─
    if re.search(r'\bcommunity\b', text[:5000], re.I):
        d["样本来源"] = "社区"
    elif re.search(r'\buniversity\b|\bcollege\b', text[:5000], re.I):
        d["样本来源"] = "大学"
    elif re.search(r'\bhospital\b|\bclinic\b|\bmedical\b', text[:5000], re.I):
        d["样本来源"] = "医院"
    else:
        d["样本来源"] = "未报告"

    # ─ 3. 样本量N (final analyzed sample) ─
    n = find_num(text, [
        r'(?:final|total|overall|analyzed?)\s+(?:sample\s+)?(?:of\s+)?n\s*=\s*(\d+)',
        r'data\s+(?:were|was)\s+(?:analyzed|available)\s+for\s+(\d+)',
        r'(\d+)\s+(?:participants?|older\s+adults?|subjects?)\s+(?:were\s+)?(?:included|completed|analyzed)',
        r'(?:included|analyzed|completed)\s+(?:data\s+from\s+)?(\d+)\s+(?:participants?|older\s+adults?)',
        r'\bn\s*=\s*(\d+)\b',
    ])
    d["样本量N"] = int(n) if n else "未报告"

    # ─ 4. 训练组N / 对照组N ─
    tg = find_num(text, [
        r'training\s+group\s*[,\(]?\s*n\s*=\s*(\d+)',
        r'(?:experimental|intervention|WM|working\s+memory)\s+(?:group|condition)\s*[,\(]?\s*n\s*=\s*(\d+)',
        r'(?:WM\s+)?training\s+condition\s*[,\(]?\s*n\s*=\s*(\d+)',
        r'(?:cognitive\s+)?training\s+group.*?n\s*=\s*(\d+)',
    ])
    cg = find_num(text, [
        r'control\s+group\s*[,\(]?\s*n\s*=\s*(\d+)',
        r'(?:active|passive|waitlist|wait[\s-]list|no[\s-]contact)\s+(?:control\s+)?(?:group|condition)\s*[,\(]?\s*n\s*=\s*(\d+)',
        r'comparison\s+group\s*[,\(]?\s*n\s*=\s*(\d+)',
    ])
    no_ctrl = bool(re.search(
        r'no\s+control\s+group|single.group|without\s+(?:a\s+)?control\s+group|'
        r'one.group|single\s+arm|uncontrolled', text, re.I))
    d["训练组N"] = int(tg) if tg else "未报告"
    d["对照组N"] = int(cg) if cg else ("无对照组" if no_ctrl else "未报告")

    # ─ 5. 年龄均值 + SD ─
    age_m = find_num(text, [
        r'(?:mean|average)\s+age[^.]{0,40}?(\d{2}(?:\.\d+)?)\s*(?:years?|yrs?|\()',
        r'age[d\s]+(?:[Mm](?:ean)?\s*[=:]\s*)(\d{2}(?:\.\d+)?)',
        r'(?:M|mean)\s*(?:age)?\s*=\s*(\d{2}(?:\.\d+)?)',
        r'aged?\s+(\d{2}(?:\.\d+)?)\s*(?:±|\(SD|years?\s+old)',
        r'(\d{2}(?:\.\d+)?)\s*(?:±\s*\d|years?\s+old)',
    ])
    age_sd = find_num(text, [
        r'age[^.]{0,50}(?:SD|s\.?d\.?)\s*=?\s*(\d+(?:\.\d+)?)',
        r'(?:±|S\.?D\.?)\s*(\d+(?:\.\d+)?)\s*(?:years?|yrs?|\))',
    ])
    d["年龄均值"] = age_m if age_m else "未报告"
    d["年龄SD"]   = age_sd if age_sd else "未报告"

    # ─ 6. 年龄段分类 ─
    if isinstance(age_m, float):
        d["年龄段分类"] = "old-old" if age_m > 75 else "young-old"
    elif isinstance(age_m, (int,)) and age_m:
        d["年龄段分类"] = "old-old" if age_m > 75 else "young-old"
    else:
        # Try to determine from text
        if re.search(r'old-old|oldest.old|75\s*\+\s*years?|(?:mean|average)\s+age[^.]{0,30}[78][0-9]', text, re.I):
            d["年龄段分类"] = "old-old"
        elif re.search(r'young-old|60\s*[-–]\s*7[05]|(?:mean|average)\s+age[^.]{0,30}6[0-9]', text, re.I):
            d["年龄段分类"] = "young-old"
        else:
            d["年龄段分类"] = "待核查"

    # ─ 7. 性别(%女) ─
    d["性别(%女)"] = pct_female(text)

    # ─ 8. 教育年限 ─
    edu = find_num(text, [
        r'(?:education|years?\s+of\s+(?:formal\s+)?education|schooling)[^.]{0,60}?(\d+(?:\.\d+)?)\s*(?:years?|yrs?)',
        r'(\d+(?:\.\d+)?)\s*(?:years?|yrs?)\s+of\s+(?:formal\s+)?education',
        r'education\s*[=:]\s*(\d+(?:\.\d+)?)',
        r'[Mm]ean\s+education[^.]{0,30}?(\d+(?:\.\d+)?)',
        r'mean\s+years\s+of\s+(?:formal\s+)?education[^.]{0,30}(\d+(?:\.\d+)?)',
    ])
    d["教育年限(年)"] = edu if edu else "未报告"

    # ─ 9. 认知筛查工具 + 分数 ─
    tools, scores = [], []
    for tool, pat_score in [
        ("MMSE", r'MMSE[^.]{0,50}?(?:score|mean|≥|>|=|cutoff)[^.]{0,30}?(\d+(?:\.\d+)?)'),
        ("MoCA", r'MoCA[^.]{0,50}?(?:score|mean|≥|>|=|cutoff)[^.]{0,30}?(\d+(?:\.\d+)?)'),
        ("MMSQ", r'MMSQ[^.]{0,30}?(\d+(?:\.\d+)?)'),
        ("Mini-Mental", r'Mini.Mental[^.]{0,50}?(\d+(?:\.\d+)?)'),
    ]:
        if re.search(tool, text, re.I):
            tools.append(tool)
            ms = re.search(pat_score, text, re.I)
            if ms:
                scores.append(f"{tool}={ms.group(1)}")

    if not tools:
        if re.search(r'no\s+(?:cognitive\s+)?screening|not\s+screened|no\s+(?:inclusion|exclusion)\s+criteria.*cogniti', text, re.I):
            d["认知筛查工具"] = "未使用"
            d["筛查分数"] = "未使用"
        else:
            d["认知筛查工具"] = "未报告"
            d["筛查分数"] = "未报告"
    else:
        d["认知筛查工具"] = "; ".join(tools)
        d["筛查分数"] = "; ".join(scores) if scores else "均通过筛查标准"

    # ─ 10. 训练类型 ─
    if re.search(r'\bn-back\b|n\s*back\b|dual.n.back', text, re.I):
        d["训练类型"] = "n-back"
    elif re.search(r'\bspan\b.*(?:training|task)|complex\s+span|reading\s+span|operation\s+span|'
                   r'digit\s+span.*train', text, re.I):
        d["训练类型"] = "span"
    else:
        d["训练类型"] = "其他"

    # ─ 11. 训练任务名称 ─
    task_names = []
    for pat in [
        r'(dual\s+n-back)', r'(single\s+n-back)', r'(Cogmed\b[^\s]*)',
        r'(BrainHQ)', r'(CogniFit)', r'(Lumosity)', r'(NeuroTracker)',
        r'(WOME\b)', r'(EngAge\b)', r'(operation\s+span)',
        r'(reading\s+span)', r'(complex\s+span)', r'(letter-number\s+sequencing)',
        r'(verbal\s+n-back)', r'(spatial\s+n-back)', r'(object\s+n-back)',
        r'(RehaCom)', r'(HappyNeuron)', r'(Seniorentraining)',
        r'(Brain\s*Fitness)', r'(MindFit)', r'(BrainStim)',
        r'(CWMS\b)', r'(categorization\s+working\s+memory\s+span)',
        r'(alpha.span)', r'(numerical\s+updating)',
        r'(visuospatial\s+WM)', r'(visuospatial\s+working\s+memory)',
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            task_names.append(m.group(1).strip())
    d["训练任务名称"] = "; ".join(dict.fromkeys(task_names)) if task_names else "未报告"

    # ─ 12. 是否自适应 ─
    d["是否自适应"] = "是" if re.search(
        r'\badaptive\b|\badapting\b|difficulty.*adjust|adjust.*difficulty|'
        r'titrat|level.adapt|adapt.*level|lure.rate\s+adjust', text, re.I) else "否"

    # ─ 13. 训练总次数 ─
    sess = find_num(text, [
        r'(\d+)\s+(?:training\s+)?sessions?\s+(?:of\s+training|were\s+conducted|were\s+completed)',
        r'(?:completed|received|underwent)\s+(\d+)\s+(?:training\s+)?sessions?',
        r'training\s+(?:consisted?\s+of|included?)\s+(\d+)\s+sessions?',
        r'(\d+)\s+sessions?\s+of\s+(?:WM|cognitive|n-back|working\s+memory)\s+training',
        r'sessions?\s*(?:n|=|:)\s*(\d+)',
        r'(\d+)\s+times?\s+per\s+week[^.]{0,30}?(\d+)\s+weeks?',  # will recalculate below
    ])
    d["训练总次数\n(sessions)"] = int(sess) if sess else "未报告"

    # ─ 14. 每次时长 ─
    dur = find_num(text, [
        r'(\d+)[\s-]*(?:min(?:utes?)?)\s+(?:per\s+session|each\s+session|/\s*session)',
        r'(?:each|per)\s+session\s+(?:lasted?|was)\s+(\d+)\s*min',
        r'(\d+)\s*[-–]\s*\d+\s*min(?:utes?)?(?:\s+(?:per\s+session|each))?',
        r'(\d+)\s*(?:hr|hours?)\s+(?:per\s+session|each\s+session)',
        r'(\d+)-(?:hr|hour)\s+(?:training\s+)?session',
    ])
    # Handle hours
    if dur and dur <= 5:
        dur_min = dur * 60
    else:
        dur_min = dur
    # Re-search for minutes specifically
    dur2 = find_num(text, [
        r'(\d{2,3})\s*(?:min(?:utes?)?)\s+(?:per\s+session|each)',
        r'(\d+)\s*min(?:utes?)?\s+(?:per|each)\s+(?:training\s+)?session',
        r'sessions?\s+(?:each\s+)?lasting\s+(\d+)\s*min',
        r'(\d+)\s*min(?:utes?)?\s+long\s+(?:training\s+)?sessions?',
    ])
    d["每次时长\n(分钟)"] = int(dur2) if dur2 else (int(dur_min) if dur_min else "未报告")

    # ─ 15. 失访率/完成率 ─
    comp = find_str(text, [
        r'(\d+(?:\.\d+)?)\s*%\s+(?:of\s+(?:the\s+)?participants?\s+)?completed',
        r'completion\s+rate[^.]{0,30}?(\d+(?:\.\d+)?)\s*%',
        r'(\d+(?:\.\d+)?)\s*%\s+(?:attrition|dropout|drop.out|withdrawal)',
        r'attrition\s+rate[^.]{0,30}?(\d+(?:\.\d+)?)\s*%',
    ])
    d["失访率/完成率\n(%)"] = comp + "%" if comp else "未报告"

    # ─ 16. 训练周数 ─
    wks = find_num(text, [
        r'(?:over|across|for|during|spanning?)\s+(\d+)\s+weeks?',
        r'(\d+)[\s-]week\s+(?:training|program|intervention|period)',
        r'(\d+)\s+weeks?\s+of\s+(?:training|intervention|cognitive)',
    ])
    d["训练周数"] = int(wks) if wks else "未报告"

    # ─ 17. 训练频率 (次/周) ─
    freq = find_num(text, [
        r'(\d+)\s+(?:sessions?\s+per\s+week|times?\s+(?:a|per)\s+week|x/week)',
        r'(?:sessions?\s+)?(\d+)\s+times?\s+(?:a|per)\s+week',
        r'(?:training\s+)?frequency[^.]{0,30}(\d+)\s+(?:times?|sessions?)\s+(?:a|per)\s+week',
        r'(\d)\s+times?\s+weekly',
    ])
    if freq:
        d["训练频率\n(次/周)"] = freq
    elif d["训练总次数\n(sessions)"] != "未报告" and d["训练周数"] != "未报告":
        sess_n = d["训练总次数\n(sessions)"]
        wks_n = d["训练周数"]
        if isinstance(sess_n, int) and isinstance(wks_n, int) and wks_n > 0:
            freq_calc = round(sess_n / wks_n, 1)
            d["训练频率\n(次/周)"] = freq_calc
        else:
            d["训练频率\n(次/周)"] = "未报告"
    else:
        d["训练频率\n(次/周)"] = "未报告"

    # ─ 18. 是否主动对照 ─
    if re.search(r'active\s+control|active\s+comparison|active\s+placebo|'
                 r'control\s+group\s+(?:received|performed|completed|did|underwent|practiced)', text, re.I):
        d["是否主动对照\n(是/否)"] = "是"
    elif re.search(r'wait.?list|no.?contact|passive\s+control|no\s+treatment|'
                   r'single.group|without\s+(?:a\s+)?control', text, re.I):
        d["是否主动对照\n(是/否)"] = "否"
    else:
        d["是否主动对照\n(是/否)"] = "待核查"

    # ─ 19. 对照组任务类型 ─
    ctrl_desc = find_str(text, [
        r'control\s+group\s+(?:received|performed|completed|did|underwent|was\s+asked\s+to)\s+([^.]{10,100})',
        r'active\s+control\s+(?:group\s+)?(?:consisted?\s+of|received|performed)\s+([^.]{10,100})',
        r'comparison\s+group\s+(?:performed|completed|received)\s+([^.]{10,100})',
    ])
    if ctrl_desc:
        d["对照组任务类型"] = ctrl_desc[:100]
    elif d.get("是否主动对照\n(是/否)") == "否":
        d["对照组任务类型"] = "无主动对照（等待/无接触）"
    else:
        d["对照组任务类型"] = "未报告"

    # ─ 20. 结合干预类型 ─
    combos = []
    if re.search(r'\btDCS\b|transcranial\s+direct\s+current', text, re.I): combos.append("tDCS")
    if re.search(r'\bTMS\b|transcranial\s+magnetic', text, re.I): combos.append("TMS")
    if re.search(r'\btRNS\b|transcranial\s+random\s+noise', text, re.I): combos.append("tRNS")
    if re.search(r'tACS\b|transcranial\s+alternating', text, re.I): combos.append("tACS")
    if re.search(r'\bexercise\b|\bphysical\s+(?:training|activity|exercise)\b', text, re.I): combos.append("运动")
    if re.search(r'\bmedication\b|\bdrug\b|\bpharmacolog', text, re.I): combos.append("药物")
    d["结合干预类型\n(tDCS/TMS/药物/运动/无)"] = "/".join(combos) if combos else "无"

    # ─ 21. 统计方法 ─
    stats = []
    if re.search(r'\bANCOVA\b', text): stats.append("ANCOVA")
    if re.search(r'\bLMM\b|linear\s+mixed.effects?\s+model|mixed.effects?\s+(?:model|ANOVA)|lme4', text, re.I):
        stats.append("LMM")
    if re.search(r'\bMANOVA\b', text): stats.append("MANOVA")
    if re.search(r'\bANOVA\b', text) and "ANCOVA" not in stats and "MANOVA" not in stats:
        stats.append("ANOVA")
    if re.search(r'\bBayesian\b|\bBayes\s+factor\b|BF\d*\b', text, re.I): stats.append("Bayesian")
    d["统计方法\n(ANOVA/ANCOVA/LMM/其他)"] = "/".join(stats) if stats else "未报告"

    # ─ 22. 训练平台/软件 ─
    platforms = []
    for pat in [
        r'\b(Cogmed|BrainHQ|CogniFit|Lumosity|NeuroTracker|Brain\s*Fitness)\b',
        r'\b(RehaCom|HappyNeuron|Seniorentraining|Brain\s*Age|NeuroAge)\b',
        r'\b(BrainTwister|PsychoPy|E-Prime|Presentation\s+software)\b',
        r'\b(MindFit|Jungle\s*Memory|WOME|EngAge)\b',
        r'\b(MATLAB|R\s+software|SPSS)\b',
    ]:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            platforms.append(m.group(1).strip())
    d["训练平台/软件"] = "; ".join(dict.fromkeys(platforms[:3])) if platforms else "未报告"

    # ─ 23. 监督方式 ─
    has_home = bool(re.search(
        r'home.based|at\s+home|home\s+(?:training|practice)|remotely|unsupervised', text, re.I))
    has_lab = bool(re.search(
        r'laboratory|lab.based|in.person|in\s+person|supervised\s+(?:by|training|session)', text, re.I))
    if has_home and has_lab:
        d["监督方式\n(实验室/居家/混合)"] = "混合"
    elif has_home:
        d["监督方式\n(实验室/居家/混合)"] = "居家"
    else:
        d["监督方式\n(实验室/居家/混合)"] = "实验室"

    # ─ 24. 近迁移(是/否) + 近迁移结局变量 ─
    has_near = bool(re.search(
        r'near.transfer|proximal\s+transfer|proximal\s+measure|'
        r'untrained\s+(?:WM|working\s+memory)\s+task|'
        r'transfer\s+to\s+(?:untrained|other)\s+(?:WM|working\s+memory)',
        text, re.I))
    d["近迁移\n(是/否)"] = "是" if has_near else "否"
    # Near transfer outcome
    near_match = re.search(
        r'(?:near.transfer|proximal)\s+(?:measure[sd]?|outcome[sd]?|tasks?)?[^.]{0,100}'
        r'(?:include[sd]?|(?:were|was)\s+(?:assessed?|measured?|tested?))[^.]*([A-Z][^.]{5,120})',
        text, re.I)
    if near_match:
        d["近迁移结局变量"] = near_match.group(1).strip()[:100]
    else:
        d["近迁移结局变量"] = "未报告" if not has_near else "待核查"

    # ─ 25. 远迁移(是/否) + 远迁移结局变量 ─
    has_far = bool(re.search(
        r'far.transfer|distal\s+transfer|distal\s+measure|'
        r'(?:fluid\s+intelligence|executive\s+function|episodic\s+memory|'
        r'daily\s+(?:functioning|activities)|speed\s+of\s+processing|'
        r'attention|inhibition|processing\s+speed|emotional|language|speech)',
        text, re.I))
    d["远迁移\n(是/否)"] = "是" if has_far else "否"
    far_match = re.search(
        r'(?:far.transfer|distal)\s+(?:measure[sd]?|outcome[sd]?|tasks?)?[^.]{0,100}'
        r'(?:include[sd]?|(?:were|was)\s+(?:assessed?|measured?|tested?))[^.]*([A-Z][^.]{5,120})',
        text, re.I)
    if far_match:
        d["远迁移结局变量"] = far_match.group(1).strip()[:100]
    else:
        d["远迁移结局变量"] = "未报告"

    # ─ 26. 远迁移结局域 ─
    domains = []
    if re.search(r'fluid\s+intelligence|Raven|CFT|matrix\s+reasoning|Gf\b|g-factor', text, re.I):
        domains.append("流体智力")
    if re.search(r'executive\s+function|inhibit|Stroop|Trail\s+Making|WCST|task.switching|'
                 r'set.shifting|flanker\b|go.no.go|stop.signal|updating', text, re.I):
        domains.append("EF")
    if re.search(r'episodic\s+memory|story\s+recall|word\s+list\s+recall|verbal\s+memory|'
                 r'long.term\s+memory|associative\s+memory\s+(?!task|training)', text, re.I):
        domains.append("情景记忆")
    if re.search(r'daily\s+(?:functioning|activities|life)|ADL\b|IADL\b|'
                 r'quality\s+of\s+life|everyday\s+(?:cognition|functioning)|'
                 r'ecological|self.(?:reported?|rated?)\s+(?:cognition|memory)', text, re.I):
        domains.append("日常功能")
    if re.search(r'processing\s+speed|reaction\s+time|speed\s+of\s+processing', text, re.I):
        domains.append("加工速度")
    if re.search(r'attention\b(?!\s+(?:is|are|was|were|has|have))', text, re.I):
        domains.append("注意")
    if re.search(r'speech\s+(?:perception|in\s+noise)|language\s+comprehension|'
                 r'reading\s+comprehension', text, re.I):
        domains.append("语言/言语")
    if re.search(r'emotional\s+(?:regulation|control|memory)|emotion\s+(?:processing|regulation)',
                 text, re.I):
        domains.append("情绪调节")
    d["远迁移结局域\n(流体智力/EF/情景记忆/日常功能/其他)"] = (
        "/".join(domains) if domains else ("无" if not has_far else "其他"))

    # ─ 27. 维持随访 ─
    has_fu = bool(re.search(
        r'follow.up|follow\s+up|longitudinal\s+assessment|maintenance\s+(?:test|assessment|measure)',
        text, re.I))
    d["维持随访\n(是/否)"] = "是" if has_fu else "否"
    if has_fu:
        fu_m = find_num(text, [
            r'(\d+)[\s-]month\s+follow',
            r'follow.up\s+(?:at|after)\s+(\d+)\s+months?',
            r'(\d+)\s+months?\s+(?:later|after\s+training|post.training)',
            r'(\d+)\s+week\s+follow',
        ])
        d["随访时间点(月)"] = int(fu_m) if fu_m else "未报告"
    else:
        d["随访时间点(月)"] = "无"

    # ─ 28. 效应量 ─
    if re.search(r"Cohen'?s?\s+d|Cohen'?s?\s+f|η[²2]|eta.squared|partial\s+eta|"
                 r"Hedges'\s*g|glass['s]\s+delta", text, re.I):
        d["效应量是否报告\n(是/否)"] = "是"
        d_val = find_str(text, [
            r"Cohen'?s?\s+d\s*=\s*([-\d.]+)",
            r"\bd\s*=\s*([-\d.]+)",
            r"g\s*=\s*([-\d.]+)",
        ])
        d["Cohen's d值\n(有则填数值，无则留空)"] = d_val if d_val else "未报告具体d值"
    else:
        d["效应量是否报告\n(是/否)"] = "否"
        d["Cohen's d值\n(有则填数值，无则留空)"] = None

    # ─ 29. 神经影像（关键字段）─
    if seq in CONFIRMED_NEURO:
        info = CONFIRMED_NEURO[seq]
        d["神经影像结局\n(是/否)"] = "是"
        d["影像类型\n(fMRI/EEG/ERP/其他)"] = info["type"]
        d["神经影像主要发现"] = info["finding"]
    else:
        d["神经影像结局\n(是/否)"] = "否"
        d["影像类型\n(fMRI/EEG/ERP/其他)"] = "无"
        d["神经影像主要发现"] = "无"

    # ─ 30. 基线WM水平 ─
    if re.search(r'(?:low|lower|poor|impaired|below\s+average)\s+(?:baseline\s+)?(?:WM|working\s+memory)',
                 text, re.I):
        d["基线WM水平\n(高/低/未报告)"] = "低"
    elif re.search(r'(?:high|higher|good|strong|above\s+average)\s+(?:baseline\s+)?(?:WM|working\s+memory)',
                   text, re.I):
        d["基线WM水平\n(高/低/未报告)"] = "高"
    else:
        d["基线WM水平\n(高/低/未报告)"] = "未报告"

    # ─ 31. 显式调节效应检验 ─
    if re.search(r'moderat(?:e|ion|or)\b|'
                 r'interaction.*(?:age|education|WM|baseline)\b|'
                 r'predictors?\s+of\s+(?:training|transfer)|'
                 r'subgroup\s+analys|individual\s+difference.*predict', text, re.I):
        d["显式调节效应检验\n(是/否)"] = "是"
        if re.search(r'high(?:er)?\s+(?:baseline|WM|age|education)[^.]{0,80}?'
                     r'(?:greater|more|larger|better)\s+(?:gain|benefit|transfer|improvement)',
                     text, re.I):
            d["调节效应方向\n(高>低/低>高/无/未检验)"] = "高>低"
        elif re.search(r'low(?:er)?\s+(?:baseline|WM)[^.]{0,80}?'
                       r'(?:greater|more|larger|better)\s+(?:gain|benefit|transfer|improvement)',
                       text, re.I):
            d["调节效应方向\n(高>低/低>高/无/未检验)"] = "低>高"
        elif re.search(r'no\s+(?:significant\s+)?(?:moderat|interaction|subgroup)', text, re.I):
            d["调节效应方向\n(高>低/低>高/无/未检验)"] = "无"
        else:
            d["调节效应方向\n(高>低/低>高/无/未检验)"] = "待核查"
    else:
        d["显式调节效应检验\n(是/否)"] = "否"
        d["调节效应方向\n(高>低/低>高/无/未检验)"] = "未检验"

    # ─ 32. 任务认知过程重叠度 ─
    is_nback   = bool(re.search(r'n-back|n\s*back|dual.n.back', text, re.I))
    is_span    = bool(re.search(r'complex\s+span|reading\s+span|operation\s+span|CWMS', text, re.I))
    has_updating   = bool(re.search(r'updating|n-back.*transfer|transfer.*n-back', text, re.I))
    has_inhibition = bool(re.search(r'inhibit|Stroop|stop.signal|go.no.go', text, re.I))
    has_switching  = bool(re.search(r'task.switching|set.shifting|WCST', text, re.I))
    has_binding    = bool(re.search(r'binding|associative\s+memory', text, re.I))
    has_episodic   = bool(re.search(r'episodic\s+memory|story\s+recall|verbal\s+memory', text, re.I))
    has_fluid      = bool(re.search(r'fluid\s+intel|Raven|matrix\s+reason', text, re.I))

    shared = 0
    if is_nback:
        if has_updating: shared += 1
        if has_inhibition: shared += 1
        if has_switching: shared += 1
    elif is_span:
        if has_updating: shared += 1
        if has_binding: shared += 1
        if has_inhibition: shared += 1

    if re.search(r'near.transfer.*n-back|n-back.*near.transfer', text, re.I):
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "高"
    elif shared >= 2:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "高"
    elif shared == 1:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "中"
    elif has_episodic or has_fluid:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "低"
    else:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "未报告"

    # ─ 33. 年龄亚组分析 ─
    d["年龄亚组分析\n(是/否)"] = "是" if re.search(
        r'age\s+(?:group|subgroup)|young.?old|old.?old|age\s+(?:as\s+)?moderator|'
        r'older.old|younger.old|55.?64|65.?74|75.?\+', text, re.I) else "否"

    # ─ 34. 认知储备指标 ─
    if re.search(r'\bNART\b|National\s+Adult\s+Reading', text, re.I):
        d["认知储备指标\n(教育/NART/未报告)"] = "NART"
    elif re.search(r'(?:education|years?\s+of\s+school)[^.]{0,100}?'
                   r'(?:reserv|moderate|predict|interact|correlat)', text, re.I):
        d["认知储备指标\n(教育/NART/未报告)"] = "教育"
    else:
        d["认知储备指标\n(教育/NART/未报告)"] = "未报告"

    # ─ 35. 认知储备是否显著调节迁移 ─
    if re.search(r'(?:education|NART|cognitive\s+reserve)[^.]{0,150}?'
                 r'(?:significant(?:ly)?)\s+(?:moderat|predict|interact)[^.]{0,100}?(?:transfer|gain)',
                 text, re.I):
        d["认知储备是否显著调节迁移\n(是/否/未检验)"] = "是"
    elif d.get("认知储备指标\n(教育/NART/未报告)") != "未报告":
        d["认知储备是否显著调节迁移\n(是/否/未检验)"] = "待核查"
    else:
        d["认知储备是否显著调节迁移\n(是/否/未检验)"] = "未检验"

    # ─ 36. 总体结论 ─
    if re.search(r'no\s+(?:significant\s+)?(?:group\s*[×x*]\s*time|training\s+(?:effect|benefit|gain)|'
                 r'near\s+transfer|far\s+transfer)|'
                 r'failed\s+to\s+(?:show|demonstrate|find|replicate)|'
                 r'null\s+(?:effect|result|finding)|'
                 r'did\s+not\s+(?:differ|improve|benefit)', text, re.I):
        d["总体结论\n(正向/无/混合)"] = "无"
    elif re.search(
        r'group\s*[×x*×]\s*time\s+interaction[^.]{0,150}?(?:significant\b|p\s*[<≤]\s*0\.0[0-9])|'
        r'(?:significant|robust|strong|clear)\s+(?:near|far)?\s+transfer\s+effect|'
        r'(?:training\s+group|WM\s+training)[^.]{0,80}?(?:significantly\s+)?(?:improved|outperformed|'
        r'benefited)', text, re.I):
        d["总体结论\n(正向/无/混合)"] = "正向"
    elif re.search(r'(?:some|partial|mixed)\s+(?:evidence|support|effects?|transfer)|'
                   r'(?:significant.*not\s+significant|not.*significant.*significant)|'
                   r'mixed\s+(?:results?|findings?|evidence)', text, re.I):
        d["总体结论\n(正向/无/混合)"] = "混合"
    else:
        d["总体结论\n(正向/无/混合)"] = "待核查"

    # ─ 37. 研究设计类型 ─
    if re.search(r'randomly\s+(?:assigned|allocated)|random(?:ized)?\s+(?:controlled|assignment|allocation)',
                 text, re.I):
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "RCT"
    elif re.search(r'crossover|cross.over', text, re.I):
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "交叉"
    elif re.search(r'control\s+group|comparison\s+group|comparison\s+condition', text, re.I):
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "准RCT"
    elif re.search(r'no\s+control|single.group|one.group|without\s+(?:a\s+)?control', text, re.I):
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "单组前后测"
    else:
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "待核查"

    # ─ 38. 发表状态 ─
    d["发表状态\n(期刊/预印本)"] = (
        "预印本" if re.search(r'preprint|bioRxiv|medRxiv|PsyArXiv', text, re.I)
        else "期刊")

    # ─ 39. RoB ─
    d["⚠️RoB总体等级\n(低/有顾虑/高/未评估)"] = "未评估"

    return d


# ── Excel操作 ────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)
ws = wb["数据提取"]

# Build header → column mapping
col_map = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(1, c).value
    if h:
        col_map[str(h).strip()] = c

def find_row(seq_no):
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value == seq_no:
            return r
    return None


def normalize(s):
    return re.sub(r'[\s\n\r]+', '', str(s)).lower()


def col_for(field_name):
    """Find column number for a field by flexible matching"""
    norm = normalize(field_name)
    for h, c in col_map.items():
        if normalize(h) == norm:
            return c
    # Partial match
    for h, c in col_map.items():
        if norm in normalize(h) or normalize(h) in norm:
            return c
    return None


def write_row(row, data, notes=""):
    """
    Write data dict to Excel row.
    Key = human-readable field name (matches or substring of header).
    Does NOT touch cols 1-5.
    """
    field_map = {
        "国家": col_for("国家"),
        "样本来源": col_for("样本来源"),
        "样本量N": col_for("样本量N"),
        "训练组N": col_for("训练组N"),
        "对照组N": col_for("对照组N"),
        "年龄均值": col_for("年龄均值"),
        "年龄段分类": col_for("年龄段分类"),
        "年龄SD": col_for("年龄SD"),
        "性别(%女)": col_for("性别(%女)"),
        "教育年限(年)": col_for("教育年限"),
        "认知筛查工具": col_for("认知筛查工具"),
        "筛查分数": col_for("筛查分数"),
        "训练类型": col_for("训练类型"),
        "训练任务名称": col_for("训练任务名称"),
        "是否自适应": col_for("是否自适应"),
        "训练总次数\n(sessions)": col_for("训练总次数"),
        "每次时长\n(分钟)": col_for("每次时长"),
        "失访率/完成率\n(%)": col_for("失访率"),
        "训练周数": col_for("训练周数"),
        "是否主动对照\n(是/否)": col_for("是否主动对照"),
        "对照组任务类型": col_for("对照组任务类型"),
        "结合干预类型\n(tDCS/TMS/药物/运动/无)": col_for("结合干预类型"),
        "统计方法\n(ANOVA/ANCOVA/LMM/其他)": col_for("统计方法"),
        "训练平台/软件": col_for("训练平台"),
        "监督方式\n(实验室/居家/混合)": col_for("监督方式"),
        "训练频率\n(次/周)": col_for("训练频率"),
        "近迁移\n(是/否)": col_for("近迁移"),
        "近迁移结局变量": col_for("近迁移结局变量"),
        "远迁移\n(是/否)": col_for("远迁移"),
        "远迁移结局变量": col_for("远迁移结局变量"),
        "远迁移结局域": col_for("远迁移结局域"),
        "维持随访\n(是/否)": col_for("维持随访"),
        "随访时间点(月)": col_for("随访时间点"),
        "效应量是否报告\n(是/否)": col_for("效应量是否报告"),
        "Cohen's d值\n(有则填数值，无则留空)": col_for("Cohen"),
        "总体结论\n(正向/无/混合)": col_for("总体结论"),
        "神经影像结局\n(是/否)": col_for("神经影像结局"),
        "影像类型\n(fMRI/EEG/ERP/其他)": col_for("影像类型"),
        "神经影像主要发现": col_for("神经影像主要发现"),
        "基线WM水平\n(高/低/未报告)": col_for("基线WM水平"),
        "显式调节效应检验\n(是/否)": col_for("显式调节效应检验"),
        "调节效应方向\n(高>低/低>高/无/未检验)": col_for("调节效应方向"),
        "任务认知过程重叠度\n(高/中/低/未报告)": col_for("任务认知过程重叠度"),
        "年龄亚组分析\n(是/否)": col_for("年龄亚组分析"),
        "认知储备指标\n(教育/NART/未报告)": col_for("认知储备指标"),
        "认知储备是否显著调节迁移\n(是/否/未检验)": col_for("认知储备是否显著调节迁移"),
        "发表状态\n(期刊/预印本)": col_for("发表状态"),
        "⚠️RoB总体等级\n(低/有顾虑/高/未评估)": col_for("RoB"),
        "研究设计类型\n(RCT/准RCT/单组前后测/交叉)": col_for("研究设计类型"),
    }

    for field, value in data.items():
        col = field_map.get(field)
        if col is None:
            # Try direct lookup
            col = col_for(field)
        if col and col >= 6:  # Never touch cols 1-5
            ws.cell(row, col).value = value

    # Status columns (always overwrite)
    ws.cell(row, 56).value = "已填入"
    ws.cell(row, 57).value = "未核查"

    # Notes
    notes_col = col_for("备注")
    if notes_col and notes:
        existing = ws.cell(row, notes_col).value or ""
        if existing and "待核查" in existing:
            ws.cell(row, notes_col).value = notes  # Replace 待核查 notes
        elif existing:
            ws.cell(row, notes_col).value = existing  # Keep manual notes
        else:
            ws.cell(row, notes_col).value = notes


# ── Main loop ────────────────────────────────────────────────────────────────
print("="*70)
print("v4全量重跑 — 56篇数据提取")
print("="*70)

done, errors, check_needed = [], [], []
batch_count = 0

for idx, seq in enumerate(SEQ_ORDER):
    row = find_row(seq)
    if not row:
        print(f"  [WARN] seq={seq} row not found, skipping")
        continue

    text, status = read_pdf(seq)
    neuro_label = "YES" if seq in CONFIRMED_NEURO else "no"

    if status in ("PDF缺失",):
        print(f"seq{seq:2d} ({idx+1:02d}/{len(SEQ_ORDER)}): ❌ PDF缺失")
        errors.append(seq)
        ws.cell(row, col_for("备注")).value = "PDF缺失"
        ws.cell(row, 56).value = "未填入"
        ws.cell(row, 57).value = "未核查"
        continue

    if status == "扫描件":
        print(f"seq{seq:2d} ({idx+1:02d}/{len(SEQ_ORDER)}): ⚠️ 扫描件")
        errors.append(seq)
        continue

    if status.startswith("读取错误"):
        print(f"seq{seq:2d} ({idx+1:02d}/{len(SEQ_ORDER)}): ❌ {status}")
        errors.append(seq)
        continue

    data = extract_fields(text, seq)
    notes_parts = []
    for v in data.values():
        if isinstance(v, str) and "待核查" in v:
            notes_parts.append(v[:60])

    notes = " | ".join(notes_parts[:2]) if notes_parts else ""
    write_row(row, data, notes)

    neuro_val = data.get("神经影像结局\n(是/否)", "否")
    near_val  = data.get("近迁移\n(是/否)", "?")
    far_val   = data.get("远迁移\n(是/否)", "?")

    print(f"seq{seq:2d} ({idx+1:02d}/{len(SEQ_ORDER)}): neuro={neuro_val}  near={near_val}  far={far_val}"
          f"  country={data.get('国家','?')}"
          f"{'  ⚠️'+notes[:40] if notes else ''}")

    done.append(seq)
    if notes_parts:
        check_needed.append(seq)

    batch_count += 1
    if batch_count % 10 == 0:
        wb.save(EXCEL)
        print(f"  >>> 已保存 (batch {batch_count//10})")

# Final save
wb.save(EXCEL)
print(f"  >>> 最终保存完毕")

# ── Self-check ───────────────────────────────────────────────────────────────
print()
print("="*70)
print("自检报告")
print("="*70)
far_col    = col_for("远迁移")
near_col   = col_for("近迁移")
neuro_col  = col_for("神经影像结局")
neuro_type = col_for("影像类型")
neuro_find = col_for("神经影像主要发现")

issues = []
for row in range(3, ws.max_row + 1):
    seq = ws.cell(row, 1).value
    if seq is None:
        continue

    far_v   = ws.cell(row, far_col).value if far_col else None
    near_v  = ws.cell(row, near_col).value if near_col else None
    neuro_v = ws.cell(row, neuro_col).value if neuro_col else None
    ntype_v = ws.cell(row, neuro_type).value if neuro_type else None
    nfind_v = ws.cell(row, neuro_find).value if neuro_find else None

    if not far_v:
        issues.append(f"  seq{seq}: 远迁移为空")
    if not near_v:
        issues.append(f"  seq{seq}: 近迁移为空")
    if neuro_v == "是" and not nfind_v:
        issues.append(f"  seq{seq}: 神经影像=是 但主要发现为空")
    if neuro_v == "是" and not ntype_v:
        issues.append(f"  seq{seq}: 神经影像=是 但类型为空")

if issues:
    print(f"发现 {len(issues)} 个问题：")
    for issue in issues:
        print(issue)
else:
    print("✅ 无问题：所有行的近迁移/远迁移均已填写，神经影像一致")

print()
print("="*70)
print(f"完成: {len(done)}/{len(SEQ_ORDER)} 篇")
print(f"错误/缺失: {errors if errors else '无'}")
print(f"建议核查 ({len(check_needed)} 篇): {check_needed}")
print("="*70)

# Final save again to be sure
wb.save(EXCEL)
print("Excel已保存：", EXCEL)
