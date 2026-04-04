"""
Paper-01 数据提取表 v2 — 核查报告 Excel sheet（带颜色）
- 修正教育年限列名
- 更新 md 报告
- 在 v2.xlsx 新增"核查报告"sheet（带颜色高亮）
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

BASE = os.path.dirname(os.path.abspath(__file__))
V2_PATH = os.path.join(BASE, "数据_6_数据提取表_v2.xlsx")
REPORT_PATH = os.path.join(BASE, "核查报告_异常值.md")

# ── 颜色常量 ──────────────────────────────────────────────────
C_HEADER    = "2F5496"   # 深蓝：表头
C_SECTION   = "D6E4F7"   # 淡蓝：分节标题
C_ANOMALY   = "FFD7D7"   # 粉红：异常值行
C_MISSING3  = "FFF2CC"   # 淡黄：3+缺失字段
C_MISSING1  = "FAFAFA"   # 近白：1-2缺失字段
C_OK        = "E2EFDA"   # 浅绿：无问题
C_WHITE     = "FFFFFF"
C_GRAY      = "F2F2F2"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(color="000000", size=10):
    return Font(bold=True, color=color, size=size)

def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── 读取工作簿 ────────────────────────────────────────────────
wb = load_workbook(V2_PATH)
ws = wb["数据提取"]

# ── 列名 → 列号 ───────────────────────────────────────────────
headers = {}
for col in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=col).value
    if v:
        headers[str(v).strip()] = col

def fcol(key, fallbacks=()):
    if key in headers: return headers[key]
    for fb in fallbacks:
        if fb in headers: return headers[fb]
    return None

col_seq      = fcol("序号")
col_author   = fcol("第一作者")
col_year     = fcol("年份")
col_age_mean = fcol("年龄均值")
col_age_cat  = fcol("年龄段分类\n(young-old≤75/old-old>75)")
col_sessions = fcol("训练总次数\n(sessions)")
col_weeks    = fcol("训练周数")
col_freq     = fcol("训练频率\n(次/周)")
col_n_total  = fcol("样本量N")
col_duration = fcol("每次时长\n(分钟)")
col_gender   = fcol("性别(%女)")
col_edu      = fcol("教育年限(年)")   # ← 修正后的实际列名

print(f"教育年限列：{col_edu}")

# ── 数据行 ────────────────────────────────────────────────────
DATA_START = 3
data_rows = []
for row in range(DATA_START, ws.max_row + 1):
    seq = ws.cell(row=row, column=col_seq).value if col_seq else None
    if seq is None:
        break
    data_rows.append(row)

# ── 计算训练频率（已在脚本1做过，这里重新确认写入）────────────
freq_computed = freq_missing = 0
for row in data_rows:
    if col_freq:
        try:
            s = float(ws.cell(row=row, column=col_sessions).value)
            w = float(ws.cell(row=row, column=col_weeks).value)
            if w > 0:
                ws.cell(row=row, column=col_freq).value = round(s / w, 1)
                freq_computed += 1
            else:
                freq_missing += 1
        except (TypeError, ValueError):
            freq_missing += 1

# ── 计算年龄段分类（已做，重新确认）──────────────────────────
age_young = age_old = age_missing = 0
for row in data_rows:
    if col_age_cat:
        try:
            a = float(ws.cell(row=row, column=col_age_mean).value)
            ws.cell(row=row, column=col_age_cat).value = "young-old" if a <= 75 else "old-old"
            if a <= 75: age_young += 1
            else: age_old += 1
        except (TypeError, ValueError):
            age_missing += 1

# ── 异常值检测 ────────────────────────────────────────────────
CHECKS = [
    ("年龄均值",        col_age_mean,  60,  95),
    ("样本量N",         col_n_total,   10, 500),
    ("训练总次数",      col_sessions,   3, 100),
    ("训练周数",        col_weeks,      1,  52),
    ("训练频率(次/周)", col_freq,       1,  14),
    ("每次时长(分钟)",  col_duration,  10, 120),
]

anomalies = []
for row in data_rows:
    seq    = ws.cell(row=row, column=col_seq).value
    author = ws.cell(row=row, column=col_author).value
    year   = ws.cell(row=row, column=col_year).value
    for field, col, lo, hi in CHECKS:
        if col is None: continue
        try:
            val = float(ws.cell(row=row, column=col).value)
            if val < lo or val > hi:
                anomalies.append((seq, author, year, field, val, lo, hi))
        except (TypeError, ValueError):
            pass

# ── 缺失值统计 ────────────────────────────────────────────────
# "未报告"也算有值，只有 None 才算缺失
MISS_FIELDS = [
    ("年龄均值",       col_age_mean),
    ("样本量N",        col_n_total),
    ("训练总次数",     col_sessions),
    ("训练周数",       col_weeks),
    ("每次时长(分钟)", col_duration),
    ("性别(%女)",      col_gender),
    ("教育年限(年)",   col_edu),
]

missing_stats = []
all_missing_by_row = {}  # seq -> list of missing field names

for fname, col in MISS_FIELDS:
    if col is None:
        missing_stats.append((fname, "列不存在", []))
        continue
    missing_list = []
    for row in data_rows:
        val = ws.cell(row=row, column=col).value
        if val is None or str(val).strip() == "":
            seq    = ws.cell(row=row, column=col_seq).value
            author = ws.cell(row=row, column=col_author).value
            year   = ws.cell(row=row, column=col_year).value
            missing_list.append((seq, author, year))
            all_missing_by_row.setdefault(seq, []).append(fname)
    missing_stats.append((fname, len(missing_list), missing_list))

# ── 优先核查列表 ──────────────────────────────────────────────
priority = {}
for row in data_rows:
    seq    = ws.cell(row=row, column=col_seq).value
    author = ws.cell(row=row, column=col_author).value
    year   = ws.cell(row=row, column=col_year).value
    priority[seq] = {"author": author, "year": year, "score": 0, "issues": []}

for a in anomalies:
    seq = a[0]
    if seq in priority:
        priority[seq]["score"] += 2
        priority[seq]["issues"].append(f"异常：{a[3]}={a[4]}")

for seq, fields in all_missing_by_row.items():
    critical = [f for f in fields if f in ("年龄均值","样本量N","训练总次数","训练周数","每次时长(分钟)")]
    if seq in priority:
        priority[seq]["score"] += len(critical)
        for f in critical:
            priority[seq]["issues"].append(f"缺失：{f}")

ranked = sorted([(k, v) for k, v in priority.items() if v["score"] > 0],
                key=lambda x: x[1]["score"], reverse=True)[:20]

# ════════════════════════════════════════════════════════════════
#  生成 Excel "核查报告" sheet
# ════════════════════════════════════════════════════════════════
SHEET_NAME = "核查报告"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
rws = wb.create_sheet(SHEET_NAME)

now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

def write_cell(ws, row, col, value, bg=None, font=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    if bg:    c.fill      = fill(bg)
    if font:  c.font      = font
    if align: c.alignment = align
    if border: c.border   = thin_border()
    return c

def section_title(ws, row, text, ncols=7, bg=C_SECTION):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = fill(bg)
    c.font      = bold_font("1F3864", 11)
    c.alignment = left()
    c.border    = thin_border()

# ── 标题行 ────────────────────────────────────────────────────
rws.merge_cells("A1:G1")
c = rws.cell(row=1, column=1, value="📊 数据核查报告 — Paper-01 数据提取表 v2")
c.fill      = fill(C_HEADER)
c.font      = Font(bold=True, color="FFFFFF", size=13)
c.alignment = center()

rws.merge_cells("A2:G2")
c = rws.cell(row=2, column=1, value=f"生成时间：{now_str}    数据来源：数据_6_数据提取表_v2.xlsx（第3行起，共{len(data_rows)}行）")
c.fill      = fill("EBF3FB")
c.font      = Font(italic=True, color="444444", size=10)
c.alignment = left()
rws.row_dimensions[1].height = 28
rws.row_dimensions[2].height = 18

# ── 第一节：计算字段 ──────────────────────────────────────────
r = 4
section_title(rws, r, "一、计算字段完成情况")
r += 1

headers_1 = ["字段", "成功计算", "留空（缺失）", "备注"]
col_widths_1 = [20, 12, 14, 35]
for i, h in enumerate(headers_1, 1):
    write_cell(rws, r, i, h, bg=C_HEADER, font=bold_font("FFFFFF"), align=center())
r += 1

data_1 = [
    ("训练频率(次/周)",   freq_computed, freq_missing,  "=总次数÷训练周数，保留1位小数"),
    ("年龄段分类",        age_young+age_old, age_missing,
     f"young-old(≤75岁)：{age_young}篇 / old-old(>75岁)：{age_old}篇"),
]
for row_data in data_1:
    for i, v in enumerate(row_data, 1):
        bg = C_OK if (i == 2 and v > 0) else C_WHITE
        write_cell(rws, r, i, v, bg=bg if i <= 2 else C_WHITE, align=center() if i<=3 else left())
    r += 1

# ── 第二节：异常值 ────────────────────────────────────────────
r += 1
section_title(rws, r, "二、异常值清单")
r += 1

headers_2 = ["序号", "第一作者", "年份", "异常字段", "异常值", "正常范围", "建议核查"]
for i, h in enumerate(headers_2, 1):
    write_cell(rws, r, i, h, bg=C_HEADER, font=bold_font("FFFFFF"), align=center())
r += 1

if anomalies:
    for seq, author, year, field, val, lo, hi in anomalies:
        row_data = [seq, author, year, field, val, f"{lo}–{hi}", "回原文确认"]
        for i, v in enumerate(row_data, 1):
            write_cell(rws, r, i, v, bg=C_ANOMALY, align=center() if i != 4 else left())
        r += 1
else:
    rws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = rws.cell(row=r, column=1, value="✅  未发现异常值")
    c.fill = fill(C_OK); c.font = Font(color="375623"); c.alignment = center()
    r += 1

# ── 第三节：缺失值 ────────────────────────────────────────────
r += 1
section_title(rws, r, "三、缺失值统计（关键字段）")
r += 1

headers_3 = ["字段", "缺失篇数", "缺失文献（作者 年份）"]
col_widths_3 = [20, 10, 60]
rws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
for i, h in enumerate(headers_3, 1):
    col_end = i if i < 3 else 7
    write_cell(rws, r, i, h, bg=C_HEADER, font=bold_font("FFFFFF"), align=center())
r += 1

for fname, cnt, mlist in missing_stats:
    bg = C_ANOMALY if isinstance(cnt, int) and cnt >= 5 else \
         C_MISSING3 if isinstance(cnt, int) and cnt >= 3 else \
         C_MISSING1 if isinstance(cnt, int) and cnt >= 1 else C_OK

    detail = "; ".join(f"{a} {y}" for _, a, y in mlist) if isinstance(mlist, list) and mlist else \
             ("无" if cnt == 0 else str(cnt))

    write_cell(rws, r, 1, fname,  bg=bg, align=left())
    write_cell(rws, r, 2, cnt,    bg=bg, align=center())
    rws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    write_cell(rws, r, 3, detail, bg=bg, align=left())
    r += 1

# ── 第四节：优先核查列表 ──────────────────────────────────────
r += 1
section_title(rws, r, "四、需要人工核查的优先列表（综合异常值+关键字段缺失）")
r += 1

headers_4 = ["优先级", "序号", "第一作者", "年份", "问题描述"]
rws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
for i, h in enumerate(headers_4, 1):
    write_cell(rws, r, i, h, bg=C_HEADER, font=bold_font("FFFFFF"), align=center())
r += 1

if ranked:
    for rank, (seq, info) in enumerate(ranked, 1):
        issues_str = "；".join(info["issues"])
        bg = C_ANOMALY if rank <= 3 else C_MISSING3 if rank <= 8 else C_MISSING1
        row_data = [rank, seq, info["author"], info["year"], issues_str]
        for i, v in enumerate(row_data, 1):
            col_end = i if i < 5 else 7
            if i == 5:
                rws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
            write_cell(rws, r, i, v, bg=bg, align=center() if i <= 4 else left())
        r += 1
else:
    rws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c = rws.cell(row=r, column=1, value="✅  无需人工核查")
    c.fill = fill(C_OK); c.font = Font(color="375623"); c.alignment = center()
    r += 1

# ── 列宽设置 ──────────────────────────────────────────────────
col_widths = [10, 14, 10, 30, 14, 14, 30]
for i, w in enumerate(col_widths, 1):
    rws.column_dimensions[get_column_letter(i)].width = w

# 行高
for row_i in range(1, r + 1):
    rws.row_dimensions[row_i].height = 20

# 冻结前两行
rws.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════
#  更新 md 报告
# ════════════════════════════════════════════════════════════════
lines = [
    "# 数据核查报告 — 异常值检测",
    f"生成时间：{now_str}",
    f"数据来源：数据_6_数据提取表_v2.xlsx（第3行起，共{len(data_rows)}行）",
    "",
    "## 一、计算字段完成情况",
    f"- 训练频率：成功计算 {freq_computed} 篇，{freq_missing} 篇因数据缺失留空",
    f"- 年龄段分类：成功计算 {age_young + age_old} 篇，{age_missing} 篇因数据缺失留空",
    f"  - young-old（≤75岁）：{age_young} 篇",
    f"  - old-old（>75岁）：{age_old} 篇",
    "",
    "## 二、异常值清单",
]
if anomalies:
    lines += [
        "| 序号 | 第一作者 | 年份 | 异常字段 | 异常值 | 正常范围 | 建议核查 |",
        "|------|---------|------|---------|-------|---------|---------|",
    ]
    for seq, author, year, field, val, lo, hi in anomalies:
        lines.append(f"| {seq} | {author} | {year} | {field} | {val} | {lo}–{hi} | 回原文确认 |")
else:
    lines.append("未发现异常值")

lines += ["", "## 三、缺失值统计",
          "| 字段 | 缺失篇数 | 缺失文献（作者+年份）|",
          "|------|---------|------------------|"]
for fname, cnt, mlist in missing_stats:
    detail = "; ".join(f"{a} {y}" for _, a, y in mlist) if isinstance(mlist, list) and mlist else \
             ("无" if cnt == 0 else str(cnt))
    lines.append(f"| {fname} | {cnt} | {detail} |")

lines += ["", "## 四、需要人工核查的优先列表",
          "（按问题数排序，最多20篇）", "",
          "| 优先级 | 序号 | 第一作者 | 年份 | 问题描述 |",
          "|--------|------|---------|------|---------|"]
for rank, (seq, info) in enumerate(ranked, 1):
    issues_str = "；".join(info["issues"])
    lines.append(f"| {rank} | {seq} | {info['author']} | {info['year']} | {issues_str} |")
if not ranked:
    lines.append("无需人工核查（无异常值且关键字段完整）")

with open(REPORT_PATH, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

# ── 保存 v2 ───────────────────────────────────────────────────
wb.save(V2_PATH)
print(f"✅ v2 已保存（含「核查报告」sheet）：{V2_PATH}")
print(f"✅ md 报告已更新：{REPORT_PATH}")
print(f"\n=== 汇总 ===")
print(f"  教育年限(年) 缺失：{next(cnt for fn,cnt,_ in missing_stats if fn=='教育年限(年)')} 篇")
print(f"  训练频率成功计算：{freq_computed} 篇")
print(f"  young-old：{age_young} / old-old：{age_old} / 缺失：{age_missing}")
print(f"  异常值：{len(anomalies)} 条")
print(f"  优先核查列表：{len(ranked)} 篇")
