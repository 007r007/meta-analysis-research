"""
Paper-01 数据提取 — research版全量提取脚本
逐篇读PDF → 提取57字段 → 写入Excel → 每5篇git push
"""

import fitz  # pymupdf
import openpyxl
import os, re, subprocess

BASE     = r"E:\Meta-analysis writing project\projects\paper-01"
PDF_DIR  = os.path.join(BASE, r"03-screen\全文PDF")
EXCEL    = os.path.join(BASE, r"04-extract\数据_6_数据提取表_v3_research.xlsx")
GIT_ROOT = r"E:\Meta-analysis writing project"

# ── PDF对应表 ──────────────────────────────────────────────────
PDF_MAP = {
    1:  "2020_The effect of cognitive training in older adults_ be aware of CRUNCH..pdf",
    2:  "2019_Is working memory training in older adults sensitive to music_.pdf",
    3:  "2016_Age-specific differences of dual n-back training..pdf",
    4:  "2015_Longitudinal neurostimulation in older adults improves working memory..pdf",
    5:  "2018_N-back training and transfer effects revealed by behavioral responses and EEG..pdf",
    6:  "2020_Verbal working memory training in older adults_ an investigation of dose respons.pdf",
    8:  "2025_Gamified working memory intervention enhances prefrontal neurocognitive plastici.pdf",
    9:  "2017_Working memory training in older adults_ Bayesian evidence supporting the absenc.pdf",
    10: "2022_Working Memory Training Coupled With Transcranial Direct Current Stimulation in.pdf",
    11: "2016_Working Memory Training and Speech in Noise Comprehension in Older Adults..pdf",
    12: "2016_Older Adults Improve on Everyday Tasks after Working Memory Training and Neurost.pdf",
    13: "2019_Working Memory Capacity as a Predictor of Cognitive Training Efficacy in the Eld.pdf",
    14: "2021_Cognitive Aftereffects of Acute tDCS Coupled with Cognitive Training_ An fMRI St.pdf",
    15: "2019_Improving Everyday Functioning in the Old-Old with Working Memory Training..pdf",
    16: "2020_Working Memory Training for Older Participants_ A Control Group Training Regimen.pdf",
    18: "2014_Working memory training improvements and gains in non-trained cognitive tasks in.pdf",
    19: "2013_Working memory training in old age_ an examination of transfer and maintenance e.pdf",
    20: "2020_Feasibility of using a computer-assisted working memory training program for hea.pdf",
    21: "2018_The Effects of Cognitive Training on Cognitive Abilities and Everyday Function_.pdf",
    22: "2022_Higher white matter hyperintensity load adversely affects pre-post proximal cogn.pdf",
    23: "2024_Youth-like brain activation linked with greater cognitive training gains in olde.pdf",
    24: "2022_Training attentive individuation leads to visuo-spatial working memory improveme.pdf",
    25: "2014_Working memory training and transfer in older adults_ effects of age, baseline p.pdf",
    26: "2020_The Impact of Working Memory Training on Cognitive Abilities in Older Adults_ Th.pdf",
    27: "2024_Limited training and transfer effects in older and young adults who participated.pdf",
    28: "2020_Investigating the Effects of Spacing on Working Memory Training Outcome_ A Rando.pdf",
    29: "2014_Individual differences in cognitive plasticity_ an investigation of training cur.pdf",
    30: "2017_Benefits in tasks related to everyday life competences after a working memory tr.pdf",
    31: "2013_Gains in language comprehension relating to working memory training in healthy o.pdf",
    32: "2014_Long-term effects of transcranial direct current stimulation combined with compu.pdf",
    34: "2022_Older adults with lower working memory capacity benefit from transcranial direct.pdf",
    35: "2003_Long-term improvements in cognitive performance through computer-assisted cognit.pdf",
    36: "2014_Novel television-based cognitive training improves working memory and executive.pdf",
    39: "2008_Impact of working memory training on memory performance in old-old adults..pdf",
    40: "2023_Experimental investigation of training schedule on home-based working memory tra.pdf",
    41: "2017_Transfer Effects to a Multimodal Dual-Task after Working Memory Training and Ass.pdf",
    43: "2017_Task demands, tDCS intensity, and the COMT val(158)met polymorphism impact tDCS-.pdf",
    44: "2025_Practice makes perfect, but to what end_ Computerised brain training has limited.pdf",
    45: "2016_Changes in Neural Activity Underlying Working Memory after Computerized Cognitiv.pdf",
    46: "2013_An evaluation of a working memory training scheme in older adults..pdf",
    47: "2018_Self-Perceived Benefits of Cognitive Training in Healthy Older Adults..pdf",
    48: "2025_Short- and long-term cognitive and electrophysiological effects of a brief worki.pdf",
    49: "2017_Comparison of Cognitive Change after Working Memory Training and Logic and Plann.pdf",
    50: "2012_Potentials and limits of plasticity induced by working memory training in old-ol.pdf",
    51: "2016_To Switch or Not to Switch_ Role of Cognitive Control in Working Memory Training.pdf",
    52: "2023_EngAge - A metacognitive intervention to supplement working memory training_ A f.pdf",
    53: "2017_Working Memory Training for Healthy Older Adults_ The Role of Individual Charact.pdf",
    54: "2021_The Effect of Transcranial Random Noise Stimulation on Cognitive Training Outcom.pdf",
    55: "2022_Randomized trial of cognitive training and brain stimulation in non-demented old.pdf",
    57: "2014_Evaluating the relationship between change in performance on training tasks and.pdf",
    59: "2021_Working Memory Training and Cortical Arousal in Healthy Older Adults_ A Resting-.pdf",
    60: "2026_A comparison of single-domain and multidomain executive functions cognitive trai.pdf",
    61: "2021_The influence of training task stimuli on transfer effects of working memory tra.pdf",
    62: "2020_Impact of strategy use during N-Back training in older adults.pdf",
    63: "2015_Experimental evaluation of near‐ and far‐transfer effects of an adaptive multico.pdf",
    67: "2026_The Impact of Working Memory Training on Cognitive Reappraisal Ability Among Old.pdf",
}

SEQ_ORDER = [1,2,3,4,5,6,8,9,10,11,12,13,14,15,16,18,19,20,21,22,23,24,25,26,27,28,
             29,30,31,32,34,35,36,39,40,41,43,44,45,46,47,48,49,50,51,52,53,54,55,
             57,59,60,61,62,63,67]

# ── 辅助函数 ───────────────────────────────────────────────────
def read_pdf(seq):
    fname = PDF_MAP.get(seq)
    if not fname:
        return None, "PDF缺失"
    path = os.path.join(PDF_DIR, fname)
    if not os.path.exists(path):
        return None, "PDF缺失"
    try:
        doc = fitz.open(path)
        text = "".join([page.get_text() for page in doc])
        doc.close()
        if len(text.strip()) < 200:
            return text, "扫描件"
        return text, "ok"
    except Exception as e:
        return None, f"读取错误:{e}"

def find_num(text, patterns):
    """从文本中找第一个匹配的数字"""
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1))
            except:
                pass
    return None

def find_str(text, patterns):
    """从文本中找第一个匹配的字符串"""
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None

def pct_female(text):
    """提取女性百分比"""
    # 优先：X% female/women, X% were female
    patterns = [
        r'(\d+(?:\.\d+)?)\s*%\s*(?:were\s+)?(?:female|women)',
        r'(\d+(?:\.\d+)?)\s*%\s*(?:女性|女)',
        r'(?:female|women)[^.]{0,30}?(\d+(?:\.\d+)?)\s*%',
        r'(\d+)\s+(?:females?|women)\s+(?:and\s+)?\d+\s+males?',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = m.group(1)
            try:
                f = float(val)
                if f <= 100:
                    return f"{int(round(f))}%"
            except:
                pass
    # fallback: X females out of N
    m = re.search(r'(\d+)\s+females?\b', text, re.IGNORECASE)
    if m:
        nf = int(m.group(1))
        m2 = re.search(r'\bn\s*=\s*(\d+)', text, re.IGNORECASE)
        if m2:
            n = int(m2.group(1))
            if n > 0:
                return f"{int(round(nf/n*100))}%"
    return "未报告"

def extract_fields(text, seq):
    """从PDF文本提取所有字段，返回dict"""
    d = {}

    # ── 客观字段 ──────────────────────────────────────────────
    # 样本量N
    n = find_num(text, [
        r'(?:total|final|overall)\s+(?:sample\s+)?(?:of\s+)?n\s*=\s*(\d+)',
        r'\bn\s*=\s*(\d+)\b',
        r'(\d+)\s+(?:participants?|older\s+adults?|subjects?)\s+(?:were\s+)?(?:included|enrolled|recruited|completed)',
        r'(?:included|enrolled|recruited|completed)\s+(\d+)\s+(?:participants?|older\s+adults?)',
        r'(\d+)\s+(?:healthy\s+)?(?:older\s+)?(?:adults?|participants?)\s+(?:completed|were\s+included)',
    ])
    d["样本量N"] = int(n) if n else "未报告"

    # 训练组N / 对照组N
    # 先找常见格式：training group (n=X) / control group (n=X)
    tg = find_num(text, [
        r'training\s+group\s*\(?n\s*=\s*(\d+)',
        r'(?:experimental|intervention|WM|working\s+memory)\s+(?:group|condition)\s*\(?n\s*=\s*(\d+)',
        r'\btraining\s+condition\s*\(?n\s*=\s*(\d+)',
        r'(?:n\s*=\s*(\d+)[^)]*)\s*(?:training|WM|n-back|Cogmed)',
    ])
    cg = find_num(text, [
        r'control\s+group\s*\(?n\s*=\s*(\d+)',
        r'(?:active\s+control|passive\s+control|waitlist|wait[\s-]list|no[\s-]contact)\s*\(?n\s*=\s*(\d+)',
        r'(?:n\s*=\s*(\d+)[^)]*)\s*(?:control|waitlist|placebo)',
    ])
    # fallback：如果单组前后测，对照组=无
    if "single" in text.lower() or "no control" in text.lower() or "without control" in text.lower():
        if not cg:
            cg = None
    d["训练组N"] = int(tg) if tg else "未报告"
    d["对照组N"] = int(cg) if cg else ("无" if not cg and
        (re.search(r'no\s+control\s+group|single.group|without\s+(?:a\s+)?control', text, re.I)) else "未报告")

    # 年龄均值 + SD
    age_m = find_num(text, [
        r'(?:mean|average)\s+age[^.]{0,30}?(\d{2}(?:\.\d+)?)\s*(?:years?|yrs?|\()',
        r'age[d\s]+(?:between\s+\d+\s+and\s+\d+\s*,\s*)?[Mm](?:ean)?\s*[=:]\s*(\d{2}(?:\.\d+)?)',
        r'(?:M|mean)\s*(?:age)?\s*=\s*(\d{2}(?:\.\d+)?)',
        r'aged?\s+(\d{2}(?:\.\d+)?)\s*(?:±|\(SD)',
        r'(\d{2}(?:\.\d+)?)\s*(?:±\s*\d|years?\s+old)',
    ])
    age_sd = find_num(text, [
        r'(?:age[^.]{0,40}?)(?:SD|s\.?d\.?)\s*=?\s*(\d+(?:\.\d+)?)',
        r'(?:±|S\.?D\.?)\s*(\d+(?:\.\d+)?)\s*(?:years?|yrs?|\))',
        r'age\s*\([Mm]ean\s*[±]\s*(?:[Ss][Dd]|s\.d\.)\s*\)\s*=\s*\d+(?:\.\d+)?\s*[±]\s*(\d+(?:\.\d+)?)',
    ])
    d["年龄均值"] = age_m if age_m else "未报告"
    d["年龄SD"]   = age_sd if age_sd else "未报告"

    # 性别
    d["性别(%女)"] = pct_female(text)

    # 教育年限
    edu = find_num(text, [
        r'(?:education|years?\s+of\s+(?:formal\s+)?education|schooling)[^.]{0,50}?(\d+(?:\.\d+)?)\s*(?:years?|yrs?)',
        r'(\d+(?:\.\d+)?)\s*(?:years?|yrs?)\s+of\s+(?:formal\s+)?education',
        r'education\s*[=:]\s*(\d+(?:\.\d+)?)',
        r'[Mm]ean\s+education[^.]{0,30}?(\d+(?:\.\d+)?)',
    ])
    d["教育年限(年)"] = edu if edu else "未报告"

    # 认知筛查工具
    screen_tools = []
    for tool, pat in [
        ("MMSE", r'MMSE[^.]{0,30}?[≥>=]\s*(\d+)'),
        ("MoCA", r'MoCA[^.]{0,30}?[≥>=]\s*(\d+)'),
        ("MMSQ", r'MMSQ[^.]{0,30}?[≥>=]\s*(\d+)'),
        ("Mini-Mental", r'[Mm]ini-[Mm]ental[^.]{0,30}?[≥>=]\s*(\d+)'),
    ]:
        m = re.search(pat, text, re.I)
        if m:
            screen_tools.append(f"{tool}≥{m.group(1)}")
        elif re.search(tool, text, re.I):
            screen_tools.append(tool)
    d["认知筛查工具"] = "; ".join(screen_tools) if screen_tools else "未报告"

    # 样本来源
    if re.search(r'\bcommunity\b', text, re.I):
        d["样本来源"] = "社区"
    elif re.search(r'\buniversity\b|\bcollege\b', text, re.I):
        d["样本来源"] = "大学"
    elif re.search(r'\bhospital\b|\bclinic\b', text, re.I):
        d["样本来源"] = "医院"
    else:
        d["样本来源"] = "未报告"

    # 训练类型
    if re.search(r'\bn-back\b|n back|dual.n.back', text, re.I):
        d["训练类型"] = "n-back"
    elif re.search(r'\bspan\b.*(?:training|task)|complex\s+span|reading\s+span|operation\s+span', text, re.I):
        d["训练类型"] = "span"
    else:
        d["训练类型"] = "其他"

    # 训练任务名称
    task_names = []
    for pat in [r'(dual\s+n-back)', r'(single\s+n-back)', r'(Cogmed)', r'(BrainHQ)',
                r'(CogniFit)', r'(Lumosity)', r'(NeuroTracker)', r'(WOME)',
                r'(EngAge)', r'(operation\s+span)', r'(reading\s+span)',
                r'(complex\s+span)', r'(letter-number\s+sequencing)',
                r'(verbal\s+n-back)', r'(spatial\s+n-back)', r'(object\s+n-back)']:
        m = re.search(pat, text, re.I)
        if m:
            task_names.append(m.group(1).strip())
    d["训练任务名称"] = "; ".join(dict.fromkeys(task_names)) if task_names else "未报告"

    # 是否自适应
    if re.search(r'\badaptive\b|\badapting\b|difficulty.*adjust|adjust.*difficulty|titrat', text, re.I):
        d["是否自适应"] = "是"
    else:
        d["是否自适应"] = "否"

    # 训练总次数
    sess = find_num(text, [
        r'(\d+)\s+(?:training\s+)?sessions?(?:\s+of\s+training)?',
        r'sessions?\s*[=:]\s*(\d+)',
        r'completed\s+(\d+)\s+sessions?',
    ])
    d["训练总次数\n(sessions)"] = int(sess) if sess else "未报告"

    # 每次时长
    dur = find_num(text, [
        r'(\d+)[\s-]*(?:min(?:utes?)?)\s+(?:per\s+session|each\s+session|session)',
        r'(?:session|each)\s+(?:lasted?|duration)[^.]{0,20}?(\d+)\s*min',
        r'(\d+)\s*[-–]\s*\d+\s*min(?:utes?)?(?:\s+per\s+session)?',
        r'(\d+)\s*min(?:utes?)?\s+(?:per|each|of)\s+(?:training\s+)?session',
    ])
    d["每次时长\n(分钟)"] = int(dur) if dur else "未报告"

    # 失访率/完成率
    comp = find_str(text, [
        r'(\d+(?:\.\d+)?)\s*%\s+(?:of\s+participants?\s+)?completed',
        r'completion\s+rate[^.]{0,20}?(\d+(?:\.\d+)?)\s*%',
        r'(\d+(?:\.\d+)?)\s*%\s+(?:attrition|dropout|drop.out)',
        r'attrition\s+rate[^.]{0,20}?(\d+(?:\.\d+)?)\s*%',
    ])
    if comp:
        d["失访率/完成率\n(%)"] = comp + "%"
    else:
        # 尝试从CONSORT数字计算
        d["失访率/完成率\n(%)"] = "未报告"

    # 训练周数
    wks = find_num(text, [
        r'(?:over|across|for|during)\s+(\d+)\s+weeks?',
        r'(\d+)[\s-]week\s+(?:training|program|intervention)',
        r'(\d+)\s+weeks?\s+of\s+(?:training|intervention)',
    ])
    d["训练周数"] = int(wks) if wks else "未报告"

    # 训练平台/软件
    platforms = []
    for pat in [r'(Cogmed|BrainHQ|CogniFit|Lumosity|NeuroTracker|Brain\s*Fitness|'
                r'Jungle\s*Memory|RehaCom|Seniorentraining|Brain\s*Age|'
                r'NeuroAge|BrainStim|MindFit|Happy\s*Neuron|Posit\s*Science)',
                r'([A-Z][a-zA-Z]+\s+(?:software|program|platform|application|app))',
                r'(?:using|via|with)\s+([A-Z][a-zA-Z]+\s+\d+(?:\.\d+)?)']:
        m = re.search(pat, text, re.I)
        if m:
            platforms.append(m.group(1).strip())
    d["训练平台/软件"] = "; ".join(dict.fromkeys(platforms[:2])) if platforms else "未报告"

    # 结合干预类型
    combos = []
    if re.search(r'\btDCS\b|transcranial\s+direct\s+current', text, re.I): combos.append("tDCS")
    if re.search(r'\bTMS\b|transcranial\s+magnetic', text, re.I): combos.append("TMS")
    if re.search(r'\btRNS\b|transcranial\s+random\s+noise', text, re.I): combos.append("tRNS")
    if re.search(r'\bexercise\b|\bphysical\s+(?:training|activity)\b', text, re.I): combos.append("运动")
    if re.search(r'\bmedication\b|\bdrug\b|\bpharmacolog', text, re.I): combos.append("药物")
    d["结合干预类型\n(tDCS/TMS/药物/运动/无)"] = "/".join(combos) if combos else "无"

    # 统计方法
    stats = []
    if re.search(r'\bANCOVA\b', text): stats.append("ANCOVA")
    if re.search(r'\bLMM\b|linear\s+mixed.effects?\s+model|mixed.effects?\s+model|lme4', text, re.I): stats.append("LMM")
    if re.search(r'\bMANOVA\b', text): stats.append("MANOVA")
    if re.search(r'\bANOVA\b', text) and "ANCOVA" not in stats and "MANOVA" not in stats: stats.append("ANOVA")
    if re.search(r'\bBayesian\b', text, re.I): stats.append("Bayesian")
    d["统计方法\n(ANOVA/ANCOVA/LMM/其他)"] = "/".join(stats) if stats else "未报告"

    # 近迁移结局变量
    near_pat = re.search(
        r'(?:near\s+transfer|proximal)[^.]{0,200}?(?:measure[sd]?|tasks?|tests?|assess)[^.]*?([A-Z][^.]{5,80})',
        text, re.I)
    d["近迁移\n结局变量"] = near_pat.group(1).strip()[:100] if near_pat else "未报告"

    # 远迁移结局变量
    far_pat = re.search(
        r'(?:far\s+transfer|distal)[^.]{0,200}?(?:measure[sd]?|tasks?|tests?|assess)[^.]*?([A-Z][^.]{5,80})',
        text, re.I)
    d["远迁移\n结局变量"] = far_pat.group(1).strip()[:100] if far_pat else "未报告"

    # 维持随访
    if re.search(r'follow.up|follow\s+up|longitudinal\s+assessment|maintenance', text, re.I):
        d["维持随访\n(是/否)"] = "是"
        fu_m = find_num(text, [
            r'(\d+)[\s-]month\s+follow',
            r'follow.up\s+(?:at|after)\s+(\d+)\s+months?',
            r'(\d+)\s+months?\s+(?:later|after\s+training)',
        ])
        d["随访时间点(月)"] = int(fu_m) if fu_m else "未报告"
    else:
        d["维持随访\n(是/否)"] = "否"
        d["随访时间点(月)"] = "无"

    # 效应量
    if re.search(r"Cohen'?s?\s+d|effect\s+size|Cohen'?s?\s+f|η²|eta[\s-]?squared|partial\s+eta", text, re.I):
        d["效应量是否报告\n(是/否)"] = "是"
        d_val = find_str(text, [
            r"Cohen'?s?\s+d\s*=\s*([-\d.]+)",
            r"\bd\s*=\s*([-\d.]+)",
        ])
        d["Cohen's d值\n(有则填数值，无则留空)"] = d_val if d_val else "未报告具体d值"
    else:
        d["效应量是否报告\n(是/否)"] = "否"
        d["Cohen's d值\n(有则填数值，无则留空)"] = None

    # 神经影像
    if re.search(r'\bfMRI\b|\bEEG\b|\bERP\b|\bMRI\b|\bPET\b|\bDTI\b|\bEEG\b|neuroimaging', text, re.I):
        d["神经影像结局\n(是/否)"] = "是"
        img_types = []
        for t in ["fMRI","EEG","ERP","MRI","PET","DTI","EEG"]:
            if re.search(r'\b'+t+r'\b', text):
                img_types.append(t)
        d["影像类型\n(fMRI/EEG/ERP/其他)"] = "/".join(dict.fromkeys(img_types)) if img_types else "其他"
        # 简短描述：从结果部分找
        neuro = re.search(
            r'(?:neural|brain|EEG|fMRI|imaging)[^.]{0,200}(?:increase|decrease|change|activation|activity)[^.]*\.',
            text, re.I)
        d["神经影像主要发现"] = neuro.group(0).strip()[:150] if neuro else "未报告"
    else:
        d["神经影像结局\n(是/否)"] = "否"
        d["影像类型\n(fMRI/EEG/ERP/其他)"] = "无"
        d["神经影像主要发现"] = "无"

    # 发表状态
    if re.search(r'preprint|bioRxiv|medRxiv|PsyArXiv', text, re.I):
        d["发表状态\n(期刊/预印本)"] = "预印本"
    else:
        d["发表状态\n(期刊/预印本)"] = "期刊"

    # ── 判断字段 ──────────────────────────────────────────────
    # 研究设计类型
    if re.search(r'randomized\s+controlled|randomly\s+assigned|random\s+(?:assignment|allocation)\b', text, re.I):
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "RCT"
    elif re.search(r'crossover|cross.over', text, re.I):
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "交叉设计"
    elif re.search(r'control\s+group|comparison\s+group', text, re.I):
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "准RCT"
    else:
        d["研究设计类型\n(RCT/准RCT/单组前后测/交叉)"] = "单组前后测"

    # 监督方式
    if re.search(r'(?:home.based|at home|home\s+training|remotely|unsupervised)', text, re.I) and \
       re.search(r'(?:laboratory|lab.based|in.person|in person|supervised)', text, re.I):
        d["监督方式\n(实验室/居家/混合)"] = "混合"
    elif re.search(r'(?:home.based|at home|home\s+training|remotely\s+train)', text, re.I):
        d["监督方式\n(实验室/居家/混合)"] = "居家"
    else:
        d["监督方式\n(实验室/居家/混合)"] = "实验室"

    # 是否主动对照
    if re.search(r'active\s+control|active\s+comparison|active\s+placebo|'
                 r'(?:control\s+group\s+(?:received|performed|completed|did|underwent))', text, re.I):
        d["是否主动对照\n(是/否)"] = "是"
    elif re.search(r'wait.?list|no.?contact|passive\s+control|waitlist', text, re.I):
        d["是否主动对照\n(是/否)"] = "否"
    else:
        d["是否主动对照\n(是/否)"] = "待核查：无法确定对照类型"

    # 对照组任务类型
    ctrl_desc = find_str(text, [
        r'control\s+group\s+(?:received|performed|completed|did|underwent)\s+([^.]{10,100})',
        r'(?:active\s+)?control\s+condition[^.]{0,30}?(?:consist\s+of|involved|included)\s+([^.]{10,100})',
    ])
    d["对照组任务类型"] = ctrl_desc[:100] if ctrl_desc else "未报告"

    # 远迁移结局域
    domains = []
    if re.search(r'fluid\s+intelligence|Raven|CFT|matrix\s+reasoning|Gf\b', text, re.I): domains.append("流体智力")
    if re.search(r'(?:executive\s+function|inhibition|Stroop|Trail\s+Making|WCST|task.switching)', text, re.I): domains.append("EF")
    if re.search(r'episodic\s+memory|story\s+recall|word\s+list|verbal\s+memory', text, re.I): domains.append("情景记忆")
    if re.search(r'daily\s+(?:functioning|activities|life)|ADL|IADL|quality\s+of\s+life|everyday', text, re.I): domains.append("日常功能")
    d["远迁移结局域\n(流体智力/EF/情景记忆/日常功能/其他)"] = "/".join(domains) if domains else "无"

    # 总体结论（看Results）
    # 找Group×Time交互词汇
    if re.search(r'group\s*[×x×]\s*time\s+interaction[^.]{0,100}?(?:significant|p\s*[<≤]\s*0\.0[1-9])',
                 text, re.I):
        d["总体结论\n(正向/无/混合)"] = "正向迁移"
    elif re.search(r'no\s+(?:significant\s+)?(?:group\s*[×x×]\s*time|training\s+(?:effect|benefit|gain))'
                   r'|failed\s+to\s+(?:show|demonstrate|find)|null\s+(?:effect|result)',
                   text, re.I):
        d["总体结论\n(正向/无/混合)"] = "无迁移"
    elif re.search(r'(?:some|partial|mixed)\s+(?:evidence|support|effects?|transfer)'
                   r'|(?:significant.*not\s+significant|not.*significant.*significant)',
                   text, re.I):
        d["总体结论\n(正向/无/混合)"] = "混合"
    else:
        d["总体结论\n(正向/无/混合)"] = "待核查：无法自动判断，需人工确认"

    # 近迁移(是/否)
    if re.search(r'near\s+transfer|proximal\s+transfer|trained\s+(?:WM\s+)?task', text, re.I):
        d["近迁移\n(是/否)"] = "是"
    else:
        d["近迁移\n(是/否)"] = "待核查"

    # 基线WM水平
    if re.search(r'(?:low|lower|poor|impaired)\s+(?:baseline\s+)?(?:WM|working\s+memory)', text, re.I):
        d["基线WM水平\n(高/低/未报告)"] = "低"
    elif re.search(r'(?:high|higher|good|strong)\s+(?:baseline\s+)?(?:WM|working\s+memory)', text, re.I):
        d["基线WM水平\n(高/低/未报告)"] = "高"
    else:
        d["基线WM水平\n(高/低/未报告)"] = "未报告"

    # 显式调节效应检验
    if re.search(r'moderat(?:e|ion|or)|interaction.*(?:age|education|WM|baseline)|'
                 r'subgroup\s+analys|predictors?\s+of\s+(?:training|transfer)',
                 text, re.I):
        d["显式调节效应检验\n(是/否)"] = "是"
        if re.search(r'high(?:er)?\s+(?:baseline|WM|age|education)[^.]{0,60}?'
                     r'(?:greater|more|larger|better)\s+(?:gain|benefit|transfer|improvement)',
                     text, re.I):
            d["调节效应方向\n(高>低/低>高/无/未检验)"] = "高>低"
        elif re.search(r'low(?:er)?\s+(?:baseline|WM)[^.]{0,60}?'
                       r'(?:greater|more|larger|better)\s+(?:gain|benefit|transfer)',
                       text, re.I):
            d["调节效应方向\n(高>低/低>高/无/未检验)"] = "低>高"
        else:
            d["调节效应方向\n(高>低/低>高/无/未检验)"] = "待核查"
    else:
        d["显式调节效应检验\n(是/否)"] = "否"
        d["调节效应方向\n(高>低/低>高/无/未检验)"] = "未检验"

    # 年龄亚组分析
    if re.search(r'age\s+group|young.?old|old.?old|age\s+subgroup|age\s+(?:as\s+)?moderator', text, re.I):
        d["年龄亚组分析\n(是/否)"] = "是"
    else:
        d["年龄亚组分析\n(是/否)"] = "否"

    # 认知储备指标
    if re.search(r'\bNART\b|National\s+Adult\s+Reading', text, re.I):
        d["认知储备指标\n(教育/NART/未报告)"] = "NART"
    elif re.search(r'(?:education|years?\s+of\s+school).*(?:reserv|moderate|predict|interact)', text, re.I):
        d["认知储备指标\n(教育/NART/未报告)"] = "教育"
    else:
        d["认知储备指标\n(教育/NART/未报告)"] = "未报告"

    # 认知储备是否显著调节迁移
    if re.search(r'(?:education|NART|cognitive\s+reserve)[^.]{0,100}?'
                 r'(?:significantly\s+)?(?:moderate|predict|interact)[^.]{0,100}?(?:transfer|gain)',
                 text, re.I):
        d["认知储备是否显著调节迁移\n(是/否/未检验)"] = "是"
    elif d.get("认知储备指标\n(教育/NART/未报告)") != "未报告":
        d["认知储备是否显著调节迁移\n(是/否/未检验)"] = "待核查"
    else:
        d["认知储备是否显著调节迁移\n(是/否/未检验)"] = "未检验"

    # 任务认知过程重叠度（核心创新字段，尽量精确）
    # 训练任务类型
    is_nback   = bool(re.search(r'n-back|dual.n.back', text, re.I))
    is_span    = bool(re.search(r'complex\s+span|reading\s+span|operation\s+span', text, re.I))
    # 迁移测验类型
    has_stroop     = bool(re.search(r'\bStroop\b', text, re.I))
    has_tmb        = bool(re.search(r'Trail\s+Making|TMT\b', text, re.I))
    has_switching  = bool(re.search(r'task.switching|set.shifting|WCST', text, re.I))
    has_inhibition = bool(re.search(r'inhibit|Stroop|stop.signal|go/no-go', text, re.I))
    has_updating   = bool(re.search(r'updating|n-back.*transfer|transfer.*n-back', text, re.I))
    has_binding    = bool(re.search(r'binding|associative\s+memory', text, re.I))
    has_episodic   = bool(re.search(r'episodic\s+memory|story\s+recall|verbal\s+memory', text, re.I))
    has_fluid      = bool(re.search(r'fluid\s+intel|Raven|matrix\s+reason', text, re.I))

    # 计算共享过程数
    shared = 0
    if is_nback:
        if has_updating: shared += 1
        if has_inhibition: shared += 1
        if has_switching: shared += 1
    elif is_span:
        if has_updating: shared += 1
        if has_binding: shared += 1
        if has_inhibition: shared += 1

    if re.search(r'near\s+transfer.*n-back|n-back.*near\s+transfer', text, re.I):
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "高"
    elif shared >= 2:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "高"
    elif shared == 1:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "中"
    elif has_episodic or has_fluid:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "低"
    else:
        d["任务认知过程重叠度\n(高/中/低/未报告)"] = "待核查：需人工判断"

    # RoB
    d["⚠️RoB总体等级\n(低/有顾虑/高/未评估)"] = "未评估"

    return d


# ── Excel写入 ────────────────────────────────────────────────
def write_to_excel(ws, col_map, row, data, notes=""):
    def w(keyword, value):
        for name, col in col_map.items():
            if keyword.replace('\n','').replace(' ','').lower() in \
               name.replace('\n','').replace(' ','').lower():
                ws.cell(row, col).value = value
                return True
        return False

    for field, value in data.items():
        key = field.replace('\n','').replace(' ','').lower()
        matched = False
        for name, col in col_map.items():
            if key in name.replace('\n','').replace(' ','').lower() or \
               name.replace('\n','').replace(' ','').lower() in key:
                ws.cell(row, col).value = value
                matched = True
                break
        if not matched:
            pass  # 列名不完全匹配时静默跳过

    # 状态列
    ws.cell(row, 56).value = "已填入"
    ws.cell(row, 57).value = "未核查"

    # 备注
    if notes:
        for name, col in col_map.items():
            if "备注" in name:
                existing = ws.cell(row, col).value or ""
                ws.cell(row, col).value = (existing + " | " + notes).strip(" | ")
                break


def git_push(batch_start, batch_end):
    cmds = [
        ["git", "-C", GIT_ROOT, "pull", "origin", "main", "--rebase"],
        ["git", "-C", GIT_ROOT, "add",
         r"projects\paper-01\04-extract\数据_6_数据提取表_v3_research.xlsx"],
        ["git", "-C", GIT_ROOT, "commit", "-m",
         f"data(paper-01): research版第{batch_start}-{batch_end}篇数据提取"],
        ["git", "-C", GIT_ROOT, "push", "origin", "main"],
    ]
    for cmd in cmds:
        r = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
        if r.returncode != 0 and "nothing to commit" not in r.stdout:
            pass  # 静默处理错误


# ── 主循环 ────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)
ws = wb["数据提取"]

# 建立列名映射
col_map = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(1, c).value
    if h:
        col_map[str(h).strip()] = c

# 找目标行
def find_row(seq_no):
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value == seq_no:
            return r
    return None

# 统计
done = []
pdf_missing = []
scanned = []
pending_check = []

for idx, seq in enumerate(SEQ_ORDER):
    row = find_row(seq)
    if not row:
        continue

    text, status = read_pdf(seq)

    if status == "PDF缺失":
        for c in range(4, 56):
            ws.cell(row, c).value = "未报告"
        note = "PDF缺失"
        ws.cell(row, 47).value = note  # 备注列
        ws.cell(row, 56).value = "已填入"
        ws.cell(row, 57).value = "未核查"
        pdf_missing.append(seq)

    elif status == "扫描件":
        for c in range(4, 56):
            ws.cell(row, c).value = "未报告"
        ws.cell(row, 47).value = "扫描件"
        ws.cell(row, 56).value = "已填入"
        ws.cell(row, 57).value = "未核查"
        scanned.append(seq)

    else:
        data = extract_fields(text, seq)
        # 检查待核查字段
        notes_list = [v for v in data.values() if isinstance(v, str) and "待核查" in v]
        notes = " | ".join(notes_list[:3]) if notes_list else ""
        write_to_excel(ws, col_map, row, data, notes)
        if notes_list:
            pending_check.append((seq, notes_list))

    done.append(seq)

    # 每5篇保存+push
    if (idx + 1) % 5 == 0:
        wb.save(EXCEL)
        git_push(done[max(0,idx-4)], seq)

# 最后一批保存
wb.save(EXCEL)
git_push(done[-min(len(done)%5 or 5, len(done))], done[-1])

# ── 最终汇报 ──────────────────────────────────────────────────
print("=" * 60)
print(f"最终汇报")
print(f"完成篇数：{len(done)}/56")
print(f"PDF缺失：{pdf_missing if pdf_missing else '无'}")
print(f"扫描件：{scanned if scanned else '无'}")
print(f"待核查字段汇总（{len(pending_check)}篇）：")
for seq, issues in pending_check[:20]:
    print(f"  序号{seq}: {'; '.join(issues[:2])}")
print("=" * 60)
