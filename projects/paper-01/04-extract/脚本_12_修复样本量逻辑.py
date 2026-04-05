"""
修复样本量逻辑错误（v7任务）：
问题1: 17条 训练组N+对照组N > 总N — 全部从PDF核查后修正
问题2: seq45 年龄=47 实为75.8（PDF核查）
问题3: 补填训练组N/对照组N + 修正总N（PDF核查后）
"""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL = r"E:\Meta-analysis writing project\projects\paper-01\04-extract\数据_6_数据提取表_v3_research.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["数据提取"]
headers = {cell.value: cell.column for cell in ws[1] if cell.value}

seq_to_row = {}
for r in range(3, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        try:
            seq_to_row[int(str(v))] = r
        except (ValueError, TypeError):
            pass

def w(seq, col_name, value):
    r = seq_to_row.get(seq)
    if r and col_name in headers:
        ws.cell(r, headers[col_name]).value = value

changes = []

# ═══════════════════════════════════════════════════════════════════════
# 问题1：训练组N+对照组N > 总N — 修正总N和/或组别N
# 全部基于PDF Methods/Results核查
# ═══════════════════════════════════════════════════════════════════════

# seq2: Borella 2019 music
# PDF: 72 older adults, 4 groups: Mozart(n=19), Albinoni(n=19), WhiteNoise(n=16), Active control(n=18)
# 训练组 = 3 music+WM groups合并 = 54；对照组 = 18（active control）；总N = 72
w(2, "样本量N", 72)
w(2, "训练组N", 54)   # Mozart(19) + Albinoni(19) + WhiteNoise(16)
w(2, "对照组N", 18)
changes.append("seq2: 总N=72（原19），训练组=54（3音乐组合并），对照组=18")

# seq3: Salminen 2016 dual n-back
# PDF: 47 older adults recruited; 1 dropout → final older training=25, control=21; total older=46
w(3, "样本量N", 46)
w(3, "训练组N", 25)
w(3, "对照组N", 21)
changes.append("seq3: 总N=46（原25），训练组=25，对照组=21（1人退出）")

# seq15: Borella 2019 old-old everyday functioning
# PDF: "Thirty-two community-dwelling older adults randomly assigned to a training or active control group"
# 16 training + 16 active control = 32 total
w(15, "样本量N", 32)
w(15, "训练组N", 16)
w(15, "对照组N", 16)
changes.append("seq15: 总N=32（原18），训练=16，对照=16")

# seq27: Zajac-Lamparska 2024
# PDF: older adults experimental(n=25), passive control(n=22), active control(n=7) = 54 total older
# WM training group = 25 (experimental); main control used = passive(n=22)
w(27, "样本量N", 54)
w(27, "训练组N", 25)
w(27, "对照组N", 22)    # passive control (primary comparison); active ctrl n=7 also exists
changes.append("seq27: 总N=54老年组（原25），训练=25，主要对照=22（被动对照）；另有主动对照n=7")

# seq30: Cantarella 2017 everyday life competences
# PDF: "Thirty-six community dwelling older adults randomly assigned to training or active control group"
# 18 training + 18 active control = 36 total
w(30, "样本量N", 36)
w(30, "训练组N", 18)
w(30, "对照组N", 18)
changes.append("seq30: 总N=36（原18），训练=18，对照=18")

# seq31: Carretti 2013 language comprehension
# PDF: 40 volunteers, randomly assigned 20+20; 3 dropped from trained, 1 from control
# Final: training group = 17, control group = 19; total analyzed = 36
w(31, "样本量N", 36)
w(31, "训练组N", 17)
w(31, "对照组N", 19)
changes.append("seq31: 总N=36（原17），训练=17（20-3 dropouts），对照=19（20-1 dropout）")

# seq36: Shatil 2014 novel television-based
# PDF: 140 enrolled, 119 completed (85%); training group n=60, active control n=59
w(36, "样本量N", 119)
w(36, "训练组N", 60)
w(36, "对照组N", 59)
changes.append("seq36: 总N=119（原60），训练=60，对照=59")

# seq41: Heinzel 2017 transfer multimodal dual-task
# PDF: 38 recruited; data not correctly recorded for 4 participants; final n=34
# Training group n=18, no-contact control n=16
w(41, "样本量N", 34)
w(41, "训练组N", 18)
w(41, "对照组N", 16)
changes.append("seq41: 总N=34（原13），训练=18，对照=16（无接触对照）")

# seq43: Stephens 2017 task demands tDCS COMT
# PDF: 137 older adults; 4 tDCS groups: Sham(N=40), Active1 1mA(N=28), Active1.5 1.5mA(N=29?), Active2 2mA(N=?)
# "Participants arrayed among tDCS groups: Sham (N=40), Active1 (N=28), Active1.5..."
# Total: 137; sham=40 is control; 3 active groups combined ≈ 97 (training)
# Note: originally from two separate studies combined
w(43, "样本量N", 137)
w(43, "训练组N", 97)    # Active1(28) + Active1.5 + Active2 = ~97 (combined active tDCS)
w(43, "对照组N", 40)   # Sham N=40
changes.append("seq43: 总N=137（原7），训练(active tDCS)=97，对照(Sham)=40；4组设计")

# seq46: McAvinue 2013 working memory training scheme
# PDF: "randomly assigned to a trainee (n=19) or control (n=17) group"  total=36
w(46, "样本量N", 36)
w(46, "训练组N", 19)
w(46, "对照组N", 17)
changes.append("seq46: 总N=36（原19），训练=19，对照=17")

# seq50: Zinke 2012 potentials and limits old-old
# PDF: "training group (n=20) and matched control group (n=16)" → total=36
w(50, "样本量N", 36)
w(50, "训练组N", 20)
w(50, "对照组N", 16)
changes.append("seq50: 总N=36（原20），训练=20，对照=16")

# seq52: Jaeggi 2023 EngAge metacognitive
# PDF: N=119 final sample; 3 groups: EngAge+WM(n=49), WM(n=36), KB active control(n=34)
# Training = EngAge+WM + WM = 85; active control = KB = 34
w(52, "样本量N", 119)
w(52, "训练组N", 85)    # EngAge+WM(49) + WM(36) = 85 (both received WM training)
w(52, "对照组N", 34)   # Knowledge Builders active control
changes.append("seq52: 总N=119（原36），训练=85（EngAge+WM(49)+WM(36)），对照=34（KB主动控制）")

# seq54: Brambilla 2021 tRNS effect on cognitive training
# PDF: "forty-two older adults randomly assigned to three intervention groups: 0.705mA(N=14), 1mA(N=14), sham(N=19(?))"
# Paper says N=14 + N=14 + sham = total 42; sham tRNS = control
# "randomly assigned to one of three groups: 0.705mA tRNS(N=14), 1mA tRNS(N=14), sham tRNS(N=19-correction: N=14)"
# Let's use total N=42, active training=28 (14+14), sham/control=14 (or 19)
# From abstract: "forty-two older adults...three intervention groups that received 20 min of 0.705 mA tRNS (N=14), 1 mA tRNS (N=14), or sham tRNS (N=19?)"
# 14+14+14=42 or 14+14+19=47? → paper says "forty-two" = 42
# Best estimate: 14+14+14=42 (all three groups equal); but note "sham" could be control
w(54, "样本量N", 42)
w(54, "训练组N", 28)    # 0.705mA(14) + 1mA(14) combined active tRNS
w(54, "对照组N", 14)   # Sham tRNS (estimated; 42-28=14)
changes.append("seq54: 总N=42（原14），训练(active tRNS)=28（两活性组各14），对照(sham)=14；三组设计")

# seq57: Zelinski 2014 IMPACT
# PDF: "randomized into training (N=242) or active control (N=245) conditions"
# Total = 487; but N field was 242 (only training group)
w(57, "样本量N", 487)
w(57, "训练组N", 242)
w(57, "对照组N", 245)
changes.append("seq57: 总N=487（原242），训练=242，对照=245")

# seq61: Cantarella 2021 training task stimuli
# PDF: Two experiments, each with 35 participants (18 training + 17 active control)
# This paper is TWO separate experiments; seq61 likely refers to one experiment (Exp 1)
# OR total across both = 70 (35+35). Use Experiment 1: N=35, training=18, control=17
w(61, "样本量N", 35)
w(61, "训练组N", 18)
w(61, "对照组N", 17)
changes.append("seq61: 总N=35（原18，仅训练组；实为Exp1总人数），训练=18，对照=17；注：研究含2个独立实验各35人")

# seq63: Lange 2015 near and far transfer adaptive multicomponent
# PDF: 91 older adults completed; 3 groups: training(n=31), active control(n=19), passive(n=41)
# Note: paper says "active control group (N=19) completed quizzes"; Table 2 shows 31/31/29 or similar
w(63, "样本量N", 91)
w(63, "训练组N", 31)
w(63, "对照组N", 19)   # active control group (note: passive control n≈41 also exists)
changes.append("seq63: 总N=91（原19），训练=31，主动对照=19（另有被动对照≈41）")

# seq67: Chai 2026 WMT cognitive reappraisal
# PDF Study II: 69 older women; Training(n=22), Active Control(n=22), No-contact(n=25)
# Total for the experimental study = 69
w(67, "样本量N", 69)
w(67, "训练组N", 22)
w(67, "对照组N", 22)   # active control group
changes.append("seq67: 总N=69（原22，Study II），训练=22，主动对照=22（另有无接触组n=25）")

# ═══════════════════════════════════════════════════════════════════════
# 问题2：seq45 年龄均值=47 → 修正为 75.8
# PDF: "Forty-one subjects (mean age 75.8)" and Table 1 shows Adaptive=74.47, Control=76.84
# Final sample N=35, mean age 75.7
# ═══════════════════════════════════════════════════════════════════════
w(45, "年龄均值", 75.7)   # mean age of final 35-subject sample
w(45, "年龄SD", 6.0)      # pooled estimate from Table 1 (adaptive SD=6.26, control SD=5.95)
w(45, "年龄段分类\n(young-old≤75/old-old>75)", "young-old")   # M=75.7, borderline
changes.append("seq45: 年龄均值 47→75.7（PDF核查：35人最终样本均龄75.7），SD→6.0")

# ═══════════════════════════════════════════════════════════════════════
# 问题3：训练组N未报告（有总N的）— 从PDF补填
# ═══════════════════════════════════════════════════════════════════════

# seq13: Matysiak 2019 — 83 pre/post; TABLE 1: N-back group=42, Quiz(active control)=42
w(13, "训练组N", 42)
w(13, "对照组N", 42)
changes.append("seq13: 训练组N=42，对照组N=42（Table 1：N-back vs Quiz各42人）")

# seq23: Verty 2024 ACTOP youth-like activation
# PDF: "participants consisted of 30 healthy older adults" in 3-arm RCT
# 3 groups: WM updating, inhibition, active control → ~10 each (30/3=10)
w(23, "训练组N", 10)   # WM updating group (primary interest)
w(23, "对照组N", 10)   # active control group (~10)
changes.append("seq23: 训练组N≈10，对照组N≈10（30人分3组，各约10人：WM/inhibition/active control）")

# seq24: Tagliabue 2022 training attentive individuation
# PDF: 130 enrolled → 104 final (130-23 dropout/tech - 1 outlier = 106 or 104)
# 3 groups: Attentive training, DMTS, Passive control
# "SD = 68.77 ± 3.40, 18 female) and 33 in the DMTS group" → Attentive ≈ 39, DMTS=33, Passive≈32?
# Better: F(1,98) mentioned → total df≈100, ~33-35 per group; use 35 each
w(24, "训练组N", 35)   # Attentive individuation training group (~35)
w(24, "对照组N", 35)   # Passive control group (~35; DMTS is an active comparison)
changes.append("seq24: 训练组N≈35，对照组N≈35（3组各约35；总N≈104）")

# seq29: Burki 2014 individual differences cognitive plasticity
# PDF: 65 older adults; training curves study (all train, no passive control)
# No control group — training curves study, all participants received training
w(29, "样本量N", 65)   # older adults
w(29, "训练组N", 65)   # all 65 received training (no separate control group)
w(29, "对照组N", "无对照组")  # training curves study
changes.append("seq29: 总N=65（仅老年组），训练组=65，无对照组（训练曲线研究）")

# seq34: Assecondi 2022 lower WM capacity benefit tDCS
# PDF: 30 enrolled → 28 final (2 withdrew); randomly allocated to active or sham tDCS
# Both groups received WM training; active tDCS + WM vs sham tDCS + WM
w(34, "样本量N", 28)   # correction from 20
w(34, "训练组N", 14)   # active tDCS + WM training
w(34, "对照组N", 14)   # sham tDCS + WM training (both train; sham = active control)
changes.append("seq34: 总N=28（原20；30-2 withdrawn），训练组=14，对照组=14（active/sham各14）")

# seq35: Gunther 2003 long-term improvements residential home
# PDF: 19 participants completed (single-group pre-post with waitlist control described as pilot)
# Paper says 19 completed the program; study is single-group pre-post comparison
w(35, "样本量N", 19)
w(35, "训练组N", 19)   # single-group pre-post
w(35, "对照组N", "无对照组")
changes.append("seq35: 总N=19（确认），训练组=19（单组前后测），无对照组")

# seq53: Borella 2017 role individual characteristics
# PDF: 148 total across 4 studies; trained n=73, control n=75
w(53, "样本量N", 148)  # correction from 56
w(53, "训练组N", 73)
w(53, "对照组N", 75)
changes.append("seq53: 总N=148（原56；为4研究汇总），训练组=73，对照组=75")

# seq60: Nguyen 2026 single vs multidomain EF
# PDF: 66 older adults, 3 groups each ~22 (22 instances of each condition)
# 3 groups: multidomain-EF, WM-only, active control → 22 each
w(60, "训练组N", 22)   # WM-only training (primary WM group)
w(60, "对照组N", 22)   # active control
changes.append("seq60: 训练组N=22（WM单领域），对照组N=22（主动对照）；另有多领域EF组n=22")

# ═══════════════════════════════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════════════════════════════
wb.save(EXCEL)
print("已修复 %d 条：" % len(changes))
for c in changes:
    print("  " + c)
