"""
Paper-01 数据提取表 v3 生成脚本
- 从 v2(51列) 迁移到 v3(55列)
- 新增3字段：失访率/完成率、统计方法、任务认知过程重叠度
- 效应量列拆分：是否报告(是/否) + Cohen's d值
- 修复：训练频率、年龄段分类计算写入
- 格式：深蓝表头、彩色说明行、冻结C3
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import os, re

BASE    = os.path.dirname(os.path.abspath(__file__))
V2_PATH = os.path.join(BASE, "数据_6_数据提取表_v2.xlsx")
V3_PATH = os.path.join(BASE, "数据_6_数据提取表_v3.xlsx")

# ══════════════════════════════════════════════════════════════
#  颜色 / 样式常量
# ══════════════════════════════════════════════════════════════
C_HEADER  = "2F5496"   # 深蓝 表头背景
C_ORIG    = "D9E1F2"   # 蓝灰 原有字段说明行
C_MUST    = "FFD7D7"   # 红   新增必须
C_SUGGEST = "FFF2CC"   # 黄   新增建议
C_ROB     = "E2EFDA"   # 绿   RoB相关
C_WHITE   = "FFFFFF"
C_DATA_ALT= "F7FAFF"   # 数据行隔行浅蓝

def pfill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font():
    return Font(bold=True, color="FFFFFF", size=10, name="微软雅黑")

def data_font():
    return Font(size=9, name="微软雅黑")

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

# ══════════════════════════════════════════════════════════════
#  v3 列定义（55列）
#  格式：(列名, 说明行分类, 说明行文字, 列宽)
#  说明行分类：orig/must/suggest/rob
# ══════════════════════════════════════════════════════════════
V3_COLS = [
    # ── 基本信息 ──────────────────────────────────────────────
    ("序号",                                             "orig",    "原有字段",  6),
    ("第一作者",                                         "orig",    "原有字段",  12),
    ("年份",                                             "orig",    "原有字段",  6),
    ("标题",                                             "orig",    "原有字段",  40),
    ("期刊",                                             "orig",    "原有字段",  20),
    ("国家",                                             "orig",    "原有字段",  10),
    ("样本来源\n(社区/大学/医院)",                        "orig",    "原有字段",  14),
    # ── 样本描述 ──────────────────────────────────────────────
    ("样本量N",                                          "orig",    "原有字段",  8),
    ("训练组N",                                          "must",    "新增必须",  8),
    ("对照组N",                                          "must",    "新增必须",  8),
    ("年龄均值",                                         "orig",    "原有字段",  8),
    ("年龄段分类\n(young-old≤75/old-old>75)",            "suggest", "新增建议",  14),
    ("年龄SD",                                           "orig",    "原有字段",  8),
    ("性别(%女)",                                        "orig",    "原有字段",  8),
    ("教育年限(年)",                                     "orig",    "原有字段",  10),
    ("认知筛查工具",                                     "orig",    "原有字段",  14),
    ("筛查分数",                                         "orig",    "原有字段",  8),
    # ── 训练设计 ──────────────────────────────────────────────
    ("训练类型\n(span/n-back/其他)",                     "orig",    "原有字段",  14),
    ("训练任务名称",                                     "orig",    "原有字段",  16),
    ("是否自适应\n(是/否)",                              "orig",    "原有字段",  10),
    ("训练总次数\n(sessions)",                           "orig",    "原有字段",  10),
    ("每次时长\n(分钟)",                                 "orig",    "原有字段",  10),
    ("失访率/完成率\n(%)",                               "must",    "新增必须",  12),   # ← 新增①
    ("训练周数",                                         "orig",    "原有字段",  8),
    ("是否主动对照\n(是/否)",                            "orig",    "原有字段",  10),
    ("对照组任务类型",                                   "must",    "新增必须",  16),
    ("结合干预类型\n(tDCS/TMS/药物/运动/无)",            "orig",    "原有字段",  16),
    ("统计方法\n(ANOVA/ANCOVA/LMM/其他)",               "suggest", "新增建议",  16),   # ← 新增②
    ("训练平台/软件",                                    "orig",    "原有字段",  14),
    ("监督方式\n(实验室/居家/混合)",                     "must",    "新增必须",  12),
    ("训练频率\n(次/周)",                                "must",    "新增必须",  10),
    # ── 结局 ──────────────────────────────────────────────────
    ("近迁移\n(是/否)",                                  "orig",    "原有字段",  8),
    ("近迁移结局变量",                                   "orig",    "原有字段",  20),
    ("远迁移\n(是/否)",                                  "orig",    "原有字段",  8),
    ("远迁移结局变量",                                   "orig",    "原有字段",  20),
    ("远迁移结局域\n(流体智力/EF/情景记忆/日常功能/其他)", "must",  "新增必须",  18),
    ("维持随访\n(是/否)",                                "orig",    "原有字段",  8),
    ("随访时间点(月)",                                   "orig",    "原有字段",  10),
    ("效应量是否报告\n(是/否)",                          "orig",    "原有字段",  12),   # ← 原字段改名
    ("Cohen's d值\n(有则填数值，无则留空)",               "orig",    "原有字段",  20),   # ← 新增（拆分）
    ("总体结论\n(正向/无/混合)",                         "orig",    "原有字段",  10),
    # ── 神经影像 ──────────────────────────────────────────────
    ("神经影像结局\n(是/否)",                            "orig",    "原有字段",  10),
    ("影像类型\n(fMRI/EEG/ERP/其他)",                   "orig",    "原有字段",  14),
    ("神经影像主要发现",                                 "orig",    "原有字段",  30),
    # ── 调节因素分析 ──────────────────────────────────────────
    ("基线WM水平\n(高/低/未报告)",                       "orig",    "原有字段",  12),
    ("显式调节效应检验\n(是/否)",                        "orig",    "原有字段",  12),
    ("调节效应方向\n(高>低/低>高/无/未检验)",            "orig",    "原有字段",  16),
    ("任务认知过程重叠度\n(高/中/低/未报告)",            "must",    "新增必须",  16),   # ← 新增③
    ("年龄亚组分析\n(是/否)",                            "orig",    "原有字段",  10),
    ("认知储备指标\n(教育/NART/未报告)",                 "orig",    "原有字段",  14),
    ("认知储备是否显著调节迁移\n(是/否/未检验)",         "suggest", "新增建议",  16),
    # ── 其他 ──────────────────────────────────────────────────
    ("发表状态\n(期刊/预印本)",                          "orig",    "原有字段",  10),
    ("备注",                                             "orig",    "原有字段",  24),
    ("⚠️RoB总体等级\n(低/有顾虑/高/未评估)",            "rob",     "RoB等级",   12),
    ("研究设计类型\n(RCT/准RCT/单组前后测/交叉)",        "must",    "新增必须",  16),
]
# 共55列
assert len(V3_COLS) == 55, f"列数={len(V3_COLS)}，应为55"

# v3列名 → 列号（1-based）
V3_HDR = {col[0]: i+1 for i, col in enumerate(V3_COLS)}

# ══════════════════════════════════════════════════════════════
#  v2 列名 → 列号映射
# ══════════════════════════════════════════════════════════════
wb2 = openpyxl.load_workbook(V2_PATH)
ws2 = wb2["数据提取"]

v2_hdr = {}
for col in range(1, ws2.max_column+1):
    v = ws2.cell(row=1, column=col).value
    if v:
        v2_hdr[str(v).strip()] = col

# v2数据行（第3行起）
v2_data_rows = []
for row in range(3, ws2.max_row+1):
    if ws2.cell(row=row, column=1).value is None:
        break
    v2_data_rows.append(row)

print(f"v2 数据行数：{len(v2_data_rows)}")

# ══════════════════════════════════════════════════════════════
#  v2列名 → v3列名 的映射
#  (v2列名, v3列名)，一一对应；特殊处理另行说明
# ══════════════════════════════════════════════════════════════
V2_TO_V3 = {
    "序号":                                              "序号",
    "第一作者":                                          "第一作者",
    "年份":                                              "年份",
    "标题":                                              "标题",
    "期刊":                                              "期刊",
    "国家":                                              "国家",
    "样本来源\n(社区/大学/医院)":                        "样本来源\n(社区/大学/医院)",
    "样本量N":                                           "样本量N",
    "训练组N":                                           "训练组N",
    "对照组N":                                           "对照组N",
    "年龄均值":                                          "年龄均值",
    # 年龄段分类：重新计算，不直接迁移
    "年龄SD":                                            "年龄SD",
    "性别(%女)":                                         "性别(%女)",
    "教育年限(年)":                                      "教育年限(年)",
    "认知筛查工具":                                      "认知筛查工具",
    "筛查分数":                                          "筛查分数",
    "训练类型\n(span/n-back/其他)":                      "训练类型\n(span/n-back/其他)",
    "训练任务名称":                                      "训练任务名称",
    "是否自适应\n(是/否)":                               "是否自适应\n(是/否)",
    "训练总次数\n(sessions)":                            "训练总次数\n(sessions)",
    "每次时长\n(分钟)":                                  "每次时长\n(分钟)",
    # 失访率：新增，留空
    "训练周数":                                          "训练周数",
    "是否主动对照\n(是/否)":                             "是否主动对照\n(是/否)",
    "对照组任务类型":                                    "对照组任务类型",
    "结合干预类型\n(tDCS/TMS/药物/运动/无)":             "结合干预类型\n(tDCS/TMS/药物/运动/无)",
    # 统计方法：新增，留空
    "训练平台/软件":                                     "训练平台/软件",
    "监督方式\n(实验室/居家/混合)":                      "监督方式\n(实验室/居家/混合)",
    # 训练频率：重新计算
    "近迁移\n(是/否)":                                   "近迁移\n(是/否)",
    "近迁移结局变量":                                    "近迁移结局变量",
    "远迁移\n(是/否)":                                   "远迁移\n(是/否)",
    "远迁移结局变量":                                    "远迁移结局变量",
    "远迁移结局域\n(流体智力/EF/情景记忆/日常功能/其他)": "远迁移结局域\n(流体智力/EF/情景记忆/日常功能/其他)",
    "维持随访\n(是/否)":                                 "维持随访\n(是/否)",
    "随访时间点(月)":                                    "随访时间点(月)",
    # 效应量：特殊处理（拆分）
    "总体结论\n(正向/无/混合)":                          "总体结论\n(正向/无/混合)",
    "神经影像结局\n(是/否)":                             "神经影像结局\n(是/否)",
    "影像类型\n(fMRI/EEG/ERP/其他)":                    "影像类型\n(fMRI/EEG/ERP/其他)",
    "神经影像主要发现":                                  "神经影像主要发现",
    "基线WM水平\n(高/低/未报告)":                        "基线WM水平\n(高/低/未报告)",
    "显式调节效应检验\n(是/否)":                         "显式调节效应检验\n(是/否)",
    "调节效应方向\n(高>低/低>高/无/未检验)":             "调节效应方向\n(高>低/低>高/无/未检验)",
    # 任务认知过程重叠度：新增，留空
    "年龄亚组分析\n(是/否)":                             "年龄亚组分析\n(是/否)",
    "认知储备指标\n(教育/NART/未报告)":                  "认知储备指标\n(教育/NART/未报告)",
    "认知储备是否显著调节迁移\n(是/否/未检验)":          "认知储备是否显著调节迁移\n(是/否/未检验)",
    "发表状态\n(期刊/预印本)":                           "发表状态\n(期刊/预印本)",
    "备注":                                              "备注",
    "⚠️RoB总体等级\n(低/有顾虑/高/未评估)":             "⚠️RoB总体等级\n(低/有顾虑/高/未评估)",
    "研究设计类型\n(RCT/准RCT/单组前后测/交叉)":         "研究设计类型\n(RCT/准RCT/单组前后测/交叉)",
}

# ══════════════════════════════════════════════════════════════
#  构建新工作簿
# ══════════════════════════════════════════════════════════════
wb3 = openpyxl.load_workbook(V2_PATH)   # 复制v2作为基础，保留RoB/字段说明sheet
# 删除旧的数据提取sheet，重建
del wb3["数据提取"]
if "核查报告" in wb3.sheetnames:
    del wb3["核查报告"]
ws3 = wb3.create_sheet("数据提取", 0)   # 插在最前面

# ── 行1：表头 ──────────────────────────────────────────────
for i, (col_name, cat, label, width) in enumerate(V3_COLS, 1):
    c = ws3.cell(row=1, column=i, value=col_name)
    c.fill      = pfill(C_HEADER)
    c.font      = hdr_font()
    c.alignment = center_align()
    c.border    = thin_border()
    ws3.column_dimensions[get_column_letter(i)].width = width

ws3.row_dimensions[1].height = 45

# ── 行2：颜色说明行 ────────────────────────────────────────
CAT_COLOR = {"orig": C_ORIG, "must": C_MUST, "suggest": C_SUGGEST, "rob": C_ROB}
for i, (col_name, cat, label, width) in enumerate(V3_COLS, 1):
    c = ws3.cell(row=2, column=i, value=label)
    c.fill      = pfill(CAT_COLOR[cat])
    c.font      = Font(size=8, color="444444", name="微软雅黑")
    c.alignment = center_align()
    c.border    = thin_border()

ws3.row_dimensions[2].height = 14

# ══════════════════════════════════════════════════════════════
#  迁移数据
# ══════════════════════════════════════════════════════════════
# 效应量列v2列号
v2_col_effect = v2_hdr.get("效应量\n(d或η²)")
# v3目标列号
v3_col_reported = V3_HDR["效应量是否报告\n(是/否)"]
v3_col_d        = V3_HDR["Cohen's d值\n(有则填数值，无则留空)"]
v3_col_freq     = V3_HDR["训练频率\n(次/周)"]
v3_col_age_cat  = V3_HDR["年龄段分类\n(young-old≤75/old-old>75)"]
v3_col_sessions = V3_HDR["训练总次数\n(sessions)"]
v3_col_weeks    = V3_HDR["训练周数"]
v3_col_age_mean = V3_HDR["年龄均值"]

# 统计
freq_ok = freq_skip = 0
age_young = age_old = age_miss = 0
effect_yes = effect_no = 0

for i_row, v2_row in enumerate(v2_data_rows):
    v3_row = i_row + 3   # v3第3行起

    # 隔行浅色背景
    row_bg = C_DATA_ALT if i_row % 2 == 1 else C_WHITE

    # ── 迁移常规字段 ──────────────────────────────────────
    for v2_name, v3_name in V2_TO_V3.items():
        if v2_name not in v2_hdr:
            continue
        val = ws2.cell(row=v2_row, column=v2_hdr[v2_name]).value
        v3_col = V3_HDR[v3_name]
        c = ws3.cell(row=v3_row, column=v3_col, value=val)
        c.font      = data_font()
        c.alignment = left_align() if v3_col >= 4 else center_align()
        c.border    = thin_border()
        c.fill      = pfill(row_bg)

    # ── 效应量拆分 ────────────────────────────────────────
    if v2_col_effect:
        raw = ws2.cell(row=v2_row, column=v2_col_effect).value
        raw_str = str(raw).strip() if raw else ""

        if not raw or raw_str == "" or raw_str == "未报告" or "Bayesian" in raw_str:
            reported = "否"
            d_val    = None
            effect_no += 1
        else:
            reported = "是"
            # 尽量提取数字
            nums = re.findall(r'[-+]?\d*\.?\d+', raw_str)
            d_val = raw_str   # 保留原始文本（格式复杂，不强行转数字）
            effect_yes += 1

        for col_i, val in [(v3_col_reported, reported), (v3_col_d, d_val)]:
            c = ws3.cell(row=v3_row, column=col_i, value=val)
            c.font = data_font(); c.alignment = center_align()
            c.border = thin_border(); c.fill = pfill(row_bg)

    # ── 修复：训练频率 ────────────────────────────────────
    try:
        sessions = float(ws3.cell(row=v3_row, column=v3_col_sessions).value)
        weeks    = float(ws3.cell(row=v3_row, column=v3_col_weeks).value)
        if weeks > 0:
            freq = round(sessions / weeks, 1)
            ws3.cell(row=v3_row, column=v3_col_freq).value = freq
            freq_ok += 1
        else:
            freq_skip += 1
    except (TypeError, ValueError):
        freq_skip += 1

    # ── 修复：年龄段分类 ──────────────────────────────────
    try:
        age = float(ws3.cell(row=v3_row, column=v3_col_age_mean).value)
        cat = "young-old" if age <= 75 else "old-old"
        ws3.cell(row=v3_row, column=v3_col_age_cat).value = cat
        if age <= 75: age_young += 1
        else: age_old += 1
    except (TypeError, ValueError):
        age_miss += 1

    # ── 新增空字段填格式（让空格也有边框背景）────────────
    new_cols = [
        V3_HDR["失访率/完成率\n(%)"],
        V3_HDR["统计方法\n(ANOVA/ANCOVA/LMM/其他)"],
        V3_HDR["任务认知过程重叠度\n(高/中/低/未报告)"],
    ]
    for col_i in new_cols:
        c = ws3.cell(row=v3_row, column=col_i)
        if c.value is None:
            c.value = None
        c.font = data_font(); c.border = thin_border(); c.fill = pfill(row_bg)
        c.alignment = center_align()

    ws3.row_dimensions[v3_row].height = 20

# ── 冻结 C3（前2列+前2行）────────────────────────────────
ws3.freeze_panes = "C3"

# ══════════════════════════════════════════════════════════════
#  更新 RoB sheet：同步序号/作者/年份
# ══════════════════════════════════════════════════════════════
ws_rob = wb3["偏倚风险评估_RoB"]
for i_row, v2_row in enumerate(v2_data_rows):
    v3_row = i_row + 3
    for col_i, v2_name in [(1, "序号"), (2, "第一作者"), (3, "年份")]:
        val = ws2.cell(row=v2_row, column=v2_hdr[v2_name]).value
        ws_rob.cell(row=v3_row, column=col_i).value = val

# ══════════════════════════════════════════════════════════════
#  保存
# ══════════════════════════════════════════════════════════════
wb3.save(V3_PATH)

print("=" * 55)
print(f"✅  v3 已保存：{V3_PATH}")
print(f"    列数：{len(V3_COLS)} 列（含说明行共{ws3.max_column}列确认）")
print(f"    数据行：{len(v2_data_rows)} 篇")
print(f"    训练频率：成功计算 {freq_ok} 篇，跳过 {freq_skip} 篇")
print(f"    年龄段：young-old={age_young} / old-old={age_old} / 缺失={age_miss}")
print(f"    效应量：有报告(是)={effect_yes} / 未报告(否)={effect_no}")
print("=" * 55)
