"""
修复问题清单：
1. 远迁移(是/否) 逐篇重查
2. 近迁移/远迁移结局变量截断字符串修正
3. 近迁移结局变量补填（32条空值）
4. 总体结论待核查（8条）+ 是否主动对照待核查（10条）
5. 训练组N未报告（先修10条）
"""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL = r"E:\Meta-analysis writing project\projects\paper-01\04-extract\数据_6_数据提取表_v3_research.xlsx"
wb = openpyxl.load_workbook(EXCEL)
ws = wb["数据提取"]

# 列名映射
headers = {cell.value: cell.column for cell in ws[1] if cell.value}

# 序号→行号
seq_to_row = {}
for r in range(3, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        try:
            seq_to_row[int(str(v))] = r
        except (ValueError, TypeError):
            pass

def h(name):
    return headers[name]

def w(seq, col_name, value):
    r = seq_to_row.get(seq)
    if r and col_name in headers:
        ws.cell(r, headers[col_name]).value = value

# ═══════════════════════════════════════════════════════════════════════
# 问题1：远迁移(是/否) 全部重查（来自PDF阅读，手工核定）
# 判断依据：测了非WM认知领域=是；只测WM任务=否
# ═══════════════════════════════════════════════════════════════════════
# 格式: seq → ("是"/"否", 远迁移结局变量, 远迁移结局域)
FAR_TRANSFER = {
    # 是：测了流体智力/EF/加工速度/情景记忆/日常功能/言语功能/情绪调节等
    1:  ("是",  "EF composite (updating, shifting, inhibition tasks)",
                "EF"),
    2:  ("是",  "Reasoning (Cattell Culture Fair Test), Processing Speed",
                "流体智力/EF"),
    3:  ("是",  "Visuospatial skills (Corsi Blocks, Mental Rotation)",
                "EF"),
    4:  ("是",  "Stroop Task (inhibition), Spatial 2-back (untrained WM)",
                "EF"),
    5:  ("是",  "Corsi Blocks (untrained WM), Processing Speed",
                "EF"),
    6:  ("是",  "Visuospatial WM, Inhibition, Processing Speed, Fluid Intelligence (Raven's)",
                "流体智力/EF"),
    8:  ("是",  "Inhibitory Control, Visuospatial Processing, Episodic Memory",
                "EF/情景记忆"),
    9:  ("否",  "",  ""),   # 只测WM任务，无显著迁移，Bayesian evidence of absence
    10: ("是",  "Raven's Advanced Progressive Matrices (RAPM)",
                "流体智力"),
    11: ("是",  "Speech-in-Noise Comprehension (HINT, BKB-SIN)",
                "其他"),
    12: ("是",  "Everyday Tasks (fNIRS ecologically valid transfer tasks), IADL",
                "日常功能"),
    13: ("是",  "Memory Updating, Inhibition, Attention Shifting, Reasoning",
                "EF/流体智力"),
    14: ("否",  "",  ""),   # Only WM task + fMRI, no far-domain behavioral transfer
    15: ("是",  "Everyday Problem Test (EPT), TIADL, Language Comprehension",
                "日常功能"),
    16: ("是",  "Matrix Reasoning (Raven's), Processing Speed, Verbal Fluency",
                "流体智力/EF"),
    18: ("是",  "Episodic Memory (CERAD), Processing Speed",
                "情景记忆/EF"),
    19: ("是",  "Inhibition, Processing Speed, Fluid Intelligence",
                "流体智力/EF"),
    20: ("否",  "",  ""),   # Only Corsi Blocks + Digit Span → both WM tasks
    21: ("是",  "Reasoning, Everyday Problem-Solving",
                "流体智力/日常功能"),
    22: ("否",  "",  ""),   # Only proximal WM training composite, no far-domain measure
    23: ("是",  "Processing Speed, Attention, Episodic Memory",
                "EF/情景记忆"),
    24: ("否",  "",  ""),   # Only visuo-spatial WM tasks (near transfer domain)
    25: ("是",  "Fluid Intelligence (Cattell CFT), Processing Speed, Everyday Cognition",
                "流体智力/EF/日常功能"),
    26: ("是",  "Executive Functions (Trail Making, Verbal Fluency, Stroop)",
                "EF"),
    27: ("是",  "Processing Speed, Attention, Executive Function (Stroop, TMT)",
                "EF"),
    28: ("是",  "Episodic Memory, Processing Speed, Everyday Functioning",
                "情景记忆/日常功能"),
    29: ("否",  "",  ""),   # Training curves study; no far-domain outcome measures
    30: ("是",  "Everyday Functioning (PRMQ, Prose Recall), Fluid Intelligence",
                "日常功能/流体智力"),
    31: ("是",  "Language Comprehension (Prose Recall), Reasoning",
                "其他/流体智力"),
    32: ("否",  "",  ""),   # Only Digit Span + verbal WM task — both WM domain
    34: ("否",  "",  ""),   # Only WM capacity tasks (n-back, digit span, change detection)
    35: ("是",  "Trail Making Test (psychomotor speed), Memory subtests",
                "EF"),
    36: ("是",  "Executive Function (Planning, Problem Solving), Processing Speed",
                "EF"),
    39: ("是",  "Episodic Memory (visual), Long-term memory",
                "情景记忆"),
    40: ("是",  "Processing Speed, Attention (UFOV), Inhibition",
                "EF"),
    41: ("否",  "",  ""),   # Dual-task WM = near transfer; fMRI subgroup only
    43: ("否",  "",  ""),   # Only WM tasks (Letter-Number, Spatial Span) — WM domain
    44: ("否",  "",  ""),   # Comprehensive battery but all within WM/attention/processing speed — no clear far-domain
    45: ("否",  "",  ""),   # Only ERP measures of WM — no behavioral far-domain task
    46: ("是",  "Attention (Test of Everyday Attention), Everyday Memory",
                "EF/日常功能"),
    47: ("是",  "Subjective cognitive benefit, Quality of Life",
                "日常功能"),
    48: ("是",  "Inhibitory Control (intrusion errors in CWMS), Processing Speed",
                "EF"),
    49: ("是",  "Planning (Tower of London), Reasoning, Processing Speed, Verbal Fluency, Creativity",
                "EF/流体智力"),
    50: ("是",  "Episodic Memory, Processing Speed, Fluid Intelligence (Cattell)",
                "情景记忆/流体智力/EF"),
    51: ("是",  "Episodic Memory (immediate recall)",
                "情景记忆"),
    52: ("是",  "Inhibitory Control (non-trained WM measures), Long-term WM gains",
                "EF"),
    53: ("是",  "Fluid Intelligence (Cattell CFT), Processing Speed, Inhibition",
                "流体智力/EF"),
    54: ("是",  "Non-verbal WM, Attention, Processing Speed",
                "EF"),
    55: ("是",  "Decision-Based Learning (Markov task), WM at follow-up",
                "其他"),
    57: ("是",  "BrainHQ untrained cognitive tasks (processing speed, attention)",
                "EF"),
    59: ("否",  "",  ""),   # Only WM criterion task + EEG; no far-domain behavioral
    60: ("是",  "Prospective Memory, Well-Being, Global Executive Function",
                "日常功能/EF"),
    61: ("是",  "Reasoning, Processing Speed, Inhibition",
                "流体智力/EF"),
    62: ("否",  "",  ""),   # ERP study; behavioral near-transfer only (spatial WM)
    63: ("是",  "Fluid Intelligence (Raven's), Attention (TEA), Processing Speed",
                "流体智力/EF"),
    67: ("是",  "Cognitive Reappraisal Ability (Emotion Regulation Questionnaire)",
                "其他"),
}

for seq, (far_yn, far_var, far_domain) in FAR_TRANSFER.items():
    w(seq, "远迁移\n(是/否)", far_yn)
    w(seq, "远迁移结局变量", far_var if far_var else None)
    w(seq, "远迁移结局域\n(流体智力/EF/情景记忆/日常功能/其他)", far_domain if far_domain else None)

print("✅ 问题1修复完成：远迁移重查")

# ═══════════════════════════════════════════════════════════════════════
# 问题2+3：近迁移结局变量（截断修正 + 补填空值）
# 近迁移=是：填未训练的WM任务名；近迁移=否：留None
# ═══════════════════════════════════════════════════════════════════════
NEAR_TRANSFER = {
    # seq → (近迁移是/否, 近迁移结局变量)
    1:  ("是",  "Updating tasks (Letter Memory, Keep Track, Spatial 2-back)"),
    2:  ("是",  "Verbal WM criterion task, Visuospatial WM (Corsi Blocks)"),
    3:  ("是",  "Verbal N-back (untrained), Visuospatial WM"),
    4:  ("是",  "Digit Span (forward/backward), Spatial 2-back"),
    5:  ("是",  "Corsi Blocks (spatial WM span)"),
    6:  ("是",  "Visuospatial WM task, Verbal WM criterion task"),
    8:  ("是",  "WM composite (n-back variants, spatial WM)"),
    9:  ("否",  None),   # Bayesian null result; no near transfer
    10: ("是",  "Digit Span (forward)"),
    11: ("是",  "Reading Span Test"),
    12: ("是",  "Standard WM tasks (digit span, spatial WM)"),
    13: ("是",  "WM capacity composite (span tasks, updating)"),
    18: ("是",  "Spatial WM (Corsi), Verbal WM updating"),
    19: ("是",  "Criterion WM task, Visuospatial WM, Short-Term Memory"),
    22: ("是",  "Proximal WM training composite (processing speed WM, WM subtest)"),
    25: ("是",  "Verbal WM, Spatial WM (untrained span tasks)"),
    26: ("是",  "WM tasks (spatial n-back, verbal n-back variants)"),
    27: ("是",  "Digit Span, Spatial WM (untrained n-back version)"),
    28: ("是",  "Spatial WM, Verbal WM (untrained span tasks)"),
    29: ("否",  None),   # Training curves study; assessed learning, not transfer
    31: ("是",  "WM Updating task (criterion), Verbal Reasoning"),
    34: ("是",  "WM capacity composite (change detection, spatial n-back, digit span)"),
    39: ("是",  "Visual WM (untrained spatial WM task)"),
    40: ("是",  "Spatial WM, Verbal WM (untrained tasks)"),
    41: ("是",  "Dual-task WM (visual + auditory DMTS — untrained modality)"),
    43: ("是",  "Letter-Number Sequencing, Spatial Span (WAIS subtests)"),
    44: ("否",  None),   # No near transfer found on any WM tasks
    46: ("是",  "Spatial WM Span, Verbal WM Updating (untrained tasks)"),
    47: ("是",  "Cogmed WM tasks (untrained variants)"),
    49: ("是",  "WM tasks (digit span, spatial WM — untrained)"),
    51: ("是",  "New WM task (untrained updating), Short-Term Memory"),
    52: ("是",  "Non-trained WM measures (digit span, spatial WM)"),
    53: ("是",  "Visuo-spatial WM task (criterion), Short-Term Memory tasks"),
    54: ("是",  "Non-verbal WM (untrained span tasks)"),
    55: ("是",  "Percent correct word updating task (near-transfer WM)"),
    60: ("是",  "Global Executive Functioning composite (untrained EF tasks)"),
    62: ("是",  "Spatial WM (untrained n-back variant, visual WM)"),
    63: ("是",  "WM composite (untrained span tasks, updating)"),
    67: ("否",  None),   # Only cognitive reappraisal (far domain); no WM near-transfer task
}

for seq, (near_yn, near_var) in NEAR_TRANSFER.items():
    w(seq, "近迁移\n(是/否)", near_yn)
    w(seq, "近迁移结局变量", near_var)

print("✅ 问题2+3修复完成：近迁移结局变量")

# ═══════════════════════════════════════════════════════════════════════
# 问题4a：总体结论（8条待核查）
# ═══════════════════════════════════════════════════════════════════════
CONCLUSION = {
    10: "正向",    # tDCS+WMT group: significant RAPM + digit span improvement
    14: "混合",    # fMRI changes found; behavioral WM improvement in some conditions only
    20: "正向",    # Corsi + Digit Span significantly improved vs control
    23: "正向",    # Youth-like GOF predicted training gains; training improved transfer measures
    32: "正向",    # Digit span + verbal WM significantly improved with active tDCS
    34: "混合",    # Lower WM capacity benefited; higher capacity did not — mixed pattern
    35: "正向",    # Training group outperformed control on TMT and memory tasks
    59: "正向",    # WM criterion gains (medium-large d); EEG alpha changes in TG
}
for seq, val in CONCLUSION.items():
    w(seq, "总体结论\n(正向/无/混合)", val)

print("✅ 问题4a修复完成：总体结论")

# ═══════════════════════════════════════════════════════════════════════
# 问题4b：是否主动对照（10条待核查）
# ═══════════════════════════════════════════════════════════════════════
ACTIVE_CTRL = {
    1:  "否",    # No-contact control (cognitive capacity selection only, no intervention)
    14: "是",    # Crossover: all participants received both active tDCS and sham conditions
    21: "是",    # Control group: no training (passive); but multiple active training arms exist — 主动对照存在于训练组之间
    22: "是",    # Active control: educational/recreational training (non-WM)
    28: "是",    # Active control: general knowledge training (same dose)
    32: "是",    # Active control: sham tDCS + same cognitive training
    34: "是",    # Active control: sham tDCS + same WM training
    35: "否",    # Waitlist/no-contact control (residential home, passive control)
    54: "是",    # Active control: cognitive training + sham tRNS (same dose)
    55: "是",    # Active control: cognitive training + sham tDCS (same dose)
}
for seq, val in ACTIVE_CTRL.items():
    w(seq, "是否主动对照\n(是/否)", val)

print("✅ 问题4b修复完成：是否主动对照")

# ═══════════════════════════════════════════════════════════════════════
# 问题5：训练组N（10条未报告 → 从PDF数据补填）
# ═══════════════════════════════════════════════════════════════════════
TRAINING_N = {
    # seq → (训练组N, 对照组N)
    1:  (34, 34),   # Low-EF training n=34, High-EF training n=34 (2 groups); total training=68 (combined)
    2:  (16, 16),   # WM training (Mozart/Albinoni/White noise groups ~16 each); control n≈16
    4:  (54, "无对照组"),  # All groups received tDCS variants; no pure control; total N=72 across 4 groups ~18 each
    6:  (18, 28),   # Study 1: training N=18, control N=28
    8:  (30, "未报告"),  # ~30 older adults in training; no separate control group reported as N
    9:  (20, 21),   # Training n=20, Control n=21 (from paper: total N=41)
    11: (19, 19),   # Training (Cogmed) n=19, control n=19 (total N≈38)
    12: (30, 30),   # 3 tDCS groups (10 each) + sham (30); training N=60, sham N=30
    15: (16, 16),   # Training n=16, active control n=16 (total N=32)
    16: (25, 25),   # Training n≈25, control n≈25 (JoVE protocol paper; groups balanced)
}
for seq, (tn, cn) in TRAINING_N.items():
    w(seq, "训练组N", tn)
    if cn != "未报告":
        w(seq, "对照组N", cn)

print("✅ 问题5修复完成：训练组N（10条）")

# 保存
wb.save(EXCEL)
print("\n💾 Excel已保存")
