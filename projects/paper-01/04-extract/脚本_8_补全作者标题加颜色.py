"""
补全 research.xlsx 中的第一作者全名、标题，
并对"数据填入状态"列上色：已填入=绿，未填入=红
"""
import openpyxl
from openpyxl.styles import PatternFill

EXCEL = r"E:\Meta-analysis writing project\projects\paper-01\04-extract\数据_6_数据提取表_v3_research.xlsx"

# ── 人工整理的作者全名 + 标题（全部从PDF第一页提取）────────────────────────────
# 格式: seq → (第一作者全名, 标题全称)
DATA = {
    1:  ("Lommen, M.J.J.",
         "The effect of cognitive training in older adults: be aware of CRUNCH"),
    2:  ("Borella, E.",
         "Is working memory training in older adults sensitive to music?"),
    3:  ("Salminen, T.",
         "Age-specific differences of dual n-back training"),
    4:  ("Tays, G.D.",
         "Longitudinal neurostimulation in older adults improves working memory"),
    5:  ("Vesterinen, H.M.",
         "N-back training and transfer effects revealed by behavioral responses and EEG"),
    6:  ("Brum, P.S.",
         "Verbal working memory training in older adults: an investigation of dose response effects"),
    8:  ("Wang, P.",
         "Gamified Working Memory Intervention Enhances Prefrontal Neurocognitive Plasticity During Aging"),
    9:  ("Guye, S.",
         "Working Memory Training in Older Adults: Bayesian Evidence Supporting the Absence of Transfer"),
    10: ("Teixeira-Santos, A.C.",
         "Working Memory Training Coupled With Transcranial Direct Current Stimulation in Older Adults: A Randomized Controlled Experiment"),
    11: ("Wayne, R.V.",
         "Working Memory Training and Speech in Noise Comprehension in Older Adults"),
    12: ("Stephens, J.A.",
         "Older Adults Improve on Everyday Tasks after Working Memory Training and Neurostimulation"),
    13: ("Matysiak, O.",
         "Working Memory Capacity as a Predictor of Cognitive Training Efficacy in the Elderly Population"),
    14: ("Šimko, P.",
         "Cognitive Aftereffects of Acute tDCS Coupled with Cognitive Training: An fMRI Study in Healthy Seniors"),
    15: ("Borella, E.",
         "Improving Everyday Functioning in the Old-Old with Working Memory Training"),
    16: ("Matysiak, O.",
         "Working Memory Training for Older Participants: A Control Group Training Regimen and Initial Intellectual Functioning Assessment"),
    18: ("Heinzel, S.",
         "Working memory training improvements and gains in non-trained cognitive tasks in young and older adults"),
    19: ("Borella, E.",
         "Working Memory Training in Old Age: An Examination of Transfer and Maintenance Effects"),
    20: ("Ghavidel, F.",
         "Feasibility of using a computer-assisted working memory training program for healthy older women"),
    21: ("Sun, S.",
         "The Effects of Cognitive Training on Cognitive Abilities and Everyday Function: A 10-Week Randomized Controlled Trial"),
    22: ("Boutzoukas, E.M.",
         "Higher white matter hyperintensity load adversely affects pre-post proximal cognitive training performance in healthy older adults"),
    23: ("Verty, L.V.",
         "Youth-like brain activation linked with greater cognitive training gains in older adults: Insights from the ACTOP study"),
    24: ("Tagliabue, C.F.",
         "Training attentive individuation leads to visuo-spatial working memory improvement in low-performing older adults: An online study"),
    25: ("Zinke, K.",
         "Working Memory Training and Transfer in Older Adults: Effects of Age, Baseline Performance, and Training Gains"),
    26: ("Mičič, S.",
         "The Impact of Working Memory Training on Cognitive Abilities in Older Adults: The Role of Cognitive Reserve"),
    27: ("Zając-Lamparska, L.",
         "Limited training and transfer effects in older and young adults who participated in 12 sessions of process-based working memory training"),
    28: ("Jaeggi, S.M.",
         "Investigating the effects of spacing on working memory training outcome – a randomized controlled multi-site trial in older adults"),
    29: ("Bürki, C.N.",
         "Individual differences in cognitive plasticity: an investigation of training curves in younger and older adults"),
    30: ("Cantarella, A.",
         "Benefits in tasks related to everyday life competences after a working memory training in older adults"),
    31: ("Carretti, B.",
         "Gains in language comprehension relating to working memory training in healthy older adults"),
    32: ("Heo, S.",
         "Long-term effects of transcranial direct current stimulation combined with computer-assisted cognitive training in healthy older adults"),
    34: ("Assecondi, S.",
         "Older adults with lower working memory capacity benefit from transcranial direct current stimulation when combined with working memory training"),
    35: ("Günther, V.K.",
         "Long-term improvements in cognitive performance through computer-assisted cognitive training: A pilot study in a residential home for older people"),
    36: ("Shatil, E.",
         "Novel Television-Based Cognitive Training Improves Working Memory and Executive Function"),
    39: ("Buschkuehl, M.",
         "Impact of Working Memory Training on Memory Performance in Old-Old Adults"),
    40: ("Booth, S.J.",
         "Experimental investigation of training schedule on home-based working memory training in healthy older adults"),
    41: ("Heinzel, S.",
         "Transfer Effects to a Multimodal Dual-Task after Working Memory Training and Associated Neural Correlates in Older Adults"),
    43: ("Stephens, J.A.",
         "Task demands, tDCS intensity, and the COMT val158met polymorphism impact tDCS-linked working memory training gains"),
    44: ("Sutton, E.",
         "Practice makes perfect, but to what end? Computerised brain training has limited effects on cognition and mood in healthy older adults"),
    45: ("Tusch, E.S.",
         "Changes in Neural Activity Underlying Working Memory after Computerized Cognitive Training in Older Adults"),
    46: ("McAvinue, L.P.",
         "An evaluation of a working memory training scheme in older adults"),
    47: ("Goghari, V.M.",
         "Self-Perceived Benefits of Cognitive Training in Healthy Older Adults"),
    48: ("Borella, E.",
         "Short- and long-term cognitive and electrophysiological effects of a brief working memory training in older adults: a pilot study"),
    49: ("Goghari, V.M.",
         "Comparison of Cognitive Change after Working Memory Training and Logic and Planning Training in Healthy Older Adults"),
    50: ("Zinke, K.",
         "Potentials and Limits of Plasticity Induced by Working Memory Training in Old-Old Age"),
    51: ("Basak, C.",
         "To Switch or Not to Switch: Role of Cognitive Control in Working Memory Training in Older Adults"),
    52: ("Jaeggi, S.M.",
         "EngAge – A metacognitive intervention to supplement working memory training: A feasibility study in older adults"),
    53: ("Borella, E.",
         "Working Memory Training for Healthy Older Adults: The Role of Individual Characteristics in Explaining Short- and Long-Term Training Gains"),
    54: ("Brambilla, M.",
         "The Effect of Transcranial Random Noise Stimulation on Cognitive Training Outcome in Healthy Aging"),
    55: ("Antonenko, D.",
         "Randomized trial of cognitive training and brain stimulation in non-demented older adults"),
    57: ("Zelinski, E.M.",
         "Evaluating the relationship between change in performance on training tasks and on untrained outcomes"),
    59: ("Spironelli, C.",
         "Working Memory Training and Cortical Arousal in Healthy Older Adults: A Resting-State EEG Pilot Study"),
    60: ("Nguyen, L.",
         "A comparison of single-domain and multidomain executive functions cognitive training for enhancing cognition and well-being in older adults"),
    61: ("Cantarella, A.",
         "The influence of training task stimuli on transfer effects of working memory training in aging"),
    62: ("Pergher, V.",
         "Impact of strategy use during N-Back training in older adults"),
    63: ("Lange, S.",
         "Experimental Evaluation of Near- and Far-Transfer Effects of an Adaptive Multicomponent Working Memory Training"),
    67: ("Chai, Q.",
         "The Impact of Working Memory Training on Cognitive Reappraisal Ability Among Older Women"),
}

# 颜色定义
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ── 打开 Excel ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)
ws = wb["数据提取"]

# 列名映射
headers = {cell.value: cell.column for cell in ws[1] if cell.value}
col_author  = headers["第一作者"]
col_title   = headers["标题"]
col_filled  = headers["数据填入状态\n(已填入/未填入)"]

# 序号 → 行号映射
seq_to_row = {}
for r in range(3, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is not None:
        try:
            seq_to_row[int(v)] = r
        except (ValueError, TypeError):
            pass

# ── 写入数据 + 上色 ───────────────────────────────────────────────────────
updated = 0
for seq, (author, title) in DATA.items():
    row = seq_to_row.get(seq)
    if not row:
        print(f"  [WARN] seq{seq} 行未找到")
        continue
    ws.cell(row, col_author).value = author
    ws.cell(row, col_title).value  = title
    updated += 1

# 所有行上色（数据填入状态列）
colored = 0
for r in range(3, ws.max_row + 1):
    v = ws.cell(r, 1).value
    if v is None:
        continue
    status_cell = ws.cell(r, col_filled)
    if status_cell.value == "已填入":
        status_cell.fill = GREEN
    else:
        status_cell.fill = RED
    colored += 1

wb.save(EXCEL)
print(f"✅ 完成：更新 {updated} 篇作者/标题，{colored} 行上色")
print(f"   绿=已填入，红=未填入")
