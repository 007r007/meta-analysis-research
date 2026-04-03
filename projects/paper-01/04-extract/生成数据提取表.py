import openpyxl, os, sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

# ── 读取55篇纳入文献 ──
src = openpyxl.load_workbook('E:/Meta-analysis writing project/projects/paper-01/03-screen/数据_5_第三轮全文筛选.xlsx')
ws_src = src['第三轮全文筛选']
included = []
for r in range(2, ws_src.max_row + 1):
    if ws_src.cell(r, 8).value == '纳入':
        included.append({
            'seq':    len(included) + 1,
            'author': ws_src.cell(r, 4).value,
            'year':   ws_src.cell(r, 3).value,
            'title':  ws_src.cell(r, 5).value,
            'journal':ws_src.cell(r, 6).value,
        })
print(f'纳入文献：{len(included)} 篇')

# ── 样式 ──
thin  = Side(style='thin',   color='FFB8CCE4')
thick = Side(style='medium', color='FF7F9EB5')
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
bd_left = Border(left=thick, right=thin, top=thin, bottom=thin)

def fill(hex6):
    return PatternFill('solid', fgColor='FF' + hex6)

# ── 列定义：(label, width, group) ──
COLS = [
    # A-E 文献标识（深灰）
    ('序号',                              4,  'ID'),
    ('第一作者',                          11, 'ID'),
    ('年份',                               5, 'ID'),
    ('标题',                              44, 'ID'),
    ('期刊',                              22, 'ID'),
    # F-N 基本信息（蓝）
    ('国家',                               9, 'INFO'),
    ('样本来源\n（社区/大学/医院）',       12, 'INFO'),
    ('样本量 N',                            8, 'INFO'),
    ('年龄均值',                            9, 'INFO'),
    ('年龄 SD',                             8, 'INFO'),
    ('性别（%女）',                         9, 'INFO'),
    ('教育年限（年）',                     10, 'INFO'),
    ('认知筛查工具',                       12, 'INFO'),
    ('筛查分数',                            9, 'INFO'),
    # O-W 训练特征（绿）
    ('训练类型\n（span/n-back/其他）',     12, 'TRAIN'),
    ('训练任务名称',                       15, 'TRAIN'),
    ('是否自适应\n（是/否）',               9, 'TRAIN'),
    ('训练总次数\n（sessions）',            9, 'TRAIN'),
    ('每次时长\n（分钟）',                  9, 'TRAIN'),
    ('训练周数',                            8, 'TRAIN'),
    ('是否有主动\n对照组（是/否）',         9, 'TRAIN'),
    ('结合干预类型\n（tDCS/TMS/药物/运动/无）', 14, 'TRAIN'),
    ('训练平台/软件',                      13, 'TRAIN'),
    # X-AE 迁移结局（橙）
    ('是否报告近迁移\n（是/否）',           9, 'TRANS'),
    ('近迁移结局变量',                     18, 'TRANS'),
    ('是否报告远迁移\n（是/否）',           9, 'TRANS'),
    ('远迁移结局变量',                     18, 'TRANS'),
    ('是否有维持随访\n（是/否）',           9, 'TRANS'),
    ('随访时间点（月）',                   10, 'TRANS'),
    ("效应量\n（Cohen's d 或 η²）",        11, 'TRANS'),
    ('总体结论\n（正向/无迁移/混合）',     12, 'TRANS'),
    # AF-AH 神经影像（紫）
    ('是否有神经\n影像结局（是/否）',       9, 'NEURO'),
    ('影像类型\n（fMRI/EEG/ERP/其他）',   12, 'NEURO'),
    ('主要发现（简述）',                   22, 'NEURO'),
    # AI-AN 调节因素（黄）
    ('基线WM水平\n（高/低/未报告）',       11, 'MOD'),
    ('是否显式检验\n调节效应（是/否）',    10, 'MOD'),
    ('调节效应方向\n（高>低/低>高/无/未检验）', 13, 'MOD'),
    ('年龄亚组分析\n（是/否）',             9, 'MOD'),
    ('认知储备指标\n（教育/NART/未报告）', 13, 'MOD'),
    ('发表状态\n（期刊/预印本）',          10, 'MOD'),
    # AO 备注
    ('备注',                              22, 'NOTE'),
]

# 分组：表头背景色、数据行底色、分组标签
GROUP = {
    'ID':    ('2F3F4F', 'F2F2F2', '文献标识'),
    'INFO':  ('1F6091', 'DEEAF1', '研究基本信息'),
    'TRAIN': ('375623', 'E2EFDA', '训练特征'),
    'TRANS': ('843C0C', 'FCE4D6', '迁移结局'),
    'NEURO': ('5B2D8E', 'EAD1F5', '神经影像结局'),
    'MOD':   ('7D6608', 'FFF2CC', '调节因素'),
    'NOTE':  ('404040', 'F2F2F2', '备注'),
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '数据提取'

# ── 第1行：分组标签（合并单元格）──
group_ranges = []
cur_group = COLS[0][2]
start_col = 1
for ci, (_, _, g) in enumerate(COLS, 1):
    if g != cur_group or ci == len(COLS):
        end_col = ci if g != cur_group else ci
        if g == cur_group and ci == len(COLS):
            end_col = ci
        group_ranges.append((cur_group, start_col, end_col - (1 if g != cur_group else 0)))
        cur_group = g
        start_col = ci

# 重新扫描分组范围
group_ranges = []
i = 0
while i < len(COLS):
    g = COLS[i][2]
    j = i
    while j < len(COLS) and COLS[j][2] == g:
        j += 1
    group_ranges.append((g, i+1, j))
    i = j

for g, c1, c2 in group_ranges:
    hdr_bg, _, label = GROUP[g]
    cell = ws.cell(1, c1, label)
    cell.fill = fill(hdr_bg)
    cell.font = Font(bold=True, color='FFFFFFFF', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    if c2 > c1:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
ws.row_dimensions[1].height = 18

# ── 第2行：列表头 ──
for ci, (label, width, g) in enumerate(COLS, 1):
    hdr_bg = GROUP[g][0]
    c = ws.cell(2, ci, label)
    c.fill = fill(hdr_bg)
    c.font = Font(bold=True, color='FFFFFFFF', size=9)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bd
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[2].height = 44

# ── 数据行（3 起）──
for ri, doc in enumerate(included, 3):
    for ci, (_, _, g) in enumerate(COLS, 1):
        data_bg = GROUP[g][1]
        c = ws.cell(ri, ci)
        if   ci == 1: c.value = doc['seq']
        elif ci == 2: c.value = doc['author']
        elif ci == 3: c.value = doc['year']
        elif ci == 4: c.value = doc['title']
        elif ci == 5: c.value = doc['journal']
        c.fill = fill(data_bg)
        c.font = Font(size=9)
        wrap = ci == 4
        halign = 'center' if ci in (1,2,3) else 'left'
        c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
        c.border = bd
    ws.row_dimensions[ri].height = 30

last_row = 2 + len(included)

# ── 数据验证 ──
def dv(col_letter, formula):
    d = DataValidation(type='list', formula1=formula,
                       allow_blank=True, showDropDown=False)
    d.sqref = f'{col_letter}3:{col_letter}{last_row}'
    ws.add_data_validation(d)

dv('G',  '"社区,大学,医院,混合"')
dv('O',  '"span,n-back,mixed,其他"')
dv('Q',  '"是,否"')
dv('U',  '"是,否"')
dv('V',  '"tDCS,TMS,药物,运动,无"')
dv('X',  '"是,否"')
dv('Z',  '"是,否"')
dv('AB', '"是,否"')
dv('AE', '"正向迁移,无迁移,混合"')
dv('AF', '"是,否"')
dv('AG', '"fMRI,EEG,ERP,fMRI+EEG,其他"')
dv('AI', '"高,低,未报告"')
dv('AJ', '"是,否"')
dv('AK', '"高>低,低>高,无,未检验"')
dv('AL', '"是,否"')
dv('AM', '"教育年限,NART,词汇测试,其他,未报告"')
dv('AN', '"期刊,预印本"')

ws.freeze_panes = 'F3'
ws.auto_filter.ref = f'A2:{get_column_letter(len(COLS))}{last_row}'

# ════════════════════════════════════════
# Sheet 2: 编码说明
# ════════════════════════════════════════
ws2 = wb.create_sheet('编码说明')
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 48
ws2.column_dimensions['D'].width = 28

title = ws2.cell(1, 1, 'Paper-01 数据提取编码说明')
title.font = Font(bold=True, size=12, color='FF1F3864')
ws2.merge_cells('A1:D1')
ws2.row_dimensions[1].height = 24

rows = [
    ('列',   '字段',                '填写说明',                                        '示例',             True,  None),
    ('',     '── 研究基本信息 ──',  '',                                                '',                 True,  '1F6091'),
    ('F',    '国家',                '研究实施国，英文缩写',                             'USA',              False, None),
    ('G',    '样本来源',            '社区招募/大学/医院/混合',                          '社区',             False, None),
    ('H',    '样本量 N',            '完成训练的总样本量（训练+对照合计）',              '48',               False, None),
    ('I',    '年龄均值',            '训练组年龄均值，精确0.1',                         '68.3',             False, None),
    ('J',    '年龄 SD',             '训练组年龄标准差',                                '5.2',              False, None),
    ('K',    '性别（%女）',         '训练组女性比例，整数',                             '62',               False, None),
    ('L',    '教育年限',            '平均受教育年数',                                  '13.5',             False, None),
    ('M',    '认知筛查工具',        '如MMSE/MoCA/MOCA；无则填"未报告"',               'MMSE',             False, None),
    ('N',    '筛查分数',            '均值，如未报告填"未报告"',                        '28.4',             False, None),
    ('',     '── 训练特征 ──',      '',                                                '',                 True,  '375623'),
    ('O',    '训练类型',            'span=广度任务; n-back; mixed=多任务混合',        'n-back',           False, None),
    ('P',    '任务名称',            '具体任务名，多个用"/"分隔',                       'dual n-back',      False, None),
    ('Q',    '是否自适应',          '难度随表现自动调整则"是"',                        '是',               False, None),
    ('R',    '训练总次数',          '总session数',                                     '20',               False, None),
    ('S',    '每次时长',            '分钟，范围用"-"',                                 '30',               False, None),
    ('T',    '训练周数',            '持续周数',                                        '4',                False, None),
    ('U',    '是否主动对照',        '有其他认知训练对照组则"是"',                      '是',               False, None),
    ('V',    '结合干预类型',        'tDCS/TMS/药物/运动/无',                           'tDCS',             False, None),
    ('W',    '训练平台/软件',       '软件或平台名称；无则填"自制"',                    'Cogmed',           False, None),
    ('',     '── 迁移结局 ──',      '',                                                '',                 True,  '843C0C'),
    ('X',    '是否报告近迁移',      '有未训练WM任务作为结局则"是"',                    '是',               False, None),
    ('Y',    '近迁移结局变量',      '具体测量任务，多个用";"分隔',                     'Reading span',     False, None),
    ('Z',    '是否报告远迁移',      '有非WM认知能力测量则"是"',                        '是',               False, None),
    ('AA',   '远迁移结局变量',      '具体测量任务，多个用";"分隔',                     'Raven; Stroop',    False, None),
    ('AB',   '是否有维持随访',      '训练结束后有随访则"是"',                          '是',               False, None),
    ('AC',   '随访时间点',          '月数，多个用";"',                                 '3;6',              False, None),
    ('AD',   '效应量',              "Cohen's d 或 η²，注明类型；无填'未报告'",        'd=0.45',           False, None),
    ('AE',   '总体结论',            '正向迁移=≥1项显著; 无迁移=均不显著; 混合=部分',  '正向迁移',         False, None),
    ('',     '── 神经影像 ──',      '',                                                '',                 True,  '5B2D8E'),
    ('AF',   '是否有神经影像',      '有fMRI/EEG/ERP等结局则"是"',                     '是',               False, None),
    ('AG',   '影像类型',            'fMRI/EEG/ERP/fMRI+EEG/其他',                     'EEG',              False, None),
    ('AH',   '主要发现',            '一句话简述，如"前额叶激活减少"',                  '前额叶激活减少',   False, None),
    ('',     '── 调节因素 ──',      '',                                                '',                 True,  '7D6608'),
    ('AI',   '基线WM水平',          '文章是否按高/低WM分组报告；未分析填"未报告"',    '低',               False, None),
    ('AJ',   '是否显式检验调节',    '是否做调节/交互效应统计检验',                     '是',               False, None),
    ('AK',   '调节效应方向',        '高>低=高WM获益更多; 低>高=低WM获益更多',         '低>高',            False, None),
    ('AL',   '年龄亚组分析',        '是否做年龄相关亚组或连续变量分析',               '否',               False, None),
    ('AM',   '认知储备指标',        '教育年限/NART/词汇测试/其他/未报告',             '教育年限',         False, None),
    ('AN',   '发表状态',            '期刊正式发表或预印本',                            '期刊',             False, None),
    ('AO',   '备注',                '特殊样本、方法学问题、EEG/fMRI补充说明等',       'EEG同步测量',      False, None),
]

thin2 = Side(style='thin', color='FFCCCCCC')
bd2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

for ri, (col, field, desc, ex, is_hdr, grp_bg) in enumerate(rows, 2):
    vals = [col, field, desc, ex]
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(ri, ci, v)
        if grp_bg:
            c.fill = fill(grp_bg)
            c.font = Font(bold=True, color='FFFFFFFF', size=9)
        else:
            c.font = Font(bold=is_hdr, size=9)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = bd2
    ws2.row_dimensions[ri].height = 15 if not (is_hdr or grp_bg) else 17

ws2.freeze_panes = 'A2'

# ── 保存 ──
out_dir = 'E:/Meta-analysis writing project/projects/paper-01/04-extract/'
os.makedirs(out_dir, exist_ok=True)
out = out_dir + '数据_6_数据提取表.xlsx'
wb.save(out)
print(f'保存成功：{out}')
print(f'共 {len(included)} 行，{len(COLS)} 列（A–{get_column_letter(len(COLS))}）')
