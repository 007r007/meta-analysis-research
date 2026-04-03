import openpyxl, sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

src = openpyxl.load_workbook('E:/Meta-analysis writing project/projects/paper-01/03-screen/数据_4_第三轮全文下载追踪.xlsx')
ws_src = src.active
rows = []
for r in range(2, ws_src.max_row + 1):
    rows.append([ws_src.cell(r, c).value for c in range(1, 14)])

wb = openpyxl.Workbook()

thin = Side(style='thin', color='FFB8CCE4')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def style(cell, bg=None, bold=False, align='left', fc='FF000000', wrap=False, size=10):
    if bg:
        cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fc, size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = border

def first_author(s):
    if not s:
        return ''
    p = s.split(';')[0].split(',')[0].strip().split()
    return p[-1] if p else s.split(';')[0][:15]

# ── Sheet 1: 第三轮全文筛选 ──
ws = wb.active
ws.title = '第三轮全文筛选'

cols = [
    ('#',           5),
    ('第二轮状态',   9),
    ('年份',         6),
    ('第一作者',    14),
    ('标题',        52),
    ('期刊',        26),
    ('DOI',         30),
    ('全文筛选结果', 13),
    ('排除原因',     13),
    ('备注',        35),
]

for ci, (h, w) in enumerate(cols, 1):
    c = ws.cell(1, ci, h)
    style(c, bg='FF2F5496', bold=True, align='center', fc='FFFFFFFF', size=10)
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 22
ws.freeze_panes = 'A2'

last_row = len(rows) + 1

for ri, row in enumerate(rows, 2):
    status2 = row[1]
    year    = row[4]
    title   = row[2]
    author  = row[3]
    journal = row[5]
    doi     = row[6]
    fa = first_author(str(author) if author else '')
    row_bg = 'FFFFEB9C' if status2 == '不确定' else 'FFFFFFFF'
    vals = [ri-1, status2, year, fa, title, journal, doi, None, None, None]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(ri, ci, v)
        align = 'center' if ci in (1,2,3,8,9) else 'left'
        style(c, bg=row_bg, align=align, wrap=(ci==5))
    ws.row_dimensions[ri].height = 30

dv1 = DataValidation(type='list', formula1='"纳入,排除,不确定"', allow_blank=True, showDropDown=False)
dv1.sqref = f'H2:H{last_row}'
ws.add_data_validation(dv1)

dv2 = DataValidation(type='list',
    formula1='"E1-人群不符,E2-干预不符,E3-结局不符,E4-研究设计,E7-全文不可及,E8-重复报告"',
    allow_blank=True, showDropDown=False)
dv2.sqref = f'I2:I{last_row}'
ws.add_data_validation(dv2)

ws.conditional_formatting.add(f'A2:J{last_row}', FormulaRule(
    formula=['$H2="纳入"'], fill=fill('FFC6EFCE'), font=Font(color='FF006100')))
ws.conditional_formatting.add(f'A2:J{last_row}', FormulaRule(
    formula=['$H2="排除"'], fill=fill('FFFFC7CE'), font=Font(color='FF9C0006')))
ws.conditional_formatting.add(f'A2:J{last_row}', FormulaRule(
    formula=['$H2="不确定"'], fill=fill('FFFFEB9C'), font=Font(color='FF7F6000')))

ws.auto_filter.ref = f'A1:J{last_row}'

# ── Sheet 2: 进度统计 ──
ws2 = wb.create_sheet('进度统计')
ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 28

c = ws2.cell(1, 1, '第三轮全文筛选 — 进度统计')
c.font = Font(bold=True, size=13, color='FF2F5496')
ws2.merge_cells('A1:C1')
ws2.row_dimensions[1].height = 28

sr = '第三轮全文筛选'
stats = [
    ('指标',         '数量',                                         '说明',                  True),
    ('输入总数',      f'=COUNTA({sr}!A2:A{last_row})',               '进入第三轮的条目',       False),
    ('已完成判决',    f'=COUNTA({sr}!H2:H{last_row})',               '已填写筛选结果',         False),
    ('待判决',        '=B4-B5',                                      '尚未填写',               False),
    ('完成进度',      '=IFERROR(B5/B4,0)',                           '',                      False),
    ('',             '',                                             '',                      False),
    ('纳入',          f'=COUNTIF({sr}!H$2:H${last_row},"纳入")',     '',                      False),
    ('排除',          f'=COUNTIF({sr}!H$2:H${last_row},"排除")',     '',                      False),
    ('不确定',        f'=COUNTIF({sr}!H$2:H${last_row},"不确定")',   '待进一步确认',           False),
    ('',             '',                                             '',                      False),
    ('排除原因明细',  '',                                             '',                      True),
    ('E1 人群不符',   f'=COUNTIF({sr}!I$2:I${last_row},"E1-人群不符")',  '年龄<60或疾病人群',  False),
    ('E2 干预不符',   f'=COUNTIF({sr}!I$2:I${last_row},"E2-干预不符")',  'WM非主要目标',       False),
    ('E3 结局不符',   f'=COUNTIF({sr}!I$2:I${last_row},"E3-结局不符")',  '无认知行为结局',     False),
    ('E4 研究设计',   f'=COUNTIF({sr}!I$2:I${last_row},"E4-研究设计")',  '综述/protocol/评论', False),
    ('E7 全文不可及', f'=COUNTIF({sr}!I$2:I${last_row},"E7-全文不可及")', '先补下载再标E7',    False),
    ('E8 重复报告',   f'=COUNTIF({sr}!I$2:I${last_row},"E8-重复报告")',  '同数据集多篇',       False),
]

for ri, (a, b, c_, hdr) in enumerate(stats, 2):
    ca = ws2.cell(ri, 1, a)
    cb = ws2.cell(ri, 2, b)
    cc = ws2.cell(ri, 3, c_)
    bg = 'FFD9E1F2' if hdr else None
    for cell in (ca, cb, cc):
        cell.font = Font(bold=hdr, size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        if bg:
            cell.fill = fill(bg)
    if b == '=IFERROR(B5/B4,0)':
        cb.number_format = '0.0%'
        cb.fill = fill('FFE2EFDA')
        cb.font = Font(bold=True, size=10, color='FF375623')
    ws2.row_dimensions[ri].height = 18

ws2.freeze_panes = 'A2'

# ── Sheet 3: PICOS速查 ──
ws3 = wb.create_sheet('PICOS速查')
ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 32

c = ws3.cell(1, 1, 'PICOS 第三轮全文筛选速查（v1.0，2026-04-02）')
c.font = Font(bold=True, size=12, color='FF2F5496')
ws3.merge_cells('A1:C1')
ws3.row_dimensions[1].height = 24

picos_data = [
    ('维度',     '纳入标准',  '排除 → 代码',  True,  18),
    ('P 人群',
     '健康老年人 ≥60岁，认知正常\n有老年+青年对比组时老年组须≥60岁',
     '<60岁（无老年组）→ E1\nMCI/痴呆/帕金森/卒中 → E1\n精神疾病/躯体疾病患者 → E1',
     False, 52),
    ('I 干预',
     'WM为主要目标的认知训练\n（n-back/complex span/Cogmed/WOME等）\n>=2次session；tDCS联合WM训练可纳入（备注"联合脑刺激"）',
     '记忆策略训练（记忆术）→ E2\n多域训练且未单独报告WM效果 → E2\n纯体力/药物/单独脑刺激/商业游戏 → E2',
     False, 52),
    ('C 对照',
     '主动对照/被动对照/单组前后测\n无对照但含调节因素分析的事后分析',
     '（无特定排除）',
     False, 36),
    ('O 结局',
     '直接训练效应、近迁移（未训练WM任务）\n远迁移（流体智力/执行功能/处理速度/注意/情景记忆）',
     '仅主观感受/生活质量/情绪 → E3\n仅神经影像（无行为结局）→ E3',
     False, 42),
    ('S 设计',
     'RCT、准实验、单组前后测、事后分析',
     '综述/Meta/系统综述 → E4\nProtocol/评论/社论/信件 → E4\n横断面研究 → E4',
     False, 42),
    ('', '', '', False, 10),
    ('排除代码', '含义',  '操作提示', True, 18),
    ('E1', '人群不符',    'Participants：年龄均值>=60，无MCI/疾病',   False, 18),
    ('E2', '干预不符',    'Intervention：WM是否为训练主要目标',        False, 18),
    ('E3', '结局不符',    'Results：必须有认知行为测量数据',           False, 18),
    ('E4', '研究设计不符','Protocol/综述/评论 → 排除',                False, 18),
    ('E7', '全文不可及',  '先补下载；确实无法获取再标E7',             False, 18),
    ('E8', '重复报告',    '同数据集多篇 → 保留主要结果文，其余标E8', False, 18),
    ('', '', '', False, 10),
    ('快速流程', '① 年龄>=60且认知正常？   否 → E1', '', False, 18),
    ('',         '② WM为主的认知训练？     否 → E2', '', False, 18),
    ('',         '③ 有认知行为结局？       否 → E3', '', False, 18),
    ('',         '④ 实证研究（非综述）？   否 → E4', '', False, 18),
    ('',         '⑤ 与已纳入文献重复？     是 → E8', '', False, 18),
    ('',         '⑥ 全通过 → 纳入',               '', False, 18),
]

for ri, (a, b, c_, hdr, h) in enumerate(picos_data, 2):
    ca = ws3.cell(ri, 1, a)
    cb = ws3.cell(ri, 2, b)
    cc = ws3.cell(ri, 3, c_)
    bg = 'FFD9E1F2' if hdr else None
    for cell in (ca, cb, cc):
        cell.font = Font(bold=hdr, size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        if bg:
            cell.fill = fill(bg)
    ws3.row_dimensions[ri].height = h

ws3.freeze_panes = 'A2'

out = 'E:/Meta-analysis writing project/projects/paper-01/03-screen/数据_5_第三轮全文筛选.xlsx'
wb.save(out)
print(f'生成成功：{out}')
print(f'数据行数：{len(rows)} 条')
