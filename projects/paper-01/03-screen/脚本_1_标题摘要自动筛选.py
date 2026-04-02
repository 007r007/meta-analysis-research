"""
脚本_2_标题摘要自动筛选.py
对去重后4168条文献做第一轮自动筛选，输出带颜色标注的Excel
"""

import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 路径配置 ─────────────────────────────────────────────────
BASE = Path(r"E:\Meta-analysis writing project\projects\paper-01")
RIS_FILE = BASE / "02-search" / "数据_1_四库合并去重后.ris"
OUT_DIR = BASE / "03-screen"
OUT_EXCEL = OUT_DIR / "数据_1_第一轮标题摘要筛选.xlsx"
OUT_STATS = OUT_DIR / "结果_1_筛选统计.json"

OUT_DIR.mkdir(exist_ok=True)

# ── 颜色定义 ─────────────────────────────────────────────────
COLORS = {
    "保留":  "C6EFCE",  # 绿色
    "E1":    "FFCCCC",  # 红色   — 非老年人群
    "E2":    "FFD7A8",  # 橙色   — 非WM训练
    "E3":    "FFEB9C",  # 黄色   — 无迁移/认知结局
    "E4":    "D9D9D9",  # 灰色   — 综述/非实验设计
    "E5":    "BDD7EE",  # 浅蓝   — 非英文
    "E6":    "E2EFDA",  # 浅绿灰 — 时间范围外
}

# ── 筛选规则词表 ─────────────────────────────────────────────

# E1：非老年人群
# 出现以下词，且同时没有老年词 → 排除
E1_EXCLUDE = [
    r'\bchildren\b', r'\bchild\b', r'\badolescent\b', r'\badolescents\b',
    r'\binfant\b', r'\binfants\b', r'\bpediatric\b', r'\bpaediatric\b',
    r'\byouth\b', r'\bundergraduate\b', r'\bstudent\b', r'\bstudents\b',
]
E1_ANIMAL = [
    r'\brat\b', r'\brats\b', r'\bmouse\b', r'\bmice\b',
    r'\banimal\b', r'\banimals\b', r'\brodent\b', r'\brodents\b',
    r'\bprimate\b', r'\bprimates\b',
]
E1_ELDERLY = [
    r'\bolder adult', r'\bolder people\b', r'\bolderly\b',
    r'\baging\b', r'\bageing\b', r'\baged\b', r'\bsenior\b',
    r'\bgeriatric\b', r'\bdementia\b', r'\balzheimer\b',
    r'\bmild cognitive impairment\b', r'\bmci\b',
]

# E2：非WM训练（标题+摘要中完全没有WM训练词）
E2_WM_TRAINING = [
    r'working memory training', r'working memory intervention',
    r'\bn-back\b', r'\bdual.?n.?back\b', r'complex span',
    r'\bcogmed\b', r'cognitive training', r'memory training',
    r'working memory program', r'working memory exercise',
    r'working memory practice', r'working memory rehabilitation',
]

# E3：无迁移/认知结局（标题+摘要中完全没有）
E3_TRANSFER = [
    r'\btransfer\b', r'\bgenerali[sz]ation\b',
    r'fluid intelligence', r'fluid cognition',
    r'executive function', r'executive control',
    r'cognitive outcome', r'cognitive performance', r'cognitive function',
    r'far transfer', r'near transfer',
    r'processing speed', r'\battention\b', r'\breasoning\b',
    r'\binhibition\b', r'episodic memory', r'\bintelligence\b',
    r'cognitive gain', r'cognitive benefit', r'cognitive improvement',
    r'cognitive effect', r'neuropsychological',
]

# E4：综述/非实验设计（标题字段精确匹配）
E4_TITLE_PATTERNS = [
    r'^a systematic review\b', r'\bsystematic review\b',
    r'\bmeta.?analysis\b', r'\bnarrative review\b',
    r'\bscoping review\b', r'\bliterature review\b',
    r'\beditorial\b', r'\bletter to\b', r'\bcommentary\b',
    r'\bcase report\b', r'\bcase study\b',
    r'\bqualitative\b',
]
# 摘要字段也检查，但用更严格的短语
E4_ABSTRACT_PATTERNS = [
    r'this systematic review', r'this meta.?analysis',
    r'this scoping review', r'this narrative review',
    r'we conducted a systematic review', r'we performed a meta.?analysis',
]

# E5：非英文（语言字段）
E5_LANG_EXCLUDE = ['chinese', 'german', 'french', 'spanish', 'japanese',
                   'korean', 'portuguese', 'italian', 'dutch', 'russian',
                   'polish', 'turkish', 'arabic', 'hebrew']

# ── 解析RIS ──────────────────────────────────────────────────

def parse_ris(filepath):
    records = []
    current = {}
    with open(filepath, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\n')
            if line.startswith('ER  -'):
                if current:
                    records.append(current)
                    current = {}
            elif len(line) >= 6 and line[2:6] == '  - ':
                tag = line[:2].strip()
                val = line[6:].strip()
                if tag in current:
                    if isinstance(current[tag], list):
                        current[tag].append(val)
                    else:
                        current[tag] = [current[tag], val]
                else:
                    current[tag] = val
    if current:
        records.append(current)
    return records


def get_field(rec, *tags):
    for tag in tags:
        v = rec.get(tag, '')
        if isinstance(v, list):
            v = ' '.join(v)
        if v:
            return v.strip()
    return ''


def get_year(rec):
    y = get_field(rec, 'PY', 'Y1', 'DA')
    m = re.search(r'\d{4}', y)
    return int(m.group()) if m else 0


def get_lang(rec):
    return get_field(rec, 'LA', 'LG').lower()


# ── 筛选逻辑 ─────────────────────────────────────────────────

def any_match(text, patterns):
    t = text.lower()
    return any(re.search(p, t) for p in patterns)


def screen(rec):
    title    = get_field(rec, 'TI', 'T1').lower()
    abstract = get_field(rec, 'AB', 'N2').lower()
    combined = title + ' ' + abstract
    lang     = get_lang(rec)
    year     = get_year(rec)

    # E5：非英文
    if lang and lang not in ('', 'english', 'en'):
        if any(l in lang for l in E5_LANG_EXCLUDE):
            return 'E5', '非英文'

    # E6：时间范围外
    if year and year < 2000:
        return 'E6', f'发表年份{year}，早于2000年'

    # E4：综述/非实验（标题优先，再看摘要）
    if any_match(title, E4_TITLE_PATTERNS):
        return 'E4', '标题显示为综述/非实验设计'
    if any_match(abstract, E4_ABSTRACT_PATTERNS):
        return 'E4', '摘要显示为综述/非实验设计'

    # E1：非老年人群
    # 动物词直接排除
    if any_match(combined, E1_ANIMAL):
        return 'E1', '动物研究'
    # 儿童/青少年词，且没有老年词
    if any_match(combined, E1_EXCLUDE) and not any_match(combined, E1_ELDERLY):
        return 'E1', '非老年人群（儿童/青少年/学生）'

    # E2：非WM训练
    if not any_match(combined, E2_WM_TRAINING):
        return 'E2', '无工作记忆训练相关词'

    # E3：无迁移/认知结局
    if not any_match(combined, E3_TRANSFER):
        return 'E3', '无迁移或认知结局测量词'

    return '保留', ''


# ── 主流程 ───────────────────────────────────────────────────

print('读取文献...')
records = parse_ris(RIS_FILE)
print(f'共 {len(records)} 条')

print('开始筛选...')
results = []
for rec in records:
    decision, reason = screen(rec)
    results.append({
        'rec': rec,
        'decision': decision,
        'reason': reason,
    })

# 统计
from collections import Counter
counts = Counter(r['decision'] for r in results)
print('\n筛选结果：')
for k, v in sorted(counts.items()):
    print(f'  {k}: {v}')
print(f'  总计: {len(results)}')

# ── 写Excel ──────────────────────────────────────────────────

print('\n生成Excel...')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '第一轮筛选'

# 表头
headers = ['#', '筛选结果', '排除原因', '标题', '作者', '年份',
           '期刊', '来源数据库', '摘要']
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 30

# 数据行
for i, item in enumerate(results, 1):
    rec = item['rec']
    decision = item['decision']
    reason = item['reason']

    title    = get_field(rec, 'TI', 'T1')
    authors  = get_field(rec, 'AU', 'A1')
    if isinstance(rec.get('AU') or rec.get('A1'), list):
        au_list = rec.get('AU') or rec.get('A1')
        authors = '; '.join(au_list[:3])
        if len(au_list) > 3:
            authors += ' et al.'
    year     = get_year(rec) or ''
    journal  = get_field(rec, 'JO', 'JF', 'T2')
    source   = get_field(rec, 'DB')
    abstract = get_field(rec, 'AB', 'N2')
    if len(abstract) > 500:
        abstract = abstract[:500] + '...'

    row_data = [i, decision, reason, title, authors, year,
                journal, source, abstract]

    fill = PatternFill('solid', fgColor=COLORS.get(decision, 'FFFFFF'))
    row_num = i + 1

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = fill
        cell.alignment = Alignment(vertical='top', wrap_text=(col in (4, 9)))
        cell.border = border
        if col in (2, 3):
            cell.alignment = Alignment(horizontal='center', vertical='top')

    ws.row_dimensions[row_num].height = 60

# 列宽
col_widths = [6, 10, 25, 60, 30, 8, 30, 12, 80]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# 冻结首行
ws.freeze_panes = 'A2'

# 自动筛选
ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'

# 图例sheet
ws2 = wb.create_sheet('图例说明')
legend = [
    ('颜色', '筛选结果', '排除原因说明'),
    ('绿色',  '保留',    '进入第二轮人工全文筛选'),
    ('红色',  'E1',      '非老年人群（儿童/青少年/动物）'),
    ('橙色',  'E2',      '非工作记忆训练'),
    ('黄色',  'E3',      '无迁移或认知结局测量'),
    ('灰色',  'E4',      '综述/元分析/非实验设计'),
    ('浅蓝',  'E5',      '非英文文献'),
    ('浅绿灰','E6',      '发表年份早于2000年'),
]
legend_colors = ['4472C4', 'C6EFCE', 'FFCCCC', 'FFD7A8',
                 'FFEB9C', 'D9D9D9', 'BDD7EE', 'E2EFDA']
for r, (row, color) in enumerate(zip(legend, legend_colors), 1):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill = PatternFill('solid', fgColor=color)
        cell.font = Font(bold=(r == 1), color='FFFFFF' if r == 1 else '000000')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    ws2.row_dimensions[r].height = 22
for col, width in enumerate([12, 12, 45], 1):
    ws2.column_dimensions[get_column_letter(col)].width = width

wb.save(OUT_EXCEL)
print(f'Excel已保存：{OUT_EXCEL}')

# ── 保存统计JSON ─────────────────────────────────────────────
stats = {
    'total_input': len(records),
    'results': dict(counts),
    'retained': counts.get('保留', 0),
    'excluded': len(records) - counts.get('保留', 0),
}
with open(OUT_STATS, 'w', encoding='utf-8') as f:
    json.dump(stats, f, ensure_ascii=False, indent=2)
print(f'统计已保存：{OUT_STATS}')
print('\n完成。')
