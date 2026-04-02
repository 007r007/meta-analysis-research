"""
脚本_2_第二轮人工摘要筛选表生成.py
从第一轮保留的445条文献生成人工摘要筛选Excel
供研究者逐篇阅读摘要，标注纳入/排除/不确定

输出：数据_3_第二轮人工摘要筛选.xlsx
"""

import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 路径配置 ─────────────────────────────────────────────────
BASE    = Path(r"E:\Meta-analysis writing project\projects\paper-01")
SCREEN_DIR = BASE / "03-screen"
EXCEL_V2   = SCREEN_DIR / "数据_2_第一轮标题摘要筛选_v2.xlsx"
RIS_FILE   = BASE / "02-search" / "数据_1_四库合并去重后.ris"
OUT_EXCEL  = SCREEN_DIR / "数据_3_第二轮人工摘要筛选.xlsx"

# ── 颜色 ─────────────────────────────────────────────────────
COL_HEADER  = "4472C4"
COL_INCLUDE = "C6EFCE"   # 绿
COL_EXCLUDE = "FFCCCC"   # 红
COL_UNSURE  = "FFEB9C"   # 黄
COL_BLANK   = "FFFFFF"

thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 解析RIS ──────────────────────────────────────────────────

def parse_ris(filepath):
    records = []
    current = {}
    with open(filepath, encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            if line.startswith("ER  -"):
                if current:
                    records.append(current)
                    current = {}
            elif len(line) >= 6 and line[2:6] == "  - ":
                tag = line[:2].strip()
                val = line[6:].strip()
                if tag in current:
                    if isinstance(current[tag], list):
                        current[tag].append(val)
                    else:
                        current[tag] = [current[tag], val]
                else:
                    current[tag] = val
    if current:
        records.append(current)
    return records


def get_field(rec, *tags):
    for tag in tags:
        v = rec.get(tag, "")
        if isinstance(v, list):
            v = " ".join(v)
        if v:
            return v.strip()
    return ""


def get_year(rec):
    y = get_field(rec, "PY", "Y1", "DA")
    m = re.search(r"\d{4}", y)
    return int(m.group()) if m else 0


# ── 读取第一轮保留列表 ────────────────────────────────────────

print("读取第一轮筛选结果...")
wb_v2 = openpyxl.load_workbook(EXCEL_V2)
ws_v2 = wb_v2["第一轮筛选"]

# 收集保留行的序号（第1列=#，第2列=筛选结果）
retained_ids = set()
for row in ws_v2.iter_rows(min_row=2, values_only=True):
    if row[1] == "保留":
        retained_ids.add(row[0])  # row[0] = #（1-based index in original RIS）

print(f"第一轮保留：{len(retained_ids)} 条")

# ── 读取RIS，提取保留文献 ─────────────────────────────────────

print("读取RIS文件...")
records = parse_ris(RIS_FILE)
print(f"RIS总条数：{len(records)}")

retained_records = []
for i, rec in enumerate(records, 1):
    if i in retained_ids:
        retained_records.append((i, rec))

print(f"匹配到保留文献：{len(retained_records)} 条")

# ── 生成Excel ─────────────────────────────────────────────────

print("生成Excel...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "第二轮摘要筛选"

# 表头
headers = [
    "#",           # A
    "筛选决定",    # B  ← 下拉菜单
    "排除原因",    # C  ← 下拉菜单（仅排除时填）
    "备注",        # D
    "标题",        # E
    "作者",        # F
    "年份",        # G
    "期刊",        # H
    "摘要（完整）",# I
]

header_fill = PatternFill("solid", fgColor=COL_HEADER)
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

ws.row_dimensions[1].height = 32

# 数据验证：筛选决定下拉
dv_decision = DataValidation(
    type="list",
    formula1='"纳入,排除,不确定"',
    allow_blank=True,
    showDropDown=False,
)
dv_decision.sqref = f"B2:B{len(retained_records)+1}"
ws.add_data_validation(dv_decision)

# 数据验证：排除原因下拉
exclude_reasons = (
    '"E1-非老年人群,'
    'E2-非WM训练,'
    'E3-无认知结局,'
    'E4-综述/非实验,'
    'E5-非英文,'
    'E6-年份范围外,'
    'E7-无对照组,'
    'E8-重复报告"'
)
dv_reason = DataValidation(
    type="list",
    formula1=exclude_reasons,
    allow_blank=True,
    showDropDown=False,
)
dv_reason.sqref = f"C2:C{len(retained_records)+1}"
ws.add_data_validation(dv_reason)

# 数据行
for seq, (orig_id, rec) in enumerate(retained_records, 1):
    title    = get_field(rec, "TI", "T1")
    au_raw   = rec.get("AU") or rec.get("A1") or ""
    if isinstance(au_raw, list):
        authors = "; ".join(au_raw[:3])
        if len(au_raw) > 3:
            authors += " et al."
    else:
        authors = au_raw
    year     = get_year(rec) or ""
    journal  = get_field(rec, "JO", "JF", "T2")
    abstract = get_field(rec, "AB", "N2")

    row_data = [orig_id, "", "", "", title, authors, year, journal, abstract]
    row_num  = seq + 1

    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=COL_BLANK)
        cell.border = border
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=(col in (5, 9)),  # 标题和摘要自动换行
        )
        if col in (1, 2, 3, 7):
            cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.row_dimensions[row_num].height = 80

# 列宽
col_widths = [6, 10, 18, 20, 60, 28, 7, 28, 90]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# 冻结首行+首列
ws.freeze_panes = "E2"

# 自动筛选
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ── 说明sheet ────────────────────────────────────────────────

ws2 = wb.create_sheet("筛选说明")
instructions = [
    ("操作说明", ""),
    ("1. 逐行阅读标题+摘要", ""),
    ("2. 在B列「筛选决定」选择：纳入 / 排除 / 不确定", ""),
    ("3. 排除时在C列选择排除原因", ""),
    ("4. 不确定的文献进入第三轮全文筛选", ""),
    ("", ""),
    ("排除代码说明", ""),
    ("E1", "非老年人群（<60岁，无老年组）"),
    ("E2", "非WM训练（记忆策略、多域训练且未单独报告WM效果）"),
    ("E3", "无认知结局（仅主观感受/生活质量/神经影像无行为数据）"),
    ("E4", "综述/元分析/非实验设计/protocol"),
    ("E5", "非英文"),
    ("E6", "发表年份范围外（<2000）"),
    ("E7", "无对照组（且未报告调节因素分析）"),
    ("E8", "重复报告（同一研究多篇，保留主要结果文章）"),
    ("", ""),
    ("PICOS标准参考", "见文档_1_第二轮筛选PICOS标准.md"),
    ("目标", "445条 → 预计保留100-150条进入全文筛选"),
]

for r, (key, val) in enumerate(instructions, 1):
    c1 = ws2.cell(row=r, column=1, value=key)
    c2 = ws2.cell(row=r, column=2, value=val)
    if key in ("操作说明", "排除代码说明", "PICOS标准参考", "目标"):
        c1.font = Font(bold=True)
    c1.alignment = Alignment(vertical="top")
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    ws2.row_dimensions[r].height = 18

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 55

wb.save(OUT_EXCEL)
print(f"\nExcel已保存：{OUT_EXCEL}")
print(f"共 {len(retained_records)} 条待人工筛选")
print("完成。")
