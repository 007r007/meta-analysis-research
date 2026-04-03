"""
脚本_3_全文下载.py
第三轮全文筛选 - 自动下载71篇纳入/不确定文献的PDF

流程：
1. 从数据_3_第二轮人工摘要筛选.xlsx 提取纳入/不确定的71条
2. 从02-search/数据_1_四库合并去重后.ris 匹配 DOI
3. 通过 CORE API 尝试下载开放获取版本
4. CORE 失败则尝试 Sci-Hub 镜像
5. 生成数据_4_第三轮全文下载追踪.xlsx 记录状态
"""

import os
import re
import sys
import io
import time
import json
import unicodedata
import requests
import openpyxl

# Fix Windows GBK console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from datetime import datetime

# ���─ 路径配置 ──────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
SCREEN_DIR = BASE_DIR
RIS_FILE   = BASE_DIR.parent / "02-search" / "数据_1_四库合并去重后.ris"
EXCEL_IN   = BASE_DIR / "数据_3_第二轮人工摘要筛选.xlsx"
PDF_DIR    = BASE_DIR / "全文PDF"
EXCEL_OUT  = BASE_DIR / "数据_4_第三轮全文下载追踪.xlsx"

PDF_DIR.mkdir(exist_ok=True)

# ── Sci-Hub 镜像列表（按优先级） ──────────────────────────────────────────
SCIHUB_MIRRORS = [
    "https://sci-hub.st",
    "https://sci-hub.ru",
    "https://sci-hub.box",
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/122.0.0.0 Safari/537.36"
}


# ── 工具函数 ──────────────────────────────────────────────────────────────

def normalize_title(t: str) -> str:
    """小写+去标点，用于模糊匹配"""
    if not t:
        return ""
    t = unicodedata.normalize("NFKD", t)
    t = re.sub(r"[^\w\s]", " ", t.lower())
    return re.sub(r"\s+", " ", t).strip()


def load_ris_doi_map(ris_path: Path) -> dict:
    """解析 RIS 文件，返回 title_normalized -> doi 字典"""
    doi_map = {}
    current_title = None
    current_doi = None

    with open(ris_path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.rstrip("\n")
            if line.startswith("TI  - "):
                current_title = line[6:].strip()
            elif line.startswith("DO  - "):
                raw = line[6:].strip()
                # 去掉末尾的 " [doi]"
                current_doi = re.sub(r"\s*\[doi\]$", "", raw, flags=re.I).strip()
            elif line.startswith("ER  -"):
                if current_title and current_doi:
                    key = normalize_title(current_title)
                    doi_map[key] = current_doi
                current_title = None
                current_doi = None

    print(f"[RIS] 解析完成，共 {len(doi_map)} 条有 DOI 的记录")
    return doi_map


def find_doi(title: str, doi_map: dict) -> str | None:
    """先精确匹配，再宽松匹配（前100字符）"""
    key = normalize_title(title)
    if key in doi_map:
        return doi_map[key]
    # 前100字符宽松匹配
    prefix = key[:100]
    for k, v in doi_map.items():
        if k[:100] == prefix:
            return v
    return None


def safe_filename(title: str, year, max_len=80) -> str:
    """生成安全的文件名"""
    name = re.sub(r'[\\/:*?"<>|]', '_', title)
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > max_len:
        name = name[:max_len].strip()
    return f"{year}_{name}.pdf"


# ── CORE 下载 ─────────────────────────────────────────────────────────────

def try_core(doi: str, save_path: Path) -> bool:
    """通过 CORE API 查找并下载 PDF"""
    try:
        api_url = f"https://api.core.ac.uk/v3/works/doi:{doi}"
        r = requests.get(api_url, headers=HEADERS, timeout=20)
        if r.status_code != 200:
            return False
        data = r.json()
        dl_url = data.get("downloadUrl") or data.get("fullTextLink")
        if not dl_url:
            # 尝试从 links 列表里找
            for lk in data.get("links", []):
                if lk.get("type") in ("download", "pdf") and lk.get("url"):
                    dl_url = lk["url"]
                    break
        if not dl_url:
            return False

        pdf_r = requests.get(dl_url, headers=HEADERS, timeout=40, stream=True)
        if pdf_r.status_code == 200 and "pdf" in pdf_r.headers.get("Content-Type", ""):
            with open(save_path, "wb") as f:
                for chunk in pdf_r.iter_content(8192):
                    f.write(chunk)
            return save_path.stat().st_size > 10_000  # 至少10KB才算成功
    except Exception as e:
        print(f"    [CORE] 异常: {e}")
    return False


# ── Sci-Hub 下载 ──────────────────────────────────────────────────────────

def try_scihub(doi: str, save_path: Path) -> tuple[bool, str]:
    """依次尝试各 Sci-Hub 镜像下载 PDF，返回 (成功, 来源)"""
    from bs4 import BeautifulSoup

    for mirror in SCIHUB_MIRRORS:
        try:
            url = f"{mirror}/{doi}"
            r = requests.get(url, headers=HEADERS, timeout=30)
            if r.status_code != 200:
                continue

            soup = BeautifulSoup(r.text, "html.parser")

            # 找 PDF 链接：<iframe id="pdf"> 或 embed 或 a 标签
            pdf_url = None
            iframe = soup.find("iframe", id="pdf")
            if iframe and iframe.get("src"):
                pdf_url = iframe["src"]
            if not pdf_url:
                embed = soup.find("embed", type="application/pdf")
                if embed and embed.get("src"):
                    pdf_url = embed["src"]
            if not pdf_url:
                for a in soup.find_all("a", href=True):
                    if ".pdf" in a["href"].lower():
                        pdf_url = a["href"]
                        break

            if not pdf_url:
                continue

            # 补全协议
            if pdf_url.startswith("//"):
                pdf_url = "https:" + pdf_url
            elif pdf_url.startswith("/"):
                pdf_url = mirror + pdf_url

            pdf_r = requests.get(pdf_url, headers=HEADERS, timeout=60, stream=True)
            if pdf_r.status_code == 200:
                content_type = pdf_r.headers.get("Content-Type", "")
                if "pdf" in content_type or "octet-stream" in content_type:
                    with open(save_path, "wb") as f:
                        for chunk in pdf_r.iter_content(8192):
                            f.write(chunk)
                    if save_path.stat().st_size > 10_000:
                        return True, mirror

        except Exception as e:
            print(f"    [Sci-Hub {mirror}] 异常: {e}")
            continue

    return False, ""


# ── 主流程 ────────────────────────────────────────────────────────────────

def main():
    # 1. 读取纳入/不确定的71条
    wb_in = openpyxl.load_workbook(EXCEL_IN)
    ws_in = wb_in.active

    papers = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        status = str(row[1]) if row[1] else ""
        if "纳入" in status or "不确定" in status:
            papers.append({
                "id":      row[0],
                "status":  status,
                "title":   str(row[4]) if row[4] else "",
                "authors": str(row[5]) if row[5] else "",
                "year":    row[6],
                "journal": str(row[7]) if row[7] else "",
            })

    print(f"\n共 {len(papers)} 篇需要下载（纳入+不确定）\n")

    # 2. 加载 DOI 字典
    doi_map = load_ris_doi_map(RIS_FILE)

    # 3. 逐篇处理
    results = []
    for i, p in enumerate(papers, 1):
        title   = p["title"]
        year    = p["year"] or "unknown"
        doi     = find_doi(title, doi_map)
        fname   = safe_filename(title, year)
        fpath   = PDF_DIR / fname

        print(f"[{i:02d}/{len(papers)}] {title[:60]}...")
        print(f"    DOI: {doi or '未找到'}")

        if fpath.exists() and fpath.stat().st_size > 10_000:
            print("    → 已存在，跳过")
            results.append({**p, "doi": doi, "filename": fname,
                             "download_status": "已存在", "source": "缓存"})
            continue

        if not doi:
            print("    → 无 DOI，需手动下载")
            results.append({**p, "doi": "", "filename": fname,
                             "download_status": "无DOI-手动", "source": ""})
            continue

        # CORE 优先
        if try_core(doi, fpath):
            print("    → ✅ CORE 下载成功")
            results.append({**p, "doi": doi, "filename": fname,
                             "download_status": "已下载", "source": "CORE"})
            time.sleep(1)
            continue

        # Sci-Hub
        ok, mirror = try_scihub(doi, fpath)
        if ok:
            print(f"    → ✅ Sci-Hub 下载成功 ({mirror})")
            results.append({**p, "doi": doi, "filename": fname,
                             "download_status": "已下载", "source": f"Sci-Hub({mirror})"})
        else:
            print("    → ❌ 未能自动下载，需手动")
            results.append({**p, "doi": doi, "filename": fname,
                             "download_status": "手动下载", "source": ""})

        time.sleep(2)  # 礼貌延迟

    # 4. 生成追踪 Excel
    wb_out = openpyxl.Workbook()
    ws     = wb_out.active
    ws.title = "全文下载追踪"

    headers = ["#", "第二轮状态", "标题", "作者", "年份", "期刊",
               "DOI", "文件名", "下载状态", "来源", "全文筛选结果", "排除原因", "备注"]
    ws.append(headers)

    # 颜色定义
    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red    = PatternFill("solid", fgColor="FFC7CE")
    grey   = PatternFill("solid", fgColor="D9D9D9")
    header_fill = PatternFill("solid", fgColor="2E75B6")

    # 表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # 下载状态→颜色
    status_fill = {
        "已下载":    green,
        "已存在":    green,
        "手动下载":  yellow,
        "无DOI-手动": yellow,
    }

    for r in results:
        row_data = [
            r["id"], r["status"], r["title"], r["authors"],
            r["year"], r["journal"], r["doi"], r["filename"],
            r["download_status"], r["source"], "", "", ""
        ]
        ws.append(row_data)
        last_row = ws.max_row
        fill = status_fill.get(r["download_status"], grey)
        for col in range(1, len(headers) + 1):
            ws.cell(last_row, col).fill = fill

    # 列宽
    col_widths = [5, 10, 50, 30, 6, 25, 35, 45, 14, 20, 14, 16, 20]
    for col_i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 图例
    ws2 = wb_out.create_sheet("说明")
    ws2.append(["颜色", "含义"])
    green_row  = ws2.max_row + 1
    ws2.append(["绿色", "PDF 已自动下载"])
    ws2.cell(green_row, 1).fill = green
    yellow_row = ws2.max_row + 1
    ws2.append(["黄色", "需手动下载（无DOI或自动失败）"])
    ws2.cell(yellow_row, 1).fill = yellow

    # 统计
    n_downloaded = sum(1 for r in results if r["download_status"] in ("已下载", "已存在"))
    n_manual     = len(results) - n_downloaded
    ws2.append([])
    ws2.append(["统计", ""])
    ws2.append(["自动下载成功", n_downloaded])
    ws2.append(["需手动下载",   n_manual])
    ws2.append(["总计",        len(results)])
    ws2.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")])

    wb_out.save(EXCEL_OUT)

    print(f"\n{'='*60}")
    print(f"完成！自动下载: {n_downloaded}/{len(results)}")
    print(f"需手动下载: {n_manual} 篇")
    print(f"Excel 已保存: {EXCEL_OUT}")
    print(f"PDF 目录: {PDF_DIR}")


if __name__ == "__main__":
    main()
